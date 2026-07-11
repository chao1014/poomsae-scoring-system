import tkinter as tk
from tkinter import ttk, font, filedialog, messagebox
import socket
import threading
import time
import pandas as pd
import traceback
import uuid
import random
import webbrowser
import os
import sqlite3
import urllib.parse
import urllib.request
from datetime import datetime

import config
import database
from projection import ProjectionWindow
import gui_dialogs

try:
    from screeninfo import get_monitors
except ImportError:
    get_monitors = None

def get_web_server():
    import web_server
    return web_server



system_settings = config.system_settings
current_state = config.current_state
colors = config.colors

try:
    from watchdog.observers import Observer
    from watchdog.events import FileSystemEventHandler
except ImportError:
    Observer = None
    FileSystemEventHandler = object

class MatchFileHandler(FileSystemEventHandler):
    def __init__(self, app_ui):
        self.app = app_ui
    def _process(self, event):
        if event.is_directory: return
        filename = event.src_path
        if not filename.endswith(('.xlsx', '.xls')): return
        self.app.root.after(0, lambda: self.app.handle_file_change(event.event_type, filename))
    def on_created(self, event): self._process(event)
    def on_modified(self, event): self._process(event)
    def on_deleted(self, event): self._process(event)

class PoomsaeReplicaGUI:
    instance = None
    
    def invalidate_leaderboard_cache(self, uid=None):
        """使排行榜快取失效 (dirty-flag 機制)。
        uid: 指定使特定選手的快取失效。若為 None，則清除目前組別所有同組選手的快取。"""
        if uid is not None:
            if uid in self.imported_matches:
                mdata = self.imported_matches[uid]
                for k in ["final_score", "final_score_0", "final_score_1",
                           "presentation_score", "presentation_score_0", "presentation_score_1",
                           "raw_total_score", "raw_total_score_0", "raw_total_score_1"]:
                    mdata.pop(k, None)
            return
        if not self.current_match_data:
            return
        current_cat = self.current_match_data.get("Category", "")
        current_div = self.current_match_data.get("Division", "")
        current_phase = self.current_match_data.get("Phase", "")
        for uid, mdata in self.imported_matches.items():
            if (mdata.get("Category") == current_cat and
                mdata.get("Division") == current_div and
                mdata.get("Phase") == current_phase):
                for k in ["final_score", "final_score_0", "final_score_1",
                           "presentation_score", "presentation_score_0", "presentation_score_1",
                           "raw_total_score", "raw_total_score_0", "raw_total_score_1"]:
                    mdata.pop(k, None)

    def query_leaderboard_data(self, cat=None, div=None, phase=None, round_num=None):
        """從資料庫與快取中載入並計算目前組別所有完賽選手的排行榜資料，支援 WT 同分打破與並列名次"""
        if cat is not None and div is not None and phase is not None:
            current_cat = cat
            current_div = div
            current_phase = phase
            rounds = round_num if round_num is not None else 2
        else:
            if not self.current_match_data: return []
            current_cat = self.current_match_data.get("Category", "")
            current_div = self.current_match_data.get("Division", "")
            current_phase = self.current_match_data.get("Phase", "")
            try:
                rounds = int(self.current_match_data.get("Round", 2))
            except:
                rounds = 2

        group_players = []
        for uid, mdata in self.imported_matches.items():
            if (mdata.get("Category") == current_cat and 
                mdata.get("Division") == current_div and 
                mdata.get("Phase") == current_phase):
                
                if mdata.get("Status") == "End":
                    group_players.append((uid, mdata))
                elif uid == self.current_match_uuid:
                    # 當前選手：必須確認已經完成該賽事要求的總輪數，才可列入大螢幕排行榜
                    if rounds == 1:
                        score_1r_text = self.left_labels[3][0].cget("text") if hasattr(self, 'left_labels') else ""
                        has_1r_score = (score_1r_text not in ["", "-"])
                        if has_1r_score:
                            group_players.append((uid, mdata))
                    else:
                        score_2r_text = self.left_labels[3][1].cget("text") if hasattr(self, 'left_labels') else ""
                        has_2r_score = (score_2r_text not in ["", "-"])
                        if self.current_stage == 2 and has_2r_score:
                            group_players.append((uid, mdata))
                
        leaderboard = []
        for uid, mdata in group_players:
            score_val = self.get_final_score(uid, mdata)
            if score_val >= 0:
                leaderboard.append({
                    "name": mdata.get("C_Name", ""),
                    "team": mdata.get("C_Team", ""),
                    "score": score_val,
                    "presentation_score": mdata.get("presentation_score", 0.0),
                    "raw_total_score": mdata.get("raw_total_score", 0.0)
                })
                
        # 依據 WT 規則排序：最終得分降序 -> 表現力去尾平均降序 -> 原始分數加總總分降序
        leaderboard.sort(key=lambda x: (x["score"], x["presentation_score"], x["raw_total_score"]), reverse=True)
        
        # 計算名次，只有當三個關鍵分數皆相同時才算並列
        # 使用 round(x, 4) 避免浮點數精度差異（10.00000000001 != 10.0）導致應並列卻不並列
        def scores_eq(a, b):
            return round(a, 4) == round(b, 4)
        
        for idx, item in enumerate(leaderboard):
            if idx > 0 and (
                scores_eq(item["score"], leaderboard[idx - 1]["score"]) and
                scores_eq(item["presentation_score"], leaderboard[idx - 1]["presentation_score"]) and
                scores_eq(item["raw_total_score"], leaderboard[idx - 1]["raw_total_score"])
            ):
                item["rank"] = leaderboard[idx - 1]["rank"]
            else:
                item["rank"] = idx + 1
        return leaderboard

    def __init__(self, root):
        PoomsaeReplicaGUI.instance = self
        self.root = root
        self.qr_popup_window = None
        self.root.title("品勢電子計分系統 (Poomsae Scoring System)")
        
        # 綁定視窗關閉事件
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        
        # 檢查並清理過大的日誌
        self.check_log_rotation()
        
        # === 修改: 程式啟動時載入設定 ===
        config.load_settings()
        
        try: self.root.state('zoomed')
        except: self.root.attributes('-fullscreen', True)
        self.root.configure(bg="#f0f0f0")

        # === 設定 Treeview 樣式 (加大行高與字型) ===
        style = ttk.Style()
        style.configure("Treeview", rowheight=25, font=("Microsoft JhengHei", 9))
        style.configure("Treeview.Heading", font=("Microsoft JhengHei", 9, "bold"))

        self.colors = {
            "header_bg": "#ffffff", "bar_blue": "#0099cc", "bar_green": "#669900", "bar_red": "#cc0000",
            "table_purple": "#e6ccff", "table_header": "#d9b3ff", "cyan_header": "#008fb3", "judge_blue": "#0099cc",
            "judge_done": "#CC9933", 
            "lime_green": "#99cc00", "lime_dark": "#668000", "lime_val_bg": "#f2ffcc",
            "btn_yellow": "#ffff99", "btn_orange": "#ffcc00", "btn_pink": "#ffcccc",
            "info_bg": "#e6f7ff", "timer_bg": "#000000", "timer_fg": "#ffff00"
        }
        self.font_bold = ("Microsoft JhengHei", 9, "bold")
        self.font_info = ("Microsoft JhengHei", 9)
        self.font_timer = ("Microsoft JhengHei", 18, "bold")

        self.timer_running = False
        self.timer_seconds = system_settings["countdown_sec"]
        self.result_vars = {} 
        self.imported_matches = {} 
        self.score_cells = [] 
        
        self.current_match_data = None 
        self.current_match_uuid = None
        self.is_locked = False         
        self.current_stage = 1         
        self.proj_window = None 
        self.current_proj_status = "Waiting..."
        self.last_proj_score_slide = 0
        self.last_proj_slide_finished = False
        self.temp_scores_to_save = {}
        
        self.score_1r_avg = 0.0; self.score_2r_avg = 0.0
        self.score_1r_raw = 0.0; self.score_2r_raw = 0.0
        
        # === 新增: 暫存變數 (用於扣分調整時重算) ===
        self.temp_avg_acc = 0.0
        self.temp_avg_pres = 0.0
        self.temp_raw_sum = 0.0
        
        self.left_labels = {}; self.left_merged_labels = {}; self.right_labels = {}; self.right_merged_labels = {}; self.center_stats_labels = {}
        
        self.observer = None
        self.watch_directory = None
        self._writing_files = set()  # 回寫 Excel 時，暫時標記以避免 Watchdog 重觸發
        
        self.poomsae_list = self._load_poomsae_list()

        self.build_header()
        main_frame = tk.Frame(root, bg="white")
        main_frame.pack(fill="both", expand=True, padx=5, pady=5)
        self.build_left_panel(main_frame)
        self.build_center_panel(main_frame)
        self.build_right_panel(main_frame)
        self.build_bottom_panel()
        
        database.set_tournament_db(system_settings["tournament_name"])
        self.refresh_judge_slots()
        
        # 自動載入並監控上一次所設定的 Excel 資料夾
        last_dir = system_settings.get("last_excel_directory", "")
        if last_dir and os.path.isdir(last_dir):
            self.watch_directory = last_dir
            self.scan_folder(last_dir)
            self.update_tree_columns()
            if Observer:
                event_handler = MatchFileHandler(self)
                self.observer = Observer()
                self.observer.schedule(event_handler, last_dir, recursive=True)
                self.observer.start()
            self.update_excel_dir_label(last_dir)
            print(f"已自動載入並監控上一次的 Excel 資料夾：{last_dir}")
        else:
            self.update_excel_dir_label("")

    def check_log_rotation(self):
        """檢查 error.log 大小，若大於 5MB 則進行輪轉備份"""
        try:
            log_file = "error.log"
            if os.path.exists(log_file) and os.path.getsize(log_file) > 5 * 1024 * 1024:
                backup_file = "error_old.log"
                if os.path.exists(backup_file):
                    try: os.remove(backup_file)
                    except: pass
                os.rename(log_file, backup_file)
        except:
            pass

    def on_closing(self):
        from tkinter import messagebox
        import os
        
        # 檢查是否有比賽正在準備或進行中
        if hasattr(self, 'btn_ready') and self.btn_ready.cget("text") == "取消":
            messagebox.showwarning("無法關閉", "目前有比賽正在準備或進行中，請先取消或結束比賽後再關閉系統！", parent=self.root)
            return
            
        if messagebox.askokcancel("關閉系統", "確定要關閉品勢計分系統嗎？", parent=self.root):
            # 停止資料夾監控
            if hasattr(self, 'observer') and self.observer:
                try:
                    self.observer.stop()
                    self.observer.join()
                except:
                    pass
            # 停止 Ngrok 安全通道子進程
            if hasattr(self, 'stop_tunnel_callback') and self.stop_tunnel_callback:
                try:
                    self.stop_tunnel_callback()
                except:
                    pass
            self.root.destroy()
            os._exit(0)

    def _load_poomsae_list(self):
        default_poomsae = [
            "太極一章 Taegeuk1", "太極二章 Taegeuk2", "太極三章 Taegeuk3", "太極四章 Taegeuk4",
            "太極五章 Taegeuk5", "太極六章 Taegeuk6", "太極七章 Taegeuk7", "太極八章 Taegeuk8",
            "高麗型 Koryo", "金剛型 Keumgang", "太白型 Taebaek", "平原型 Pyongwon",
            "十進型 Sipjin", "地跆 Jitae", "天拳 Chonkwon", "漢水 Hansu", "一如 Ilyeo"
        ]
        
        # 預先填入預設型場作為 fallback 時使用
        self.poomsae_data_list = []
        for idx, name in enumerate(default_poomsae):
            self.poomsae_data_list.append({
                "no": idx + 1,
                "display": name
            })
        
        path = system_settings.get("poomsae_excel_path", "")
        if not path or not os.path.exists(path):
            print(f"型場 Excel 檔案路徑不存在或未設定: {path}，使用系統預設型場。")
            return [""] + default_poomsae
            
        try:
            df = pd.read_excel(path)
            
            # 防禦性讀取：優先找 '編號'、'型場' 與 '型場英文' 欄位
            col_no = None
            col_zh = None
            col_en = None
            
            for col in df.columns:
                col_str = str(col)
                if '編號' in col_str or 'id' in col_str.lower() or 'no' in col_str.lower():
                    col_no = col
                elif '型場' in col_str and '英文' not in col_str:
                    col_zh = col
                elif '型場英文' in col_str or ('英文' in col_str and '型場' in col_str):
                    col_en = col
            
            # 若欄位名稱沒有對上，則依據欄位 index 來決定：
            if col_no is None and len(df.columns) > 0:
                col_no = df.columns[0]
            if col_zh is None and len(df.columns) > 1:
                col_zh = df.columns[1]
            if col_en is None and len(df.columns) > 2:
                col_en = df.columns[2]
                
            if col_zh is None:
                raise ValueError("無法解析 Excel 檔案中的型場欄位。")
                
            poomsaes = []
            excel_data_list = []
            for idx, row in df.iterrows():
                raw_no = row[col_no] if col_no is not None else (idx + 1)
                try:
                    no_val = int(float(raw_no))
                except:
                    no_val = idx + 1
                    
                zh_val = str(row[col_zh]).strip() if pd.notna(row[col_zh]) else ""
                en_val = str(row[col_en]).strip() if (col_en is not None and pd.notna(row[col_en])) else ""
                
                # 排除空值或 nan 的無效列
                if not zh_val or zh_val.lower() == 'nan':
                    continue
                
                display_name = f"{zh_val} {en_val}" if (en_val and en_val.lower() != 'nan') else zh_val
                poomsaes.append(display_name)
                
                excel_data_list.append({
                    "no": no_val,
                    "display": display_name
                })
                
            if not poomsaes:
                raise ValueError("Excel 檔案中無有效型場資料。")
                
            self.poomsae_data_list = excel_data_list
            print(f"已成功從 Excel 載入 {len(poomsaes)} 個型場。")
            return [""] + poomsaes
            
        except Exception as e:
            print(f"讀取型場 Excel 發生錯誤: {e}，回退使用系統預設型場。")
            # 出錯時 self.poomsae_data_list 依然保留最初填入的 17 個預設型場
            return [""] + default_poomsae

    def update_poomsae_list(self):
        self.poomsae_list = self._load_poomsae_list()
        
        # 更新下拉選單的選項
        if hasattr(self, 'combo_poomsae_1'):
            curr1 = self.combo_poomsae_1.get()
            self.combo_poomsae_1['values'] = self.poomsae_list
            if curr1 in self.poomsae_list:
                self.combo_poomsae_1.set(curr1)
            else:
                self.combo_poomsae_1.current(0)
                
        if hasattr(self, 'combo_poomsae_2'):
            curr2 = self.combo_poomsae_2.get()
            self.combo_poomsae_2['values'] = self.poomsae_list
            if curr2 in self.poomsae_list:
                self.combo_poomsae_2.set(curr2)
            else:
                self.combo_poomsae_2.current(0)

    def setup_hover(self, button, hover_bg, normal_bg, hover_fg=None, normal_fg=None):
        def on_enter(e):
            button.config(bg=hover_bg)
            if hover_fg:
                button.config(fg=hover_fg)
        def on_leave(e):
            button.config(bg=normal_bg)
            if normal_fg:
                button.config(fg=normal_fg)
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    def setup_dynamic_hover(self, button):
        def adjust_color(hex_color, amount):
            hex_color = hex_color.lstrip('#')
            if len(hex_color) == 3:
                hex_color = ''.join([c*2 for c in hex_color])
            if len(hex_color) != 6:
                return f"#{hex_color}"
            try:
                r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
                r = max(0, min(255, r + amount))
                g = max(0, min(255, g + amount))
                b = max(0, min(255, b + amount))
                return f"#{r:02x}{g:02x}{b:02x}"
            except:
                return f"#{hex_color}"

        def on_enter(e):
            if str(button.cget("state")) == "disabled":
                return
            bg = getattr(button, 'normal_bg', None)
            if not bg:
                bg = button.cget("bg")
                button.normal_bg = bg
            if bg.lower() in ["#e0e0e0", "#ffffff", "#f8f9fa", "#f0f0f0"]:
                hover_bg = adjust_color(bg, -10)
            else:
                hover_bg = adjust_color(bg, -25)
            button.config(bg=hover_bg)
            
        def on_leave(e):
            if str(button.cget("state")) == "disabled":
                return
            bg = getattr(button, 'normal_bg', None)
            if not bg:
                bg = button.cget("bg")
            button.config(bg=bg)
                
        button.bind("<Enter>", on_enter)
        button.bind("<Leave>", on_leave)

    def center_window(self, window, width, height):
        window.update_idletasks()
        try:
            parent = window.master
            sw = window.winfo_screenwidth()
            sh = window.winfo_screenheight()
            
            if parent and parent.winfo_viewable():
                px = parent.winfo_rootx()
                py = parent.winfo_rooty()
                pw = parent.winfo_width()
                ph = parent.winfo_height()
                
                # 當對話框寬度大於主視窗寬度時，改為相對於螢幕置中
                if width > pw:
                    x = (sw - width) // 2
                else:
                    x = px + (pw - width) // 2
                    
                # 當對話框高度大於主視窗高度時，改為相對於螢幕置中
                if height > ph:
                    y = (sh - height) // 2 - 30
                else:
                    y = py + (ph - height) // 2 - 30
            else:
                x = (sw - width) // 2
                y = (sh - height) // 2 - 40
        except Exception:
            sw = window.winfo_screenwidth()
            sh = window.winfo_screenheight()
            x = (sw - width) // 2
            y = (sh - height) // 2 - 40
            
        if y < 0: y = 0
        if x < 0: x = 0
            
        window.geometry(f"{width}x{height}+{int(x)}+{int(y)}")

    def update_court_label(self):
        court_no = system_settings.get("court_no", 1)
        enable_cloud = system_settings.get("enable_cloud", True)
        
        if not enable_cloud:
            self.lbl_court.config(text=f"第 {court_no} 場地\n(點擊顯示QR)", font=("Microsoft JhengHei", 12, "bold"), cursor="hand2")
            self.lbl_court.bind("<Button-1>", self.show_qr_popup)
        elif hasattr(self, 'cloudflare_url') and self.cloudflare_url:
            self.lbl_court.config(text=f"第 {court_no} 場地\n(點擊顯示QR)", font=("Microsoft JhengHei", 12, "bold"), cursor="hand2")
            self.lbl_court.bind("<Button-1>", self.show_qr_popup)
        else:
            self.lbl_court.config(text=f"第 {court_no} 場地\n(準備中...)", font=("Microsoft JhengHei", 12, "bold"), cursor="")
            self.lbl_court.unbind("<Button-1>")

    def build_header(self):
        header_container = tk.Frame(self.root, bg="white", pady=1)
        header_container.pack(fill="x")
        
        # 固定最上面三列的欄位長寬，防止被長文字撐開
        col_widths = [110, 190, 150, 130, 110, 190]
        for i in range(6):
            header_container.grid_columnconfigure(i, weight=1, minsize=col_widths[i], uniform="header_cols")
        for r in range(3):
            header_container.grid_rowconfigure(r, weight=1)
            
        self.lbl_type = tk.Label(header_container, text="Individual", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=12)
        self.lbl_type.grid(row=0, column=0, sticky="nsew", padx=2, pady=1)
        self.lbl_category = tk.Label(header_container, text="Senior 1 Male", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=20)
        self.lbl_category.grid(row=0, column=1, sticky="nsew", padx=2, pady=1)
        
        court_frame = tk.Frame(header_container, bg="#ddd", relief="raised", bd=2)
        court_frame.grid(row=0, column=2, rowspan=2, sticky="nsew", padx=2, pady=1)
        self.lbl_court = tk.Label(court_frame, text="", font=("Microsoft JhengHei", 12, "bold"), bg="#ddd", fg="blue", width=14)
        self.lbl_court.pack(fill="both", expand=True)
        
        no_frame = tk.Frame(header_container, bg="#ddd", relief="raised", bd=2)
        no_frame.grid(row=0, column=3, rowspan=2, sticky="nsew", padx=2, pady=1)
        self.lbl_no = tk.Label(no_frame, text="籤號：", font=("Microsoft JhengHei", 28, "bold"), bg="#ddd", fg="red", width=12)
        self.lbl_no.pack(fill="both", expand=True)
        
        self.lbl_1r_tag = tk.Label(header_container, text="1R", font=("Microsoft JhengHei", 10, "bold"), bg="white", fg="blue", relief="flat", width=12)
        self.lbl_1r_tag.grid(row=0, column=4, sticky="nsew", padx=2, pady=1)
        self.combo_poomsae_1 = ttk.Combobox(header_container, values=self.poomsae_list, font=("Microsoft JhengHei", 10), state="readonly", width=18)
        self.combo_poomsae_1.current(0) 
        self.combo_poomsae_1.grid(row=0, column=5, sticky="nsew", padx=2, pady=1)
        
        self.lbl_division = tk.Label(header_container, text="Final", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=12)
        self.lbl_division.grid(row=1, column=0, sticky="nsew", padx=2, pady=1)
        self.lbl_phase = tk.Label(header_container, text="Final Round", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=20)
        self.lbl_phase.grid(row=1, column=1, sticky="nsew", padx=2, pady=1)
        self.lbl_2r_tag = tk.Label(header_container, text="2R", font=("Microsoft JhengHei", 10, "bold"), bg="#eee", fg="#999", relief="flat", width=12)
        self.lbl_2r_tag.grid(row=1, column=4, sticky="nsew", padx=2, pady=1)
        self.combo_poomsae_2 = ttk.Combobox(header_container, values=self.poomsae_list, font=("Microsoft JhengHei", 10), state="readonly", width=18)
        self.combo_poomsae_2.current(0)
        self.combo_poomsae_2.grid(row=1, column=5, sticky="nsew", padx=2, pady=1)
        
        self.combo_poomsae_1.bind("<<ComboboxSelected>>", self.on_poomsae_combobox_changed)
        self.combo_poomsae_2.bind("<<ComboboxSelected>>", self.on_poomsae_combobox_changed)
        
        self.lbl_team_1 = tk.Label(header_container, text="TPE", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=12)
        self.lbl_team_1.grid(row=2, column=0, sticky="nsew", padx=2, pady=1)
        self.lbl_name_1 = tk.Label(header_container, text="CHEN, Xiao-Ming", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=20)
        self.lbl_name_1.grid(row=2, column=1, sticky="nsew", padx=2, pady=1)
        
        self.lbl_current_round = tk.Label(header_container, text="1R", font=("Microsoft JhengHei", 12, "bold"), bg="blue", fg="white", width=14)
        self.lbl_current_round.grid(row=2, column=2, sticky="nsew", padx=2, pady=1)
        self.lbl_timer = tk.Label(header_container, text="01:30", font=("Microsoft JhengHei", 16, "bold"), bg=self.colors["timer_bg"], fg=self.colors["timer_fg"], width=12)
        self.lbl_timer.grid(row=2, column=3, sticky="nsew", padx=2, pady=1)
        
        self.lbl_team_2 = tk.Label(header_container, text="TPE", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=12)
        self.lbl_team_2.grid(row=2, column=4, sticky="nsew", padx=2, pady=1)
        self.lbl_name_2 = tk.Label(header_container, text="CHEN, Xiao-Ming", font=self.font_info, bg=self.colors["info_bg"], relief="sunken", width=20)
        self.lbl_name_2.grid(row=2, column=5, sticky="nsew", padx=2, pady=1)
        
        # 初始化場地與連線標籤
        self.update_court_label()

    def build_left_panel(self, parent):
        left_frame = tk.Frame(parent, bg="white", width=300)
        left_frame.pack(side="left", fill="y", padx=5)
        table_frame = tk.Frame(left_frame, bg="white")
        table_frame.pack(fill="x", pady=5)
        headers = ["", "1R", "2R"]
        rows_split = ["正確性", "表現性", "扣分", "得分"]
        rows_merged = ["平均分", "總分"]
        for col, text in enumerate(headers):
            bg = self.colors["table_header"] if col > 0 else "white"
            tk.Label(table_frame, text=text, bg=bg, width=8, relief="solid", bd=1).grid(row=0, column=col, sticky="nsew")
        current_row = 1
        for i, row_text in enumerate(rows_split):
            tk.Label(table_frame, text=row_text, bg=self.colors["table_purple"], width=12, anchor="w", relief="solid", bd=1).grid(row=current_row, column=0, sticky="nsew")
            lbl_1 = tk.Label(table_frame, bg=self.colors["table_purple"], width=8, relief="solid", bd=1)
            lbl_1.grid(row=current_row, column=1, sticky="nsew")
            lbl_2 = tk.Label(table_frame, bg=self.colors["table_purple"], width=8, relief="solid", bd=1)
            lbl_2.grid(row=current_row, column=2, sticky="nsew")
            if i not in self.left_labels: self.left_labels[i] = {}
            self.left_labels[i][0] = lbl_1
            self.left_labels[i][1] = lbl_2
            current_row += 1
        for i, row_text in enumerate(rows_merged):
            idx = i + 4
            tk.Label(table_frame, text=row_text, bg=self.colors["table_purple"], width=12, anchor="w", relief="solid", bd=1).grid(row=current_row, column=0, sticky="nsew")
            lbl_merged = tk.Label(table_frame, bg=self.colors["table_purple"], width=16, relief="solid", bd=1)
            lbl_merged.grid(row=current_row, column=1, columnspan=2, sticky="nsew")
            self.left_merged_labels[idx] = lbl_merged
            current_row += 1
        
        control_frame = tk.Frame(left_frame, bg="white", pady=5)
        control_frame.pack(fill="x", side="bottom")
        
        self.btn_select_folder = tk.Button(control_frame, text="📁 設定 Excel 資料夾", font=("Microsoft JhengHei", 9, "bold"), bg=self.colors["btn_yellow"], relief="flat", bd=0, pady=6, command=self.select_folder)
        self.btn_select_folder.pack(fill="x", pady=(0, 2))
        
        self.lbl_excel_dir = tk.Label(control_frame, text="📁 未設定資料夾", font=("Microsoft JhengHei", 8), fg="#7f8c8d", bg="white", wraplength=180, justify="left", anchor="w")
        self.lbl_excel_dir.pack(fill="x", pady=(0, 5))
        
        session_frame = tk.Frame(control_frame, bg="white")
        session_frame.pack(fill="x", pady=2)
        
        # 選擇比賽場次標籤與 Combobox 放在同一行
        tk.Label(session_frame, text="選擇比賽場次：", bg="white", font=("Microsoft JhengHei", 9, "bold")).pack(side="left", padx=(0, 5))
        
        # Combobox 設定 (使用 postcommand 展開時重置為完整列表，字型 10)
        self.cb_session_select = ttk.Combobox(session_frame, font=("Microsoft JhengHei", 10), postcommand=self.reset_session_combobox_values, width=15)
        self.cb_session_select.pack(side="left", fill="x", expand=True)
        
        # === 優化: 點擊輸入框時自動全選，方便覆蓋輸入 ===
        self.cb_session_select.bind("<FocusIn>", lambda e: self.cb_session_select.select_range(0, tk.END))
        self.cb_session_select.bind("<Button-1>", lambda e: self.cb_session_select.select_range(0, tk.END) if self.cb_session_select.focus_get() == self.cb_session_select else None)
        
        # === 綁定事件 ===
        self.cb_session_select.bind("<<ComboboxSelected>>", self.on_session_selected)
        self.cb_session_select.bind("<KeyRelease>", self.on_session_type) 
        self.cb_session_select.bind("<Return>", self.on_session_submit)
        
        # 新增按鈕區 (排在下拉選單下方，包含「上一場」與「下一場」按鈕，改用 Grid 以平分寬度)
        session_btn_frame = tk.Frame(control_frame, bg="white")
        session_btn_frame.pack(fill="x", pady=2)
        
        session_btn_frame.columnconfigure(0, weight=1)
        session_btn_frame.columnconfigure(1, weight=1)
        
        # 1. 上一場次按鈕 (靠左側 column 0)
        self.btn_prev_session = tk.Button(
            session_btn_frame,
            text="⬅️ 上一場",
            font=("Microsoft JhengHei", 9, "bold"),
            bg="#bdc3c7",
            fg="#2c3e50",
            relief="flat",
            bd=0,
            pady=3,
            cursor="hand2",
            command=self.goto_prev_session
        )
        self.btn_prev_session.grid(row=0, column=0, sticky="nsew", padx=(0, 2))
        
        # 為上一場次按鈕加上 Hover 變色效果
        def on_prev_session_enter(e):
            self.btn_prev_session.config(bg="#95a5a6")
        def on_prev_session_leave(e):
            self.btn_prev_session.config(bg="#bdc3c7")
        self.btn_prev_session.bind("<Enter>", on_prev_session_enter)
        self.btn_prev_session.bind("<Leave>", on_prev_session_leave)
        
        # 2. 下一場次按鈕 (靠右側 column 1)
        self.btn_next_session = tk.Button(
            session_btn_frame,
            text="下一場 ➡️",
            font=("Microsoft JhengHei", 9, "bold"),
            bg="#f1c40f",
            fg="#2c3e50",
            relief="flat",
            bd=0,
            pady=3,
            cursor="hand2",
            command=self.goto_next_session
        )
        self.btn_next_session.grid(row=0, column=1, sticky="nsew", padx=(2, 0))
        
        # 為下一場次按鈕加上 Hover 變色效果
        def on_next_session_enter(e):
            self.btn_next_session.config(bg="#f39c12")
        def on_next_session_leave(e):
            self.btn_next_session.config(bg="#f1c40f")
        self.btn_next_session.bind("<Enter>", on_next_session_enter)
        self.btn_next_session.bind("<Leave>", on_next_session_leave)
        
        radio_frame = tk.Frame(left_frame, bg="white")
        radio_frame.pack(fill="x", pady=5)
        
        # 標題提示 (讓賽制群組看起來更有層次，改用 grid 避免幾何管理器衝突)
        tk.Label(radio_frame, text="賽制模式切換：", bg="white", font=("Microsoft JhengHei", 9, "bold"), anchor="w").grid(row=0, column=0, columnspan=2, sticky="w", pady=(0, 3))
        
        self.mode_var = tk.IntVar(value=0)
        self.mode_buttons = []
        
        opts = [
            ("Cutoff 一般", 0),
            ("PK 評分 (PK)", 1),
            ("自由品勢 (FL)", 2),
            ("快速比賽", 3)
        ]
        
        # 設定 grid 權重使按鈕平分寬度
        radio_frame.columnconfigure(0, weight=1)
        radio_frame.columnconfigure(1, weight=1)
        
        for idx, (text, val) in enumerate(opts):
            r = (idx // 2) + 1  # 從 row 1 開始，因 row 0 已被標題 Label 佔用
            c = idx % 2
            btn = tk.Button(
                radio_frame,
                text=text,
                font=("Microsoft JhengHei", 9, "bold"),
                relief="flat",
                bd=0,
                height=1,
                anchor="center",
                padx=5,
                pady=6,
                cursor="hand2"
            )
            btn.grid(row=r, column=c, sticky="ew", padx=2, pady=2)
            btn.config(command=lambda v=val: self.select_mode(v))
            self.setup_mode_btn_hover(btn, val)
            self.mode_buttons.append((btn, val))
            
        self.update_mode_button_styles()

    def on_session_type(self, event):
        """當使用者在下拉選單輸入文字時，過濾選項 (模糊搜尋)"""
        # 如果按的是 Enter 或上下鍵，不進行過濾處理，交給其他事件
        if event.keysym in ['Return', 'Up', 'Down']:
            return

        val = self.cb_session_select.get()
        
        # 取得所有 ID
        all_ids = list(set(d["SourceFile"] for d in self.imported_matches.values()))
        try: all_ids.sort(key=lambda x: int(x))
        except: all_ids.sort()
        
        if val == '':
            self.cb_session_select['values'] = all_ids
        else:
            # 過濾: 只要包含輸入字串就顯示 (Case insensitive)
            filtered = [x for x in all_ids if val.lower() in x.lower()]
            self.cb_session_select['values'] = filtered
            
            # 保持下拉選單開啟 (這行在某些系統可能無效，但不影響功能)
            # try: self.cb_session_select.event_generate('<Down>')
            # except: pass

    def on_session_submit(self, event=None):
        """當按下 Enter 或失焦時，讀取輸入框內容並更新列表"""
        target = self.cb_session_select.get().strip()
        
        # 取得所有有效的 ID
        all_ids = list(set(d["SourceFile"] for d in self.imported_matches.values()))
        
        # 如果輸入的內容是有效的場次 ID
        if target in all_ids:
            # 呼叫更新 Treeview
            self.update_tree_columns()
            
            # (選用) 將焦點移回主視窗，避免誤觸 Combobox
            self.root.focus()
        else:
            # 如果輸入無效 (找不到該場次)，清空 Treeview
            for t in [self.tree_ready, self.tree_end]:
                for item in t.get_children():
                    t.delete(item)

    def select_mode(self, val):
        old_val = self.mode_var.get()
        if old_val == val:
            return
        self.mode_var.set(val)
        self.update_mode_button_styles()
        self.update_tree_columns()
        
        # 當賽制切換且投影視窗開啟時，自動無縫重建對應賽制的投影視窗
        if hasattr(self, 'proj_window') and self.proj_window is not None and self.proj_window.winfo_exists():
            self.switch_projection_mode()

    def update_mode_button_styles(self):
        current_val = self.mode_var.get()
        for btn, val in self.mode_buttons:
            if val == current_val:
                btn.config(bg="#3498db", fg="white", activebackground="#2980b9", activeforeground="white")
            else:
                btn.config(bg="#f5f6fa", fg="#2c3e50", activebackground="#e1e8ed", activeforeground="#2c3e50")

    def setup_mode_btn_hover(self, btn, val):
        def on_enter(e):
            current_val = self.mode_var.get()
            if val != current_val:
                btn.config(bg="#e1e8ed")
        def on_leave(e):
            current_val = self.mode_var.get()
            if val != current_val:
                btn.config(bg="#f5f6fa")
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

    def build_center_panel(self, parent):
        center_frame = tk.Frame(parent, bg="white")
        center_frame.pack(side="left", fill="both", expand=True, padx=5)
        for i in range(7): center_frame.grid_columnconfigure(i, weight=1, uniform="center_cols")
        h_cols = ["正確性", "表現性", "總分", "裁判", "總分", "表現性", "正確性"]
        for col_idx, text in enumerate(h_cols):
            bg_col = self.colors["judge_blue"] if text == "裁判" else self.colors["cyan_header"]
            lbl = tk.Label(center_frame, text=text, bg=bg_col, fg="white", font=self.font_bold, relief="solid", bd=1, height=1)
            lbl.grid(row=0, column=col_idx, sticky="nsew")
        self.score_cells = []
        for r in range(1, 8):
            row_cells = []
            for c in range(7):
                bg = "#f9f9f9"
                lbl = tk.Label(center_frame, text="", bg=bg, relief="solid", bd=1, height=1)
                lbl.grid(row=r, column=c, sticky="nsew")
                row_cells.append(lbl)
            self.score_cells.append(row_cells)
            center_frame.grid_rowconfigure(r, weight=1)
        self.center_stats_labels = {} 
        stats_headers = ["總分 (Total)", "平均分 (Avg)"]
        for i, label_text in enumerate(stats_headers):
            curr_row = 8 + i
            key_prefix = "Total" if i == 0 else "Avg"
            for c in range(3):
                lbl = tk.Label(center_frame, text="0.00", bg=self.colors["lime_val_bg"], font=self.font_bold, relief="solid", bd=1)
                lbl.grid(row=curr_row, column=c, sticky="nsew", pady=1)
                self.center_stats_labels[f"{key_prefix}_L_{c}"] = lbl
            tk.Label(center_frame, text=label_text, bg=self.colors["lime_dark"], fg="white", font=self.font_bold, relief="solid", bd=1).grid(row=curr_row, column=3, sticky="nsew", pady=1)
            for c in range(3):
                col_idx = 4 + c
                lbl = tk.Label(center_frame, text="", bg=self.colors["lime_val_bg"], font=self.font_bold, relief="solid", bd=1)
                lbl.grid(row=curr_row, column=col_idx, sticky="nsew", pady=1)
                self.center_stats_labels[f"{key_prefix}_R_{c}"] = lbl
        r_ded = 10
        tk.Checkbutton(center_frame, text="Withdraw", bg="white", anchor="w").grid(row=r_ded, column=0, sticky="nsew")
        f_btns1 = tk.Frame(center_frame, bg="white")
        f_btns1.grid(row=r_ded, column=1, sticky="nsew", padx=1)
        tk.Button(f_btns1, text="-", bg=self.colors["btn_orange"], width=2, relief="raised", command=lambda: self.update_deduction(-0.1)).pack(side="left", fill="y", expand=True)
        tk.Button(f_btns1, text="+", bg=self.colors["btn_orange"], width=2, relief="raised", command=lambda: self.update_deduction(0.1)).pack(side="left", fill="y", expand=True)
        self.lbl_deduction_val = tk.Label(center_frame, text="", bg=self.colors["lime_val_bg"], relief="solid", bd=1)
        self.lbl_deduction_val.grid(row=r_ded, column=2, sticky="nsew", pady=1)
        tk.Label(center_frame, text="扣分 (Deduction)", bg=self.colors["lime_dark"], fg="white", font=self.font_bold, relief="solid", bd=1).grid(row=r_ded, column=3, sticky="nsew", pady=1)
        self.lbl_deduction_val_R = tk.Label(center_frame, text="", bg=self.colors["lime_val_bg"], relief="solid", bd=1)
        self.lbl_deduction_val_R.grid(row=r_ded, column=4, sticky="nsew", pady=1)
        f_btns2 = tk.Frame(center_frame, bg="white")
        f_btns2.grid(row=r_ded, column=5, sticky="nsew", padx=1)
        tk.Checkbutton(center_frame, text="Withdraw", bg="white", anchor="w").grid(row=r_ded, column=6, sticky="nsew")
        r_fin = 11
        tk.Checkbutton(center_frame, text="Disqualification", bg="white", anchor="w").grid(row=r_fin, column=0, sticky="nsew")
        tk.Label(center_frame, bg="white").grid(row=r_fin, column=1, sticky="nsew")
        self.lbl_final_L = tk.Label(center_frame, text="", bg=self.colors["lime_green"], font=self.font_bold, relief="solid", bd=1)
        self.lbl_final_L.grid(row=r_fin, column=2, sticky="nsew", pady=1)
        tk.Label(center_frame, text="最終得分 (Final)", bg=self.colors["lime_dark"], fg="white", font=self.font_bold, relief="solid", bd=1).grid(row=r_fin, column=3, sticky="nsew", pady=1)
        self.lbl_final_R = tk.Label(center_frame, text="", bg=self.colors["lime_green"], font=self.font_bold, relief="solid", bd=1)
        self.lbl_final_R.grid(row=r_fin, column=4, sticky="nsew", pady=1)
        tk.Label(center_frame, bg="white").grid(row=r_fin, column=5, sticky="nsew")
        tk.Checkbutton(center_frame, text="Disqualification", bg="white", anchor="w").grid(row=r_fin, column=6, sticky="nsew")
        self.result_vars["Final Points"] = self.lbl_final_L

    def build_right_panel(self, parent):
        right_frame = tk.Frame(parent, bg="white", width=300)
        right_frame.pack(side="right", fill="y", padx=5)
        table_frame = tk.Frame(right_frame, bg="white")
        table_frame.pack(fill="x", pady=5)
        headers = ["", "1R", "2R"]
        rows_split = ["正確性", "表現性", "扣分", "得分"]
        rows_merged = ["平均分", "總分"]
        for col, text in enumerate(headers):
            bg = self.colors["table_header"] if col > 0 else "white"
            tk.Label(table_frame, text=text, bg=bg, width=8, relief="solid", bd=1).grid(row=0, column=col, sticky="nsew")
        current_row = 1
        for i, row_text in enumerate(rows_split):
            tk.Label(table_frame, text=row_text, bg=self.colors["table_purple"], width=12, anchor="w", relief="solid", bd=1).grid(row=current_row, column=0, sticky="nsew")
            lbl_1 = tk.Label(table_frame, bg=self.colors["table_purple"], width=8, relief="solid", bd=1)
            lbl_1.grid(row=current_row, column=1, sticky="nsew")
            lbl_2 = tk.Label(table_frame, bg=self.colors["table_purple"], width=8, relief="solid", bd=1)
            lbl_2.grid(row=current_row, column=2, sticky="nsew")
            if i not in self.right_labels: self.right_labels[i] = {}
            self.right_labels[i][0] = lbl_1
            self.right_labels[i][1] = lbl_2
            current_row += 1
        for i, row_text in enumerate(rows_merged):
            idx = i + 4
            tk.Label(table_frame, text=row_text, bg=self.colors["table_purple"], width=12, anchor="w", relief="solid", bd=1).grid(row=current_row, column=0, sticky="nsew")
            lbl_merged = tk.Label(table_frame, bg=self.colors["table_purple"], width=16, relief="solid", bd=1)
            lbl_merged.grid(row=current_row, column=1, columnspan=2, sticky="nsew")
            self.right_merged_labels[idx] = lbl_merged
            current_row += 1
            
        btn_box = tk.Frame(right_frame, bg="white", pady=10)
        btn_box.pack(fill="x")
        
        # 語意化配色與圖示系統
        btn_defs = [
            ("⚙️ 系統設定", self.open_settings, "#3498db", "#ffffff"),
            ("➕ 建立比賽", self.create_match, "#2ecc71", "#ffffff"),
            ("📝 修改比賽", self.update_match, "#f1c40f", "#2c3e50"),
            ("❌ 刪除比賽", self.delete_match, "#e74c3c", "#ffffff"),
            ("🔌 釋放全部連線", self.release_all_connections, "#7f8c8d", "#ffffff")
        ]
        
        btn_font = ("Microsoft JhengHei", 9, "bold")
        for txt, cmd, bg_col, fg_col in btn_defs:
            btn = tk.Button(btn_box, text=txt, font=btn_font, bg=bg_col, fg=fg_col, relief="flat", height=2, command=cmd)
            btn.pack(fill="x", pady=4)
            self.setup_dynamic_hover(btn)

    def build_bottom_panel(self):
        bottom_frame = tk.Frame(self.root, bg="white", height=200)
        bottom_frame.pack(fill="x", side="bottom", padx=5, pady=5)
        self.btn_row = tk.Frame(bottom_frame, bg="white")
        self.btn_row.pack(fill="x", pady=5)
        
        btn_font = ("Microsoft JhengHei", 9, "bold")
        
        def style_btn(btn, default_bg="white", default_fg="#2c3e50"):
            btn.config(font=btn_font, relief="flat", bg=default_bg, fg=default_fg, height=2, width=10)
            self.setup_dynamic_hover(btn)
            return btn
            
        self.btn_draw = tk.Button(self.btn_row, text="抽型場", command=self.draw_poomsae)
        style_btn(self.btn_draw, self.colors["btn_pink"], "#2c3e50")
        
        self.btn_ready = tk.Button(self.btn_row, text="準備", command=self.toggle_ready)
        style_btn(self.btn_ready)
        
        self.btn_start = tk.Button(self.btn_row, text="開始", command=self.toggle_start)
        style_btn(self.btn_start)
        
        self.btn_show_score = tk.Button(self.btn_row, text="展示分數", state="disabled", command=self.show_scores)
        style_btn(self.btn_show_score, "#e0e0e0", "#7f8c8d")
        
        self.btn_next = tk.Button(self.btn_row, text="下一品", state="disabled", command=self.dispatch_next_action)
        style_btn(self.btn_next, "#e0e0e0", "#7f8c8d")
        
        self.btn_end = tk.Button(self.btn_row, text="結束", state="disabled", command=self.end_match)
        style_btn(self.btn_end, "#e0e0e0", "#7f8c8d")
        
        self.btn_next_player = tk.Button(self.btn_row, text="選取下一位", command=self.select_next_player)
        style_btn(self.btn_next_player, "#1abc9c", "#ffffff")
        
        self.btn_unselect = tk.Button(self.btn_row, text="取消選擇", command=self.unselect_match)
        style_btn(self.btn_unselect, "#bdc3c7", "#ffffff")
        
        self.btn_reshow = tk.Button(self.btn_row, text="重新展示", state="disabled", command=self.reshow_scores)
        style_btn(self.btn_reshow, "#e0e0e0", "#7f8c8d")
        
        self.btn_log = tk.Button(self.btn_row, text="Log紀錄", command=self.open_log)
        style_btn(self.btn_log, self.colors["btn_pink"], "#2c3e50")
        self.btn_re_eval = tk.Button(self.btn_row, text="本輪重評", command=self.re_evaluate_round)
        style_btn(self.btn_re_eval, self.colors["btn_orange"], "#ffffff")
        
        self.btn_proj = tk.Button(self.btn_row, text="投影畫面", command=self.toggle_projection)
        style_btn(self.btn_proj, self.colors["btn_yellow"], "#2c3e50")
        
        self.tip_lbl = tk.Label(self.btn_row, text="💡 提示：按兩下 (雙擊) 選手項目，或選取項目後按 Enter 鍵以載入比賽資料。", font=("Microsoft JhengHei", 9, "bold"), fg="#7f8c8d", bg="white", anchor="w")
        
        self.refresh_bottom_buttons()

        # 建立 Notebook 標籤頁以將 Ready 與 End 選手分開
        self.notebook = ttk.Notebook(bottom_frame)
        self.notebook.pack(fill="both", expand=True)
        
        self.tab_ready = tk.Frame(self.notebook, bg="white")
        self.tab_end = tk.Frame(self.notebook, bg="white")
        
        self.notebook.add(self.tab_ready, text=" 未比賽選手 (Ready) ")
        self.notebook.add(self.tab_end, text=" 已完賽選手 (End) ")
        
        cols = ("籤號", "狀態", "姓名", "NOC", "組別", "級別", "輪次", "單位", "類型", "比賽模式")
        
        # 1. 建立 Ready 選手 Treeview
        ready_scroll = tk.Scrollbar(self.tab_ready)
        ready_scroll.pack(side="right", fill="y")
        self.tree_ready = ttk.Treeview(self.tab_ready, columns=cols, show="headings", yscrollcommand=ready_scroll.set, height=10)
        for col in cols:
            self.tree_ready.heading(col, text=col)
            self.tree_ready.column(col, width=80, anchor="center")
        self.tree_ready.pack(fill="both", expand=True)
        ready_scroll.config(command=self.tree_ready.yview)
        
        self.tree_ready.bind("<Double-Button-1>", self.on_match_select)
        self.tree_ready.bind("<Return>", self.on_match_select)
        
        # 2. 建立 End 選手 Treeview
        end_scroll = tk.Scrollbar(self.tab_end)
        end_scroll.pack(side="right", fill="y")
        self.tree_end = ttk.Treeview(self.tab_end, columns=cols, show="headings", yscrollcommand=end_scroll.set, height=10)
        for col in cols:
            self.tree_end.heading(col, text=col)
            self.tree_end.column(col, width=80, anchor="center")
        self.tree_end.pack(fill="both", expand=True)
        end_scroll.config(command=self.tree_end.yview)
        
        self.tree_end.bind("<Double-Button-1>", self.on_match_select)
        self.tree_end.bind("<Return>", self.on_match_select)
        
        # 預設將 self.tree 指向 Ready Treeview，並綁定標籤切換虛擬事件
        self.tree = self.tree_ready
        self.notebook.bind("<<NotebookTabChanged>>", self.on_tab_changed)

    def refresh_bottom_buttons(self):
        # 先忘記所有按鈕的排版
        self.btn_draw.pack_forget()
        self.btn_ready.pack_forget()
        self.btn_start.pack_forget()
        self.btn_show_score.pack_forget()
        self.btn_next.pack_forget()
        self.btn_end.pack_forget()
        if hasattr(self, 'btn_next_player'): self.btn_next_player.pack_forget()
        if hasattr(self, 'btn_unselect'): self.btn_unselect.pack_forget()
        if hasattr(self, 'btn_reshow'): self.btn_reshow.pack_forget()
        if hasattr(self, 'btn_log'): self.btn_log.pack_forget()
        if hasattr(self, 'btn_re_eval'): self.btn_re_eval.pack_forget()
        if hasattr(self, 'btn_proj'): self.btn_proj.pack_forget()
        if hasattr(self, 'tip_lbl'): self.tip_lbl.pack_forget()
        
        # 根據設定決定是否顯示抽選按鈕
        if system_settings.get("show_draw_button", True):
            self.btn_draw.pack(side="left", padx=2)
            
        self.btn_ready.pack(side="left", padx=2)
        self.btn_start.pack(side="left", padx=2)
        self.btn_show_score.pack(side="left", padx=2)
        self.btn_next.pack(side="left", padx=2)
        self.btn_end.pack(side="left", padx=2)
        
        if hasattr(self, 'btn_next_player'): self.btn_next_player.pack(side="left", padx=2)
        if hasattr(self, 'btn_unselect'): self.btn_unselect.pack(side="left", padx=2)
        if hasattr(self, 'btn_reshow'): self.btn_reshow.pack(side="left", padx=2)
        if hasattr(self, 'btn_log'): self.btn_log.pack(side="left", padx=2)
        if hasattr(self, 'btn_re_eval'): self.btn_re_eval.pack(side="left", padx=2)
        if hasattr(self, 'tip_lbl'): self.tip_lbl.pack(side="left", fill="both", expand=True, padx=10)
        if hasattr(self, 'btn_proj'): self.btn_proj.pack(side="right", padx=2)
        
        self.update_button_states()

    def refresh_judge_slots(self):
        judge_count = int(system_settings["judge_count"])
        
        # 收集當前已連線的裁判 ID 列表
        connected_judges = set()
        for sid, jdata in current_state['judges'].items():
            jid = jdata.get('id', '')
            if jid and (jdata.get('connected', False) or sid.startswith('manual_')):
                connected_judges.add(jid)
                
        for i in range(7):
            label = self.score_cells[i][3]
            jid_str = f"J{i+1}"
            manual_jid_str = f"manual_J{i+1}"
            is_connected = (jid_str in connected_judges or manual_jid_str in connected_judges)
            
            if i < judge_count:
                fg_color = "#39ff14" if is_connected else "white"  # 已連線顯示亮綠色
                label.config(text=f"Judge {i+1}", bg=self.colors["judge_blue"], fg=fg_color, cursor="hand2")
                label.bind("<Button-1>", lambda e, idx=i: self.open_manual_judge_input(idx + 1))
            else:
                label.config(text="", bg="#f9f9f9", fg="white", cursor="")
                label.unbind("<Button-1>")
        self.update_live_scores()
    def open_manual_judge_input(self, judge_num):
        if not self.is_locked:
            jid_str = f"J{judge_num}"
            is_online = False
            for old_sid, jd in list(current_state['judges'].items()):
                if jd.get('id') == jid_str and not old_sid.startswith('manual_') and jd.get('connected', False):
                    is_online = True
                    break
            
            if is_online:
                ans = messagebox.askyesno("釋放連線", f"目前比賽尚未準備，無法手動輸入分數。\n\n是否要強制將 裁判 {jid_str} 斷線並釋放名額？", parent=self.root)
                if ans:
                    released = False
                    for old_sid, jd in list(current_state['judges'].items()):
                        if jd.get('id') == jid_str and not old_sid.startswith('manual_'):
                            jd['connected'] = False
                            jd['kicked'] = True
                            try:
                                get_web_server().socketio.emit('force_disconnect', {}, room=old_sid)  # type: ignore
                            except:
                                pass
                            released = True
                    if released:
                        messagebox.showinfo("成功", f"裁判 {jid_str} 的連線已強制釋放，該設備已被踢回登入畫面。", parent=self.root)
                        self.refresh_judge_slots()
                return
            else:
                messagebox.showwarning("警告", "請先按下「準備」並鎖定比賽後，才能輸入成績。", parent=self.root)
                return
        
        top = tk.Toplevel(self.root)
        top.title(f"手動輸入 - Judge {judge_num}")
        
        is_pk = (self.mode_var.get() == 1)
        pk_seq = int(config.system_settings.get('pk_sequence_mode', 0))
        is_pk_simul = is_pk and (pk_seq == 0)
        is_pk_seq = is_pk and (pk_seq == 1 or pk_seq == 2)
        current_side = config.current_state.get('current_player_side', 0)
        
        if is_pk_simul:
            self.center_window(top, 1850, 700)
        else:
            self.center_window(top, 950, 620)
            
        top.grab_set()
        
        main_frame = tk.Frame(top, padx=15, pady=10)
        main_frame.pack(fill="both", expand=True)
        
        fake_sid = f"manual_J{judge_num}"
        
        def_acc = 4.0
        def_p1 = 2.0
        def_p2 = 2.0
        def_p3 = 2.0
        
        def_hong_acc = 4.0
        def_hong_p1 = 2.0
        def_hong_p2 = 2.0
        def_hong_p3 = 2.0
        
        found_judge_data = None
        if fake_sid in current_state['judges']:
            found_judge_data = current_state['judges'][fake_sid]
        else:
            jid_str = f"J{judge_num}"
            for old_sid, jd in current_state['judges'].items():
                if jd.get('id') == jid_str and not old_sid.startswith('manual_'):
                    found_judge_data = jd
                    break
                    
        if found_judge_data:
            if is_pk_seq and current_side == 1:
                def_acc = found_judge_data.get('hong_acc', 4.0)
                def_p1 = found_judge_data.get('hong_p1', 2.0)
                def_p2 = found_judge_data.get('hong_p2', 2.0)
                def_p3 = found_judge_data.get('hong_p3', 2.0)
            else:
                def_acc = found_judge_data.get('acc', 4.0)
                def_p1 = found_judge_data.get('p1', 2.0)
                def_p2 = found_judge_data.get('p2', 2.0)
                def_p3 = found_judge_data.get('p3', 2.0)
            
            def_hong_acc = found_judge_data.get('hong_acc', 4.0)
            def_hong_p1 = found_judge_data.get('hong_p1', 2.0)
            def_hong_p2 = found_judge_data.get('hong_p2', 2.0)
            def_hong_p3 = found_judge_data.get('hong_p3', 2.0)
            
        var_acc = tk.DoubleVar(value=def_acc)
        var_pres1 = tk.DoubleVar(value=def_p1)
        var_pres2 = tk.DoubleVar(value=def_p2)
        var_pres3 = tk.DoubleVar(value=def_p3)
        
        var_hong_acc = tk.DoubleVar(value=def_hong_acc)
        var_hong_pres1 = tk.DoubleVar(value=def_hong_p1)
        var_hong_pres2 = tk.DoubleVar(value=def_hong_p2)
        var_hong_pres3 = tk.DoubleVar(value=def_hong_p3)
        
        total_label_var = tk.StringVar()
        
        def update_total(*args):
            t_chung = round(var_acc.get() + var_pres1.get() + var_pres2.get() + var_pres3.get(), 1)
            if is_pk_simul:
                t_hong = round(var_hong_acc.get() + var_hong_pres1.get() + var_hong_pres2.get() + var_hong_pres3.get(), 1)
                total_label_var.set(f"青方總分: {t_chung:.1f}   |   紅方總分: {t_hong:.1f}")
            else:
                total_label_var.set(f"總分: {t_chung:.1f}")
                
        update_total()
        
        def create_category_section(parent, label_text, variable, max_val, row, col):
            frame = tk.LabelFrame(parent, text="", font=("Microsoft JhengHei", 11, "bold"), padx=10, pady=5, bd=2, relief="groove")
            frame.grid(row=row, column=col, padx=8, pady=5, sticky="nsew")
            
            header = tk.Frame(frame)
            header.pack(fill="x", pady=2)
            
            lbl_title = tk.Label(header, text=label_text, font=("Microsoft JhengHei", 11, "bold"))
            lbl_title.pack(side="left")
            
            lbl_score_val = tk.Label(header, text=f"{variable.get():.1f} / {max_val:.1f}", font=("Microsoft JhengHei", 11, "bold"), fg="#16a085")
            lbl_score_val.pack(side="right")
            
            entry_var = tk.StringVar(value=f"{variable.get():.1f}")
            entry = tk.Entry(header, textvariable=entry_var, width=5, font=("Microsoft JhengHei", 10, "bold"), justify="center")
            entry.pack(side="right", padx=10)
            
            grid_frame = tk.Frame(frame)
            grid_frame.pack(fill="both", expand=True, pady=5)
            
            steps = int(max_val * 10) + 1
            buttons = []
            
            def highlight_button(val):
                for btn in buttons:
                    if abs(btn.val - val) < 0.001:
                        btn.config(bg="#2ecc71", fg="white", font=("Microsoft JhengHei", 9, "bold"))
                    else:
                        btn.config(bg="#f5f6fa", fg="#2f3640", font=("Microsoft JhengHei", 9))
            
            def on_score_update(val, from_entry=False):
                val = max(0.0, min(max_val, round(val, 1)))
                variable.set(val)
                lbl_score_val.config(text=f"{val:.1f} / {max_val:.1f}")
                highlight_button(val)
                update_total()
                if not from_entry:
                    entry_var.set(f"{val:.1f}")
            
            def on_entry_change(*args):
                try:
                    text = entry_var.get().strip()
                    if text == "": return
                    val = float(text)
                    if 0.0 <= val <= max_val:
                        on_score_update(val, from_entry=True)
                except ValueError:
                    pass
            
            entry_var.trace_add("write", on_entry_change)
            
            def on_entry_focus(e):
                entry.select_range(0, tk.END)
                entry.icursor(tk.END)
            entry.bind("<FocusIn>", on_entry_focus)
            
            for i in range(steps):
                val = round(i * 0.1, 1)
                btn = tk.Button(grid_frame, text=f"{val:.1f}", width=3, font=("Microsoft JhengHei", 9), relief="raised", takefocus=False)
                btn.val = val
                btn.config(command=lambda v=val: on_score_update(v))
                
                r = i // 10
                c = i % 10
                btn.grid(row=r, column=c, padx=1, pady=1)
                buttons.append(btn)
                
            highlight_button(variable.get())
            return entry

        if is_pk_simul:
            main_frame.columnconfigure(0, weight=1)
            main_frame.columnconfigure(1, weight=1)
            main_frame.rowconfigure(0, weight=1)
            main_frame.rowconfigure(1, weight=0)
            
            chung_container = tk.LabelFrame(main_frame, text="青方 (Chung / Blue)", font=("Microsoft JhengHei", 12, "bold"), fg="#0099cc", padx=10, pady=5)
            chung_container.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
            chung_container.columnconfigure(0, weight=1)
            chung_container.columnconfigure(1, weight=1)
            chung_container.rowconfigure(0, weight=1)
            chung_container.rowconfigure(1, weight=1)
            
            hong_container = tk.LabelFrame(main_frame, text="紅方 (Hong / Red)", font=("Microsoft JhengHei", 12, "bold"), fg="#cc0000", padx=10, pady=5)
            hong_container.grid(row=0, column=1, padx=10, pady=5, sticky="nsew")
            hong_container.columnconfigure(0, weight=1)
            hong_container.columnconfigure(1, weight=1)
            hong_container.rowconfigure(0, weight=1)
            hong_container.rowconfigure(1, weight=1)
            
            entry_acc = create_category_section(chung_container, "正確性 (Accuracy)", var_acc, 4.0, 0, 0)
            entry_p1 = create_category_section(chung_container, "速度與力量 (Speed & Power)", var_pres1, 2.0, 1, 0)
            entry_p2 = create_category_section(chung_container, "節奏與協調 (Rhythm)", var_pres2, 2.0, 0, 1)
            entry_p3 = create_category_section(chung_container, "精神表現 (Spirit/Expression)", var_pres3, 2.0, 1, 1)
            
            entry_hong_acc = create_category_section(hong_container, "正確性 (Accuracy)", var_hong_acc, 4.0, 0, 0)
            entry_hong_p1 = create_category_section(hong_container, "速度與力量 (Speed & Power)", var_hong_pres1, 2.0, 1, 0)
            entry_hong_p2 = create_category_section(hong_container, "節奏與協調 (Rhythm)", var_hong_pres2, 2.0, 0, 1)
            entry_hong_p3 = create_category_section(hong_container, "精神表現 (Spirit/Expression)", var_hong_pres3, 2.0, 1, 1)
            
            bottom_row_idx = 1
        else:
            main_frame.columnconfigure(0, weight=1)
            main_frame.rowconfigure(0, weight=1)
            main_frame.rowconfigure(1, weight=0)
            
            title_text = "手動輸入"
            fg_color = "black"
            if is_pk_seq and current_side == 0:
                title_text = "青方 (Chung / Blue) 手動輸入"
                fg_color = "#0099cc"
            elif is_pk_seq and current_side == 1:
                title_text = "紅方 (Hong / Red) 手動輸入"
                fg_color = "#cc0000"
                
            container = tk.LabelFrame(main_frame, text=title_text, font=("Microsoft JhengHei", 12, "bold"), fg=fg_color, padx=10, pady=5)
            container.grid(row=0, column=0, padx=10, pady=5, sticky="nsew")
            container.columnconfigure(0, weight=1)
            container.columnconfigure(1, weight=1)
            container.rowconfigure(0, weight=1)
            container.rowconfigure(1, weight=1)
            
            entry_acc = create_category_section(container, "正確性 (Accuracy)", var_acc, 4.0, 0, 0)
            entry_p1 = create_category_section(container, "速度與力量 (Speed & Power)", var_pres1, 2.0, 1, 0)
            entry_p2 = create_category_section(container, "節奏與協調 (Rhythm)", var_pres2, 2.0, 0, 1)
            entry_p3 = create_category_section(container, "精神表現 (Spirit/Expression)", var_pres3, 2.0, 1, 1)
            
            bottom_row_idx = 1
            
        bottom_frame = tk.Frame(main_frame, pady=10)
        bottom_frame.grid(row=bottom_row_idx, column=0, columnspan=2, sticky="ew")
        
        lbl_total = tk.Label(bottom_frame, textvariable=total_label_var, font=("Microsoft JhengHei", 18, "bold"), fg="#c0392b")
        lbl_total.pack(side="left", padx=10)
        
        def release_connection():
            jid_str = f"J{judge_num}"
            released = False
            for old_sid, jd in list(current_state['judges'].items()):
                if jd.get('id') == jid_str and not old_sid.startswith('manual_'):
                    jd['connected'] = False
                    jd['kicked'] = True
                    try:
                        get_web_server().socketio.emit('force_disconnect', {}, room=old_sid)  # type: ignore
                    except:
                        pass
                    released = True
            if released:
                messagebox.showinfo("成功", f"裁判 {jid_str} 的連線已手動釋放，現在可以更換設備重新登入。", parent=self.root)
                self.update_live_scores()
                top.destroy()
            else:
                messagebox.showinfo("提示", f"裁判 {jid_str} 目前無在線 Socket 連線，不需釋放。", parent=self.root)

        btn_release = tk.Button(bottom_frame, text="釋放連線", bg="#e67e22", fg="white", font=("Microsoft JhengHei", 11, "bold"), width=12, command=release_connection, takefocus=False)
        btn_release.pack(side="left", padx=10)

        def submit():
            acc = var_acc.get()
            pres = round(var_pres1.get() + var_pres2.get() + var_pres3.get(), 1)
            total = round(acc + pres, 1)
            jid_str = f"J{judge_num}"
            
            score_entry = {
                'id': jid_str, 'submitted': True
            }
            
            if is_pk_simul:
                hong_acc = var_hong_acc.get()
                hong_pres = round(var_hong_pres1.get() + var_hong_pres2.get() + var_hong_pres3.get(), 1)
                hong_total = round(hong_acc + hong_pres, 1)
                score_entry.update({
                    'acc': acc, 'pres': pres, 
                    'p1': var_pres1.get(), 'p2': var_pres2.get(), 'p3': var_pres3.get(),
                    'total': total,
                    'hong_acc': hong_acc,
                    'hong_pres': hong_pres,
                    'hong_p1': var_hong_pres1.get(),
                    'hong_p2': var_hong_pres2.get(),
                    'hong_p3': var_hong_pres3.get(),
                    'hong_total': hong_total,
                    'pk_mode': 'simultaneous',
                    'chung_submitted': True,
                    'hong_submitted': True
                })
            elif is_pk_seq and current_side == 1:
                # 只有紅方，將數值存入 hong_ 系列，保留原有的青方成績
                score_entry.update({
                    'hong_acc': acc, 'hong_pres': pres, 
                    'hong_p1': var_pres1.get(), 'hong_p2': var_pres2.get(), 'hong_p3': var_pres3.get(),
                    'hong_total': total,
                    'pk_mode': 'sequence',
                    'hong_submitted': True,
                    'chung_submitted': found_judge_data.get('chung_submitted', False) if found_judge_data else False
                })
                if found_judge_data:
                    score_entry['acc'] = found_judge_data.get('acc', 0.0)
                    score_entry['pres'] = found_judge_data.get('pres', 0.0)
                    score_entry['p1'] = found_judge_data.get('p1', 0.0)
                    score_entry['p2'] = found_judge_data.get('p2', 0.0)
                    score_entry['p3'] = found_judge_data.get('p3', 0.0)
                    score_entry['total'] = found_judge_data.get('total', 0.0)
            else:
                # 只有青方（或一般模式），將數值存入一般系列，保留原有的紅方成績
                score_entry.update({
                    'acc': acc, 'pres': pres, 
                    'p1': var_pres1.get(), 'p2': var_pres2.get(), 'p3': var_pres3.get(),
                    'total': total,
                    'pk_mode': 'sequence' if is_pk_seq else 'normal',
                    'chung_submitted': True,
                    'hong_submitted': found_judge_data.get('hong_submitted', False) if found_judge_data else False
                })
                if found_judge_data and 'hong_acc' in found_judge_data:
                    score_entry['hong_acc'] = found_judge_data.get('hong_acc', 0.0)
                    score_entry['hong_pres'] = found_judge_data.get('hong_pres', 0.0)
                    score_entry['hong_p1'] = found_judge_data.get('hong_p1', 0.0)
                    score_entry['hong_p2'] = found_judge_data.get('hong_p2', 0.0)
                    score_entry['hong_p3'] = found_judge_data.get('hong_p3', 0.0)
                    score_entry['hong_total'] = found_judge_data.get('hong_total', 0.0)
                
            current_state['judges'][fake_sid] = score_entry
            get_web_server().check_all_submitted()
            top.destroy()
            
        btn_cancel = tk.Button(bottom_frame, text="取消 (Esc)", bg="#95a5a6", fg="white", font=("Microsoft JhengHei", 11, "bold"), width=12, command=top.destroy, takefocus=False)
        btn_cancel.pack(side="right", padx=10)
        
        btn_submit = tk.Button(bottom_frame, text="確認送出 (Enter)", bg="#2ecc71", fg="white", font=("Microsoft JhengHei", 11, "bold"), width=16, command=submit, takefocus=False)
        btn_submit.pack(side="right", padx=10)
        
        top.bind("<Return>", lambda e: submit())
        top.bind("<Escape>", lambda e: top.destroy())
        
        def init_focus():
            entry_acc.focus_set()
            entry_acc.select_range(0, tk.END)
            entry_acc.icursor(tk.END)
        top.after(100, init_focus)

    def update_live_scores(self):
        judge_count = int(system_settings["judge_count"])
        
        # 收集當前已連線的裁判 ID 列表
        connected_judges = set()
        for sid, jdata in current_state['judges'].items():
            jid = jdata.get('id', '')
            if jid and (jdata.get('connected', False) or sid.startswith('manual_')):
                connected_judges.add(jid)
                
        for r in range(7):
            for col_idx in [0, 1, 2, 4, 5, 6]:
                self.score_cells[r][col_idx].config(text="")
            if r < judge_count:
                jid_str = f"J{r+1}"
                manual_jid_str = f"manual_J{r+1}"
                is_connected = (jid_str in connected_judges or manual_jid_str in connected_judges)
                fg_color = "#39ff14" if is_connected else "white"  # 已連線顯示亮綠色
                self.score_cells[r][3].config(bg=self.colors["judge_blue"], fg=fg_color)
                
        # 彙整裁判資料，解決 manual 覆蓋實體連線問題
        active_judges_map = {}
        for sid, data in current_state['judges'].items():
            jid_str = data.get('id', '')
            if not jid_str: continue
            if jid_str not in active_judges_map:
                active_judges_map[jid_str] = data
            else:
                existing = active_judges_map[jid_str]
                # 優先採用手動輸入，或者已經提交的資料
                if sid.startswith('manual_') or (data.get('submitted', False) and not existing.get('submitted', False)):
                    active_judges_map[jid_str] = data
                    
        is_pk = (self.mode_var.get() == 1)
        pk_seq = int(config.system_settings.get('pk_sequence_mode', 0))
        is_pk_seq = is_pk and (pk_seq == 1 or pk_seq == 2)
        current_side = config.current_state.get('current_player_side', 0)

        # 改進 submitted_count 的計算，在 PK 交叉/依序上場時，根據當前正在打分方的提交數
        if is_pk_seq:
            if current_side == 1:
                submitted_count = sum(1 for j in active_judges_map.values() if j.get('hong_submitted'))
            else:
                submitted_count = sum(1 for j in active_judges_map.values() if j.get('chung_submitted'))
        else:
            submitted_count = sum(1 for j in active_judges_map.values() if j.get('submitted'))

        if submitted_count == 0:
            for key in self.center_stats_labels:
                self.center_stats_labels[key].config(text="")
        
        for idx in range(judge_count):
            try:
                jid_key = f"J{idx+1}"
                manual_jid_key = f"manual_J{idx+1}"
                is_connected = (jid_key in connected_judges or manual_jid_key in connected_judges)
                fg_color = "#39ff14" if is_connected else "white"
                
                data = active_judges_map.get(jid_key)
                if data:
                    show_chung = False
                    show_hong = False
                    is_done = False
                    
                    if is_pk_seq:
                        show_chung = data.get('chung_submitted', False)
                        show_hong = data.get('hong_submitted', False)
                        if current_side == 1:
                            is_done = data.get('hong_submitted', False)
                        else:
                            is_done = data.get('chung_submitted', False)
                    else:
                        show_chung = data.get('submitted', False)
                        show_hong = data.get('submitted', False)
                        is_done = data.get('submitted', False)
                    
                    # 青方顯示
                    if show_chung:
                        self.score_cells[idx][0].config(text=f"{data.get('acc', 0.0):.1f}")
                        self.score_cells[idx][1].config(text=f"{data.get('pres', 0.0):.1f}")
                        self.score_cells[idx][2].config(text=f"{data.get('total', 0.0):.2f}")
                    else:
                        self.score_cells[idx][0].config(text="")
                        self.score_cells[idx][1].config(text="")
                        self.score_cells[idx][2].config(text="")
                        
                    # 紅方顯示 (PK 賽制)
                    if is_pk:
                        if show_hong:
                            self.score_cells[idx][4].config(text=f"{data.get('hong_total', 0.0):.2f}")
                            self.score_cells[idx][5].config(text=f"{data.get('hong_pres', 0.0):.1f}")
                            self.score_cells[idx][6].config(text=f"{data.get('hong_acc', 0.0):.1f}")
                        else:
                            self.score_cells[idx][4].config(text="")
                            self.score_cells[idx][5].config(text="")
                            self.score_cells[idx][6].config(text="")
                    else:
                        self.score_cells[idx][4].config(text="")
                        self.score_cells[idx][5].config(text="")
                        self.score_cells[idx][6].config(text="")
                        
                    if is_done:
                        self.score_cells[idx][3].config(bg=self.colors["judge_done"], fg="#006600")
                    else:
                        self.score_cells[idx][3].config(bg=self.colors["judge_blue"], fg=fg_color)
                else:
                    self.score_cells[idx][3].config(bg=self.colors["judge_blue"], fg=fg_color)
            except: pass
            
        if self.proj_window and self.proj_window.winfo_exists():
            self.proj_window.refresh()

    # === 關鍵修改: 扣分調整邏輯 ===
    def update_deduction(self, val):
        try:
            curr_text = self.lbl_deduction_val.cget("text")
            curr = float(curr_text) if curr_text else 0.0
            new_val = round(curr + val, 1)
            if new_val < 0: new_val = 0.0
            self.lbl_deduction_val.config(text=str(new_val))
            
            # 使用暫存的計算結果，若有值則重新計算最終分
            if self.temp_avg_acc > 0 or self.temp_avg_pres > 0:
                final_score = self.temp_avg_acc + self.temp_avg_pres - new_val
                final_score = round(final_score, 3)
                
                # 更新中間大格
                self.update_final_score(f"{final_score:.3f}")
                
                # 更新中間下方 Avg Row (Average 的 Final 欄位)
                self.center_stats_labels["Avg_L_2"].config(text=f"{final_score:.3f}")
                
                # 更新左側面板
                self.update_left_panel_scores(
                    self.temp_avg_acc, 
                    self.temp_avg_pres, 
                    new_val, 
                    final_score, 
                    self.temp_raw_sum
                )
            
            # 同步更新 SQLite 資料庫中的 deduction
            if self.current_match_uuid:
                try:
                    import sqlite3
                    conn = sqlite3.connect(database.get_db_path())
                    c = conn.cursor()
                    c.execute("""
                        UPDATE scores 
                        SET deduction = ? 
                        WHERE match_uuid = ? AND round = ?
                    """, (new_val, self.current_match_uuid, self.current_stage))
                    conn.commit()
                    conn.close()
                except Exception as db_err:
                    print(f"Error updating deduction in SQLite: {db_err}")
        except: pass

    def update_left_panel_scores(self, avg_acc, avg_pres, deduction, final_avg_score, raw_sum):
        col_idx = 0 if self.current_stage == 1 else 1
        self.left_labels[0][col_idx].config(text=f"{avg_acc:.2f}")
        self.left_labels[1][col_idx].config(text=f"{avg_pres:.2f}")
        self.left_labels[2][col_idx].config(text=f"{deduction:.1f}")
        self.left_labels[3][col_idx].config(text=f"{(avg_acc + avg_pres - deduction):.2f}")
        if self.current_stage == 1:
            self.score_1r_avg = final_avg_score
            self.score_1r_raw = raw_sum
            display_avg = final_avg_score
            display_total = raw_sum
        else:
            self.score_2r_avg = final_avg_score
            self.score_2r_raw = raw_sum
            display_avg = (self.score_1r_avg + self.score_2r_avg) / 2
            display_total = self.score_1r_raw + self.score_2r_raw
        self.left_merged_labels[4].config(text=f"{display_avg:.3f}")
        self.left_merged_labels[5].config(text=f"{display_total:.1f}")

    def clear_left_panel_scores(self, stage=None):
        if stage is None:
            for row in range(4): 
                self.left_labels[row][0].config(text="") 
                self.left_labels[row][1].config(text="")
            self.left_merged_labels[4].config(text="")
            self.left_merged_labels[5].config(text="")
            self.score_1r_avg = 0.0; self.score_2r_avg = 0.0
            self.score_1r_raw = 0.0; self.score_2r_raw = 0.0
            self.temp_avg_acc = 0.0; self.temp_avg_pres = 0.0; self.temp_raw_sum = 0.0
            
            # 清空右側面板
            if hasattr(self, 'right_labels') and self.right_labels:
                for row in range(4):
                    self.right_labels[row][0].config(text="")
                    self.right_labels[row][1].config(text="")
                self.right_merged_labels[4].config(text="")
                self.right_merged_labels[5].config(text="")
                self.score_1r_avg_R = 0.0; self.score_2r_avg_R = 0.0
                self.score_1r_raw_R = 0.0; self.score_2r_raw_R = 0.0
        elif stage == 1:
            for row in range(4): 
                self.left_labels[row][0].config(text="")
            self.score_1r_avg = 0.0
            self.score_1r_raw = 0.0
            if self.score_2r_avg > 0:
                self.left_merged_labels[4].config(text=f"{self.score_2r_avg:.3f}")
                self.left_merged_labels[5].config(text=f"{self.score_2r_raw:.1f}")
            else:
                self.left_merged_labels[4].config(text="")
                self.left_merged_labels[5].config(text="")
                
            # 清空右側面板 1R
            if hasattr(self, 'right_labels') and self.right_labels:
                for row in range(4):
                    self.right_labels[row][0].config(text="")
                self.score_1r_avg_R = 0.0
                self.score_1r_raw_R = 0.0
                s2_avg = getattr(self, 'score_2r_avg_R', 0.0) or 0.0
                s2_raw = getattr(self, 'score_2r_raw_R', 0.0) or 0.0
                if s2_avg > 0:
                    self.right_merged_labels[4].config(text=f"{s2_avg:.3f}")
                    self.right_merged_labels[5].config(text=f"{s2_raw:.1f}")
                else:
                    self.right_merged_labels[4].config(text="")
                    self.right_merged_labels[5].config(text="")
        elif stage == 2:
            for row in range(4): 
                self.left_labels[row][1].config(text="")
            self.score_2r_avg = 0.0
            self.score_2r_raw = 0.0
            if self.score_1r_avg > 0:
                self.left_merged_labels[4].config(text=f"{self.score_1r_avg:.3f}")
                self.left_merged_labels[5].config(text=f"{self.score_1r_raw:.1f}")
            else:
                self.left_merged_labels[4].config(text="")
                self.left_merged_labels[5].config(text="")
                
            # 清空右側面板 2R
            if hasattr(self, 'right_labels') and self.right_labels:
                for row in range(4):
                    self.right_labels[row][1].config(text="")
                self.score_2r_avg_R = 0.0
                self.score_2r_raw_R = 0.0
                s1_avg = getattr(self, 'score_1r_avg_R', 0.0) or 0.0
                s1_raw = getattr(self, 'score_1r_raw_R', 0.0) or 0.0
                if s1_avg > 0:
                    self.right_merged_labels[4].config(text=f"{s1_avg:.3f}")
                    self.right_merged_labels[5].config(text=f"{s1_raw:.1f}")
                else:
                    self.right_merged_labels[4].config(text="")
                    self.right_merged_labels[5].config(text="")

    def update_right_panel_scores(self, avg_acc, avg_pres, deduction, final_avg_score, raw_sum):
        col_idx = 0 if self.current_stage == 1 else 1
        if hasattr(self, 'right_labels') and self.right_labels:
            self.right_labels[0][col_idx].config(text=f"{avg_acc:.2f}")
            self.right_labels[1][col_idx].config(text=f"{avg_pres:.2f}")
            self.right_labels[2][col_idx].config(text=f"{deduction:.1f}")
            self.right_labels[3][col_idx].config(text=f"{(avg_acc + avg_pres - deduction):.2f}")
            if self.current_stage == 1:
                self.score_1r_avg_R = final_avg_score
                self.score_1r_raw_R = raw_sum
                display_avg = final_avg_score
                display_total = raw_sum
            else:
                self.score_2r_avg_R = final_avg_score
                self.score_2r_raw_R = raw_sum
                display_avg = ((getattr(self, 'score_1r_avg_R', 0.0) or 0.0) + self.score_2r_avg_R) / 2
                display_total = (getattr(self, 'score_1r_raw_R', 0.0) or 0.0) + self.score_2r_raw_R
            self.right_merged_labels[4].config(text=f"{display_avg:.3f}")
            self.right_merged_labels[5].config(text=f"{display_total:.1f}")

    def reset_judges_display(self):
        # 移除手動輸入產生的裁判 Key (以 manual_ 開頭)
        manual_keys = [k for k in current_state['judges'].keys() if k.startswith("manual_")]
        for k in manual_keys:
            current_state['judges'].pop(k, None)
            
        # 同步移除已經斷線的 socket 裁判
        disconnected_keys = [k for k, j in current_state['judges'].items() if not j.get('connected', True)]
        for k in disconnected_keys:
            current_state['judges'].pop(k, None)
            
        # 取得當前賽事模式與預設分數
        mode = self.mode_var.get() if hasattr(self, 'mode_var') else 0
        default_acc = 4.0 if mode != 2 else 6.0
        default_pres = 6.0 if mode != 2 else 4.0
        
        for j in current_state['judges'].values():
            j['submitted'] = False
            j['chung_submitted'] = False
            j['hong_submitted'] = False
            j['acc'] = default_acc
            j['pres'] = default_pres
            j['p1'] = 2.0
            j['p2'] = 2.0
            j['p3'] = 2.0
            j['total'] = default_acc + default_pres
            j['hong_acc'] = default_acc
            j['hong_pres'] = default_pres
            j['hong_p1'] = 2.0
            j['hong_p2'] = 2.0
            j['hong_p3'] = 2.0
            j['hong_total'] = default_acc + default_pres
            j['freestyle_scores'] = {}
            
        self.update_live_scores()

    # === 修改: 儲存設定邏輯 ===
    def open_settings(self):
        gui_dialogs.open_settings(self)

    def update_qr_code(self, url):
        self.cloudflare_url = url
        self.update_court_label()

    def show_qr_popup(self, event=None):
        if hasattr(self, 'qr_popup_window') and self.qr_popup_window and self.qr_popup_window.winfo_exists():
            self.qr_popup_window.lift()
            self.qr_popup_window.focus_force()
            return
        self.qr_popup_window = gui_dialogs.show_qr_popup(self, event)

    def _open_match_editor(self, match_id=None):
        gui_dialogs.open_match_editor(self, match_id)

    def create_match(self): self._open_match_editor()
    def update_match(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("警告", "請先選擇一場比賽", parent=self.root)
            return
        iid = selected[0]
        self._open_match_editor(iid)
    def delete_match(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showwarning("警告", "請先選擇一場比賽", parent=self.root)
            return
        if messagebox.askyesno("確認刪除", "確定要刪除這場比賽資料嗎？", parent=self.root):
            iid = selected[0]
            if iid in self.imported_matches:
                if self.current_match_data and self.imported_matches[iid] == self.current_match_data:
                    self.end_match()
                del self.imported_matches[iid]
                self.update_tree_columns()
    def release_all_connections(self):
        if not messagebox.askyesno("確認", "是否確定要釋放所有裁判的連線？這將會強制踢除所有目前已連線的設備並退回登入畫面。", parent=self.root):
            return
        released_count = 0
        for old_sid, jd in list(current_state['judges'].items()):
            if not old_sid.startswith('manual_'):
                jd['connected'] = False
                try:
                    get_web_server().socketio.emit('force_disconnect', {}, room=old_sid)  # type: ignore
                except:
                    pass
                released_count += 1
        messagebox.showinfo("成功", f"已成功釋放 {released_count} 個裁判設備連線。", parent=self.root)
        self.update_live_scores()
    def update_excel_dir_label(self, path):
        if not hasattr(self, 'lbl_excel_dir'): return
        if path:
            self.lbl_excel_dir.config(text=f"📁 已載入：{path}", fg="#27ae60")
        else:
            self.lbl_excel_dir.config(text="📁 未設定資料夾", fg="#7f8c8d")

    def select_folder(self):
        folder_path = filedialog.askdirectory()
        if not folder_path: return
        self.watch_directory = folder_path
        
        # 記憶路徑並儲存設定
        system_settings["last_excel_directory"] = folder_path
        config.save_settings()
        self.update_excel_dir_label(folder_path)
        
        self.imported_matches.clear()
        self.scan_folder(folder_path)
        self.update_tree_columns()
        if Observer:
            if self.observer:
                self.observer.stop()
                self.observer.join()
            event_handler = MatchFileHandler(self)
            self.observer = Observer()
            self.observer.schedule(event_handler, folder_path, recursive=True)
            self.observer.start()
        messagebox.showinfo("監控啟動", f"已讀取並開始監控資料夾：\n{folder_path}", parent=self.root)
    def scan_folder(self, root_folder):
        for root, dirs, files in os.walk(root_folder):
            for file in files:
                if file.endswith(('.xlsx', '.xls')) and not file.startswith('~$'):
                    full_path = os.path.join(root, file)
                    self.process_excel_file(full_path)
        self.update_session_combobox()
    def process_excel_file(self, file_path):
        """讀取單一 Excel 並更新資料"""
        try:
            filename = os.path.basename(file_path)
            file_id = os.path.splitext(filename)[0]
            
            df = pd.read_excel(file_path)
            
            for index, row in df.iterrows():
                unique_id = f"{file_id}_{index}"
                
                game_val = row.get("Game", 1)
                try: mode_idx = int(game_val) - 1
                except: mode_idx = 0
                
                match_data = {
                    "SourceFile": file_id,
                    "Status": "Ready",  # === 新增: 預設狀態 ===
                    "Game": mode_idx,
                    "Court": row.get("Court", ""),
                    "No": row.get("No", ""),
                    "Round": row.get("Round", 2),
                    "Type": row.get("Type", "Individual"),
                    "Category": row.get("Category", ""),
                    "Division": row.get("Division", ""),
                    "Phase": row.get("Phase", ""),
                    "C_Name": row.get("C Name", ""),
                    "C_NOC": row.get("C NOC", ""),
                    "C_Team": row.get("C Team Name", ""),
                    "H_Name": row.get("H Name", ""),
                    "H_NOC": row.get("H NOC", ""),
                    "H_Team": row.get("H Team Name", "")
                }
                
                for k, v in match_data.items():
                    if pd.isna(v): match_data[k] = ""
                
                # 如果該 ID 已存在且狀態已經是 End，則保留 End (防止 Excel 重讀時覆蓋狀態)
                # 這是進階防護，如果您希望重讀 Excel 就重置狀態，可以拿掉這段 if
                if unique_id in self.imported_matches:
                    old_status = self.imported_matches[unique_id].get("Status", "Ready")
                    if old_status == "End":
                        match_data["Status"] = "End"

                self.imported_matches[unique_id] = match_data
                
                # === 新增: 重開軟體時，從 SQLite 資料庫中恢復已完賽選手的 Status 與 final_score ===
                try:
                    import sqlite3
                    conn = sqlite3.connect(database.get_db_path())
                    c = conn.cursor()
                    c.execute("""
                        SELECT DISTINCT round FROM scores WHERE match_uuid = ?
                    """, (unique_id,))
                    completed_rounds = [r[0] for r in c.fetchall()]
                    conn.close()
                    
                    if completed_rounds:
                        try:
                            total_rounds = int(match_data.get("Round", 2))
                        except:
                            total_rounds = 2
                        
                        # 只要已完成的輪數包含最後一輪 (或是大於等於需要的輪數)，就視為完賽
                        if len(completed_rounds) >= total_rounds or total_rounds in completed_rounds:
                            match_data["Status"] = "End"
                            # 自動載入完賽分數至快取，以便排行榜與 LOG 即時查詢
                            self.get_final_score(unique_id, match_data)
                except Exception as db_err:
                    print(f"Error restoring match status from SQLite for {unique_id}: {db_err}")
                
        except Exception as e:
            print(f"Error reading {file_path}: {e}")

    def handle_file_change(self, event_type, file_path):
        """處理 Watchdog 事件"""
        filename = os.path.basename(file_path)
        if filename.startswith('~$'): return 
        
        file_id = os.path.splitext(filename)[0]
        
        # 如果此檔案正在由本程式寫回，跳過本次 Watchdog 觸發，避免重讀覆蓋修改
        if file_id in self._writing_files:
            return
        
        # 1. 先清除該檔案舊的所有資料
        # 找出所有 SourceFile 是這個檔案的 key
        keys_to_remove = [k for k, v in self.imported_matches.items() if v.get("SourceFile") == file_id]
        for k in keys_to_remove:
            del self.imported_matches[k]
            
        # 2. 如果不是刪除事件，則重新讀取
        if event_type != 'deleted':
            time.sleep(0.1)
            self.process_excel_file(file_path)
            
        self.update_session_combobox()
        
        # 如果目前選取的就是被修改的檔案，則刷新 Treeview
        if self.cb_session_select.get() == file_id:
            self.update_tree_columns()

    def write_match_back_to_excel(self, match_id, match_data):
        """
        將 match_data 回寫到原始 Excel 檔案的對應列。
        match_id 格式必須為 '{file_id}_{row_index}'（從 Excel 匯入的資料才符合此格式）。
        手動建立的比賽（SourceFile == '手動建立'）或無法解析的 ID 會直接略過。
        支援 .xlsx（openpyxl）與 .xls（pandas + xlwt）兩種格式。
        """
        source_file = match_data.get("SourceFile", "")
        if not source_file or source_file == "手動建立":
            return  # 非 Excel 匯入的資料，不回寫
        
        if not self.watch_directory:
            return  # 尚未載入任何資料夾，無法定位檔案
        
        # 解析 row_index：match_id 格式為 {file_id}_{row_index}
        # file_id 本身可能含有底線，所以從最後一個底線切割
        last_underscore = match_id.rfind('_')
        if last_underscore == -1:
            print(f"[Excel回寫] 無法解析 match_id: {match_id}")
            return
        try:
            row_index = int(match_id[last_underscore + 1:])
        except ValueError:
            print(f"[Excel回寫] match_id 末段非整數: {match_id}")
            return
        
        # 在 watch_directory 底下搜尋對應的 Excel 檔案
        excel_path = None
        file_ext = None
        for root, dirs, files in os.walk(self.watch_directory):
            for fname in files:
                if fname.startswith('~$'):
                    continue
                name_no_ext = os.path.splitext(fname)[0]
                ext = os.path.splitext(fname)[1].lower()
                if name_no_ext == source_file and ext in ('.xlsx', '.xls'):
                    excel_path = os.path.join(root, fname)
                    file_ext = ext
                    break
            if excel_path:
                break
        
        if not excel_path:
            print(f"[Excel回寫] 找不到檔案: {source_file}")
            return
        
        # Excel 欄名 -> match_data 欄位的對應 (寫回方向)
        FIELD_TO_COL = {
            "Court":        "Court",
            "No":           "No",
            "Round":        "Round",
            "Type":         "Type",
            "Category":     "Category",
            "Division":     "Division",
            "Phase":        "Phase",
            "C_Name":       "C Name",
            "C_NOC":        "C NOC",
            "C_Team":       "C Team Name",
            "H_Name":       "H Name",
            "H_NOC":        "H NOC",
            "H_Team":       "H Team Name",
        }
        
        # 加入黑名單，防止 Watchdog 重觸發
        self._writing_files.add(source_file)

        try:
            if file_ext == '.xlsx':
                # ── .xlsx：用 openpyxl 精準更新單一列（保留格式）──
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active

                # 第 1 列為標題列；row_index 為 0-base，Excel 列號 = row_index + 2
                excel_row = row_index + 2

                # 建立「欄名 -> 欄號」對應表
                header_map = {}
                for col_idx, cell in enumerate(ws[1], start=1):
                    if cell.value is not None:
                        header_map[str(cell.value).strip()] = col_idx

                # Game 欄特殊：app 為 0-base，Excel 為 1-base
                if "Game" in header_map:
                    try:
                        ws.cell(row=excel_row, column=header_map["Game"]).value = int(match_data.get("Game", 0)) + 1
                    except Exception:
                        pass

                for field_key, excel_col_name in FIELD_TO_COL.items():
                    if excel_col_name in header_map:
                        raw_val = match_data.get(field_key, "")
                        ws.cell(row=excel_row, column=header_map[excel_col_name]).value = raw_val if raw_val != "" else None

                wb.save(excel_path)

            else:
                # ── .xls：用 xlrd 讀取、xlwt 直接寫回（不經過 pandas writer）──
                import xlrd
                import xlwt

                rb = xlrd.open_workbook(excel_path)
                rs = rb.sheet_by_index(0)

                # 目標資料列（xlrd 列索引 = row_index + 1，因第 0 列是標題）
                target_xlrd_row = row_index + 1
                if target_xlrd_row >= rs.nrows:
                    print("[Excel回寫] row_index " + str(row_index) + " 超出範圍（共 " + str(rs.nrows - 1) + " 筆資料）")
                    return

                # 建立欄名 -> match_data 欄位的反向查詢
                COL_TO_FIELD = {v: k for k, v in FIELD_TO_COL.items()}

                # 建立新的 xlwt 工作簿並複製內容，對目標列套用修改
                wb_out = xlwt.Workbook(encoding='utf-8')
                ws_out = wb_out.add_sheet(rs.name)

                for r in range(rs.nrows):
                    for c in range(rs.ncols):
                        if r == target_xlrd_row:
                            col_name = str(rs.cell_value(0, c)).strip() if rs.cell_value(0, c) else ""
                            if col_name == "Game":
                                try:
                                    ws_out.write(r, c, int(match_data.get("Game", 0)) + 1)
                                except Exception:
                                    ws_out.write(r, c, rs.cell_value(r, c))
                            elif col_name in COL_TO_FIELD:
                                val = match_data.get(COL_TO_FIELD[col_name], "")
                                ws_out.write(r, c, val)
                            else:
                                ws_out.write(r, c, rs.cell_value(r, c))
                        else:
                            ws_out.write(r, c, rs.cell_value(r, c))

                wb_out.save(excel_path)

            print(f"[Excel回寫] 成功更新: {excel_path} (行 {row_index})")

        except PermissionError:
            from tkinter import messagebox
            messagebox.showerror(
                "寫入失敗",
                f"無法寫入 Excel 檔案：\n{excel_path}\n\n請確認該檔案未被其他程式（如 Microsoft Excel）開啟後再試。"
            )
        except Exception as e:
            print(f"[Excel回寫] 發生錯誤: {e}")
        finally:
            # 延遲移除黑名單，確保 Watchdog 的事件已被觸發並忽略
            def _remove_lock():
                import time as _time
                _time.sleep(0.5)
                self._writing_files.discard(source_file)
            import threading
            threading.Thread(target=_remove_lock, daemon=True).start()



    def append_match_to_excel(self, session_name, match_data):
        """
        将 match_data 作为新列 append 到对应场次的 Excel 档案。
        若找不到 Excel 档案，则在 watch_directory 建立新的 .xls 档案。
        成功时回传 row_index (0-base 资料列索引)，失败回传 None。
        """
        if not self.watch_directory:
            from tkinter import messagebox
            messagebox.showerror(
                "无法储存",
                "尚未载入任何资料夹，无法建立或更新 Excel 档案。\n请先点击「选择资料夹」後再建立新比赛。"
            )
            return None

        EXCEL_HEADERS = [
            "Game", "Court", "No", "Round", "Type", "Category", "Division", "Phase",
            "C Name", "C NOC", "C Team Name", "H Name", "H NOC", "H Team Name"
        ]
        COL_TO_FIELD = {
            "Court": "Court", "No": "No", "Round": "Round", "Type": "Type",
            "Category": "Category", "Division": "Division", "Phase": "Phase",
            "C Name": "C_Name", "C NOC": "C_NOC", "C Team Name": "C_Team",
            "H Name": "H_Name", "H NOC": "H_NOC", "H Team Name": "H_Team",
        }

        def get_cell_value(excel_col_name):
            if excel_col_name == "Game":
                try:
                    return int(match_data.get("Game", 0)) + 1
                except Exception:
                    return 1
            field_key = COL_TO_FIELD.get(excel_col_name, "")
            return match_data.get(field_key, "") if field_key else ""

        # 搜寻现有的 Excel 档案
        excel_path = None
        file_ext = None
        for root, dirs, files in os.walk(self.watch_directory):
            for fname in files:
                if fname.startswith("~$"):
                    continue
                name_no_ext = os.path.splitext(fname)[0]
                ext = os.path.splitext(fname)[1].lower()
                if name_no_ext == session_name and ext in (".xlsx", ".xls"):
                    excel_path = os.path.join(root, fname)
                    file_ext = ext
                    break
            if excel_path:
                break

        # 加入黑名单，防止 Watchdog 侦测到我们自己写入的档案
        self._writing_files.add(session_name)

        try:
            import xlwt

            if not excel_path:
                # 建立新的 .xls 档案
                excel_path = os.path.join(self.watch_directory, session_name + ".xls")
                wb = xlwt.Workbook(encoding="utf-8")
                ws = wb.add_sheet("Sheet1")
                for c, h in enumerate(EXCEL_HEADERS):
                    ws.write(0, c, h)
                for c, h in enumerate(EXCEL_HEADERS):
                    ws.write(1, c, get_cell_value(h))
                wb.save(excel_path)
                print("[Excel新增] 已建立新档案: " + excel_path)
                return 0

            elif file_ext == ".xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
                row_index = ws.max_row - 1
                new_excel_row = ws.max_row + 1
                header_map = {}
                for c, cell in enumerate(ws[1], start=1):
                    if cell.value is not None:
                        header_map[str(cell.value).strip()] = c
                for h in EXCEL_HEADERS:
                    if h in header_map:
                        ws.cell(row=new_excel_row, column=header_map[h]).value = get_cell_value(h)
                wb.save(excel_path)
                print("[Excel新增] 已 append 到 xlsx: " + excel_path + " (row_index=" + str(row_index) + ")")
                return row_index

            else:
                # .xls：读取全部後重写并 append 新列
                import xlrd
                rb = xlrd.open_workbook(excel_path)
                rs = rb.sheet_by_index(0)
                row_index = rs.nrows - 1
                header_map = {}
                for c in range(rs.ncols):
                    val = rs.cell_value(0, c)
                    if val:
                        header_map[str(val).strip()] = c
                wb_out = xlwt.Workbook(encoding="utf-8")
                ws_out = wb_out.add_sheet(rs.name)
                for r in range(rs.nrows):
                    for c in range(rs.ncols):
                        ws_out.write(r, c, rs.cell_value(r, c))
                new_row = rs.nrows
                for h in EXCEL_HEADERS:
                    if h in header_map:
                        ws_out.write(new_row, header_map[h], get_cell_value(h))
                wb_out.save(excel_path)
                print("[Excel新增] 已 append 到 xls: " + excel_path + " (row_index=" + str(row_index) + ")")
                return row_index

        except PermissionError:
            from tkinter import messagebox
            messagebox.showerror(
                "写入失败",
                "无法写入 Excel 档案：\n" + str(excel_path) + "\n\n请确认该档案未被其他程式（如 Microsoft Excel）开启後再试。"
            )
            return None
        except Exception as e:
            print("[Excel新增] 发生错误: " + str(e))
            return None
        finally:
            def _remove_lock():
                import time as _time
                _time.sleep(0.5)
                self._writing_files.discard(session_name)
            import threading
            threading.Thread(target=_remove_lock, daemon=True).start()

    def session_sort_key(self, session_str):
        if not session_str:
            return (0, 0, "")
        parts = session_str.split('-')
        try:
            primary = int(parts[0])
        except ValueError:
            primary = 999999
            
        sub = 0
        if len(parts) > 1:
            try:
                sub = int(parts[1])
            except ValueError:
                sub = 999999
                
        return (primary, sub, session_str)

    def reset_session_combobox_values(self):
        """點擊下拉選單展開時，重設為完整的排序後場次清單"""
        match_ids = list(set(d.get("SourceFile", "") for d in self.imported_matches.values()))
        match_ids = [x for x in match_ids if x]  # 過濾空值
        match_ids.sort(key=self.session_sort_key)
        self.cb_session_select['values'] = match_ids

    def goto_next_session(self):
        """切換到排序上的下一個場次"""
        if self.is_locked:
            messagebox.showwarning("警告", "比賽進行中 (已準備)，無法切換場次", parent=self.root)
            return
            
        # 1. 取得目前所有的場次
        match_ids = list(set(d.get("SourceFile", "") for d in self.imported_matches.values()))
        match_ids = [x for x in match_ids if x]  # 過濾空值
        if not match_ids:
            messagebox.showinfo("提示", "目前沒有載入任何比賽場次", parent=self.root)
            return
            
        # 2. 排序
        match_ids.sort(key=self.session_sort_key)
        
        # 3. 取得目前選取的場次
        current = self.cb_session_select.get().strip()
        
        # 4. 尋找下一個場次
        next_session = None
        if not current:
            # 如果目前沒有選取，預設選取第一個場次
            next_session = match_ids[0]
        else:
            if current in match_ids:
                idx = match_ids.index(current)
                if idx < len(match_ids) - 1:
                    next_session = match_ids[idx + 1]
                else:
                    messagebox.showinfo("提示", "已經是最後一個場次了", parent=self.root)
                    return
            else:
                # 如果輸入的字串不在列表內，尋找第一個場次
                next_session = match_ids[0]
                
        # 5. 設定新場次並更新 UI
        if next_session:
            self.cb_session_select.set(next_session)
            self.update_tree_columns()

    def goto_prev_session(self):
        """切換到排序上的上一個場次"""
        if self.is_locked:
            messagebox.showwarning("警告", "比賽進行中 (已準備)，無法切換場次", parent=self.root)
            return
            
        # 1. 取得目前所有的場次
        match_ids = list(set(d.get("SourceFile", "") for d in self.imported_matches.values()))
        match_ids = [x for x in match_ids if x]  # 過濾空值
        if not match_ids:
            messagebox.showinfo("提示", "目前沒有載入任何比賽場次", parent=self.root)
            return
            
        # 2. 排序
        match_ids.sort(key=self.session_sort_key)
        
        # 3. 取得目前選取的場次
        current = self.cb_session_select.get().strip()
        
        # 4. 尋找上一個場次
        prev_session = None
        if not current:
            # 如果目前沒有選取，預設選取第一個場次
            prev_session = match_ids[0]
        else:
            if current in match_ids:
                idx = match_ids.index(current)
                if idx > 0:
                    prev_session = match_ids[idx - 1]
                else:
                    messagebox.showinfo("提示", "已經是第一個場次了", parent=self.root)
                    return
            else:
                # 如果輸入的字串不在列表內，尋找第一個場次
                prev_session = match_ids[0]
                
        # 5. 設定新場次並更新 UI
        if prev_session:
            self.cb_session_select.set(prev_session)
            self.update_tree_columns()

    def update_session_combobox(self):
        """更新場次下拉選單的內容"""
        match_ids = list(set(d.get("SourceFile", "") for d in self.imported_matches.values()))
        match_ids = [x for x in match_ids if x]  # 過濾空值
        match_ids.sort(key=self.session_sort_key)
        self.cb_session_select['values'] = match_ids
        
        # 這裡不要強制清空使用者正在輸入的文字，只更新選項列表
        # current = self.cb_session_select.get()
        # if current and current not in match_ids:
        #     self.cb_session_select.set('')
        #     self.update_tree_columns()

    def on_session_selected(self, event): self.update_tree_columns()

    def get_row_values(self, data, mode):
        mode_names = ["Cutoff", "PK", "Freestyle", "Fast"]
        try: mode_str = mode_names[int(data["Game"])] if 0 <= int(data["Game"]) < 4 else "Unknown"
        except: mode_str = "Unknown"
        
        # === 修改: 讀取資料中的狀態，若無則預設 Ready ===
        status = data.get("Status", "Ready")
        
        if mode == 1:
            return (data["No"], status, data["Category"], data["Division"], data["Phase"], 
                    data["Type"], mode_str, data["C_Name"], data["C_NOC"], data["C_Team"],
                    data["H_Name"], data["H_NOC"], data["H_Team"])
        else:
            return (data["No"], status, data["C_Name"], data["C_NOC"], 
                    data["Category"], data["Division"], data["Phase"], 
                    data["C_Team"], data["Type"], mode_str)

    def update_tree_columns(self):
        mode = self.mode_var.get()
        # 設定欄位
        if mode == 1: new_cols = ("NO.", "狀態", "組別", "級別", "輪次", "類型", "比賽模式", "青方姓名", "青NOC", "青方單位", "紅方姓名", "紅NOC", "紅單位")
        else: new_cols = ("籤號", "狀態", "姓名", "NOC", "組別", "級別", "輪次", "單位", "類型", "比賽模式")
        
        # 對兩個 tree 都進行 columns 設定與清空
        for t in [self.tree_ready, self.tree_end]:
            t["columns"] = new_cols
            for col in new_cols:
                t.heading(col, text=col)
                w = 80
                if "姓名" in col or "單位" in col: w = 100
                t.column(col, width=w, anchor="center")
            for item in t.get_children(): t.delete(item)
            
        target_session = self.cb_session_select.get()
        if not target_session: return # 未選取則不顯示
        
        # === 篩選與排序選手資料 ===
        ready_matches = []
        end_matches = []
        for iid, data in self.imported_matches.items():
            # 篩選條件: 來源檔案相同 且 比賽模式相同
            if data.get("SourceFile") == target_session and data["Game"] == mode:
                status = data.get("Status", "Ready")
                if status == "End":
                    end_matches.append((iid, data))
                else:
                    ready_matches.append((iid, data))
        
        # 依籤號排序
        def sort_key(x):
            try: return int(x[1]["No"])
            except: return str(x[1]["No"])
            
        try: ready_matches.sort(key=sort_key)
        except: pass
        try: end_matches.sort(key=sort_key)
        except: pass
        
        # 插入各自的 Treeview
        for iid, data in ready_matches:
            self.tree_ready.insert("", "end", iid=iid, values=self.get_row_values(data, mode))
        for iid, data in end_matches:
            self.tree_end.insert("", "end", iid=iid, values=self.get_row_values(data, mode))
            
        self.update_button_states()

    def on_tab_changed(self, event):
        try:
            current_tab = self.notebook.index(self.notebook.select())
            if current_tab == 0:
                self.tree = self.tree_ready
            else:
                self.tree = self.tree_end
        except:
            pass

    def on_match_select(self, event):
        if self.is_locked: return "break"
        tree_widget = event.widget
        selected_items = tree_widget.selection()
        if not selected_items: return
        iid = selected_items[0]
        if iid not in self.imported_matches: return
        data = self.imported_matches.get(iid)
        if data:
            if data.get("Status") == "End":
                confirm = messagebox.askyesno(
                    "選手已完賽",
                    f"選手「{data['C_Name']}」已經完成比賽。\n您確定要再次選取他重新比賽嗎？\n(注意：這將會清除並覆蓋他原有的分數與紀錄！)",
                    parent=self.root
                )
                if not confirm:
                    return "break"
                else:
                    try:
                        database.clear_match_scores(iid)
                    except Exception as db_err:
                        print(f"清除資料庫分數失敗: {db_err}")
                    if hasattr(self, 'temp_scores_to_save') and self.temp_scores_to_save:
                        for r_num in list(self.temp_scores_to_save.keys()):
                            self.temp_scores_to_save[r_num] = [
                                s for s in self.temp_scores_to_save[r_num] if s.get('match_uuid') != iid
                            ]
                    if hasattr(config, 'current_state') and config.current_state.get('temp_scores'):
                        for r_num in list(config.current_state['temp_scores'].keys()):
                            config.current_state['temp_scores'][r_num] = [
                                s for s in config.current_state['temp_scores'][r_num] if s.get('match_uuid') != iid
                            ]
                    self.invalidate_leaderboard_cache(uid=iid)
                    data["Status"] = "Ready"
                    for k in ["final_score", "final_score_0", "final_score_1",
                              "presentation_score", "presentation_score_0", "presentation_score_1",
                              "raw_total_score", "raw_total_score_0", "raw_total_score_1"]:
                        data.pop(k, None)
                    self.update_tree_columns()
                    try:
                        self.export_match_log(auto_open=False)
                    except Exception as log_err:
                        print(f"退回未比賽時自動更新 HTML 失敗: {log_err}")
                    try:
                        self.notebook.select(0)
                    except:
                        pass
            self.current_match_data = data
            self.current_match_uuid = iid
            self.current_stage = 1
            config.current_state['current_player_side'] = 0
            config.current_state['pk_seq_state'] = 0  # 重置依序上場狀態機
            self.update_court_label()
            self.lbl_no.config(text=f"籤號：{data['No']}", font=("Microsoft JhengHei", 28, "bold"))
            self.lbl_type.config(text=data["Type"])
            self.lbl_category.config(text=data["Category"])
            self.lbl_division.config(text=data["Division"])
            self.lbl_phase.config(text=data["Phase"])
            self.lbl_current_round.config(text="1R")
            self.lbl_1r_tag.config(bg="white", fg="blue")
            self.lbl_2r_tag.config(bg="#eee", fg="#999")
            self.lbl_name_1.config(text=data["C_Name"])
            self.lbl_team_1.config(text=data["C_Team"])
            if data["Game"] == 1:
                self.lbl_name_2.config(text=data["H_Name"])
                self.lbl_team_2.config(text=data["H_Team"])
            else:
                self.lbl_name_2.config(text=data["C_Name"])
                self.lbl_team_2.config(text=data["C_Team"])
            try: round_count = int(data["Round"])
            except: round_count = 2
            current_state['current_player'] = data["C_Name"]
            self.lbl_final_L.config(text="")
            self.lbl_final_R.config(text="")
            self.lbl_deduction_val.config(text="")
            self.lbl_deduction_val_R.config(text="")
            for r in range(7):
                self.score_cells[r][0].config(text="")
                self.score_cells[r][1].config(text="")
                self.score_cells[r][2].config(text="")
            self.refresh_judge_slots()
            self.btn_show_score.config(state="disabled", bg="#eee")
            self.last_proj_score_slide = 0
            self.last_proj_slide_finished = False
            self.temp_scores_to_save = {}
            self.update_proj_data("Ready", data["C_Name"], data["C_Team"], "-")
            
            # 依場次與賽事自動帶入型場
            session_name = data.get("SourceFile", "")
            t_name = system_settings.get("tournament_name", "Default")
            saved_poomsaes = system_settings.get("session_poomsae", {}).get(t_name, {}).get(session_name)
            
            if saved_poomsaes and len(saved_poomsaes) >= 2:
                p1_saved, p2_saved = saved_poomsaes[0], saved_poomsaes[1]
                if p1_saved and p1_saved in self.poomsae_list:
                    self.combo_poomsae_1.set(p1_saved)
                else:
                    self.combo_poomsae_1.set("")
                    
                if p2_saved and p2_saved in self.poomsae_list:
                    self.combo_poomsae_2.set(p2_saved)
                else:
                    self.combo_poomsae_2.set("")
            else:
                self.combo_poomsae_1.set("")
                self.combo_poomsae_2.set("")
                
            # 若只需要 1R，將 2R 標示為不需選擇
            if round_count == 1:
                self.combo_poomsae_2.config(state="normal")
                self.combo_poomsae_2.set("--- 不需選擇 ---")
                self.combo_poomsae_2.config(state="disabled")
            else:
                self.combo_poomsae_2.config(state="readonly")
                
            self.update_button_states()
            
            # 自動根據選取場次的賽事類型切換主控台賽制
            game_mode = data.get("Game", 0)
            if hasattr(self, 'mode_var') and self.mode_var.get() != game_mode:
                self.select_mode(game_mode)
    def draw_poomsae(self):
        if not self.current_match_data:
            messagebox.showwarning("警告", "請先選取一場比賽", parent=self.root)
            return
        
        # 1. 判斷需要抽幾個型場
        try: 
            r = int(self.current_match_data["Round"])
        except: 
            r = 2
            
        # 2. 依據抽籤範圍篩選型場
        start_val = system_settings.get("draw_range_start", "")
        end_val = system_settings.get("draw_range_end", "")
        
        start_no = None
        end_no = None
        
        # 尋找起始與結束的編號
        for item in getattr(self, 'poomsae_data_list', []):
            if start_val and item["display"] == start_val:
                start_no = item["no"]
            if end_val and item["display"] == end_val:
                end_no = item["no"]
                
        # 向下相容：若為數字，則嘗試將其轉換為 int
        if start_no is None and start_val:
            try:
                start_no = int(start_val)
            except ValueError:
                pass
        if end_no is None and end_val:
            try:
                end_no = int(end_val)
            except ValueError:
                pass
            
        # 從 self.poomsae_data_list 進行篩選
        pool = []
        for item in getattr(self, 'poomsae_data_list', []):
            item_no = item["no"]
            if start_no is not None and item_no < start_no:
                continue
            if end_no is not None and item_no > end_no:
                continue
            pool.append(item["display"])
            
        # 如果過濾後的 pool 不夠抽，顯示警告
        if len(pool) < r:
            messagebox.showerror(
                "抽籤失敗",
                f"目前設定的抽籤範圍（{start_val or '無限制'} ~ {end_val or '無限制'}）內僅有 {len(pool)} 個型場，不足以進行 {r} 輪抽籤（至少需要 {r} 個）。\n請至「系統設定」調整抽籤範圍！",
                parent=self.root
            )
            return
            
        # 3. 抽籤邏輯
        if r == 2:
            # 使用 random.sample 一次抽出 2 個不重複的元素
            picked = random.sample(pool, 2)
            self.combo_poomsae_1.set(picked[0])
            self.combo_poomsae_2.set(picked[1])
        else:
            # 只需要抽 1 個
            p1 = random.choice(pool)
            self.combo_poomsae_1.set(p1)
            # 若只有一輪，第二個保持現狀或清空皆可，因為 UI 已經 disable 了
        
        # === 更新投影畫面以即時顯示抽完的型場 ===
        if hasattr(self, 'proj_window') and self.proj_window and self.proj_window.winfo_exists():
            self.proj_window.refresh()
            
        # 儲存場次與賽事之型場記憶
        if self.current_match_data:
            session_name = self.current_match_data.get("SourceFile", "")
            if session_name:
                t_name = system_settings.get("tournament_name", "Default")
                if "session_poomsae" not in config.system_settings:
                    config.system_settings["session_poomsae"] = {}
                if t_name not in config.system_settings["session_poomsae"]:
                    config.system_settings["session_poomsae"][t_name] = {}
                config.system_settings["session_poomsae"][t_name][session_name] = [
                    self.combo_poomsae_1.get(),
                    self.combo_poomsae_2.get()
                ]
                config.save_settings()
                
        self.update_button_states()

    def on_poomsae_combobox_changed(self, event=None):
        if self.is_locked:
            p1 = self.combo_poomsae_1.get().strip()
            is_2r_active = str(self.combo_poomsae_2['state']) != 'disabled'
            p2 = self.combo_poomsae_2.get().strip() if is_2r_active else "N/A"
            
            reverted = False
            t_name = system_settings.get("tournament_name", "Default")
            session_name = self.current_match_data.get("SourceFile", "") if self.current_match_data else ""
            saved = config.system_settings.get("session_poomsae", {}).get(t_name, {}).get(session_name, ["", ""])
            
            if not p1:
                messagebox.showwarning("警告", "比賽已準備，型場不允許為空！\n已自動恢復原設定。", parent=self.root)
                if saved[0]: self.combo_poomsae_1.set(saved[0])
                else: self.combo_poomsae_1.current(0)
                reverted = True
                
            if is_2r_active and not p2:
                messagebox.showwarning("警告", "比賽已準備，型場不允許為空！\n已自動恢復原設定。", parent=self.root)
                if len(saved) > 1 and saved[1]: self.combo_poomsae_2.set(saved[1])
                else: self.combo_poomsae_2.current(0)
                reverted = True
                
            if reverted:
                return

        if hasattr(self, 'proj_window') and self.proj_window and self.proj_window.winfo_exists():
            self.proj_window.refresh()
            
        # 儲存場次與賽事之型場記憶
        if self.current_match_data:
            session_name = self.current_match_data.get("SourceFile", "")
            if session_name:
                t_name = system_settings.get("tournament_name", "Default")
                if "session_poomsae" not in config.system_settings:
                    config.system_settings["session_poomsae"] = {}
                if t_name not in config.system_settings["session_poomsae"]:
                    config.system_settings["session_poomsae"][t_name] = {}
                config.system_settings["session_poomsae"][t_name][session_name] = [
                    self.combo_poomsae_1.get(),
                    self.combo_poomsae_2.get()
                ]
                config.save_settings()
                
        self.update_button_states()

    def toggle_ready(self):
        # --- 新增檢查邏輯開始 ---
        
        # 1. 取得目前的選擇
        p1 = self.combo_poomsae_1.get()
        p2 = self.combo_poomsae_2.get()
        
        # 2. 基本檢查：1R 不能為空
        if not p1:
            messagebox.showwarning("警告", "請先設定型場 (1R)", parent=self.root)
            return

        # 3. 進階檢查：如果 2R 是啟用的 (代表這場比賽有兩輪)，則檢查是否重複
        # 判斷 combobox 狀態是否為 'disabled'
        is_2r_active = str(self.combo_poomsae_2['state']) != 'disabled'
        
        if is_2r_active:
            if not p2:
                messagebox.showwarning("警告", "本場賽事需要兩套型場，請設定 2R。", parent=self.root)
                return
            
            if p1 == p2:
                messagebox.showwarning("規則錯誤", f"1R 與 2R 不能相同！\n目前都是：{p1}\n\n請手動更改其中一個。", parent=self.root)
                return
        current_text = self.btn_ready.cget("text")
        if current_text == "準備":
            self.is_locked = True
            self.has_clicked_start = False
            self.btn_ready.config(text="取消")
            self.timer_seconds = system_settings["countdown_sec"]
            self._update_timer_label()
            self.lbl_final_L.config(text="-")
            self.lbl_final_R.config(text="-")
            self.lbl_deduction_val.config(text="0.0")
            self.lbl_deduction_val_R.config(text="")
            name = self.current_match_data["C_Name"] if self.current_match_data else ""
            team = self.current_match_data["C_Team"] if self.current_match_data else ""
            self.update_proj_data("Ready", name, team, "-")
            self.start_scoring()
            self.update_button_states()
        else:
            if getattr(self, 'has_clicked_start', False):
                ans = messagebox.askyesno("確認取消", "比賽已經開始計分，您確定要取消準備狀態嗎？\n這將重設目前的計分並退選選手！", parent=self.root)
                if not ans:
                    return
            self.is_locked = False
            self.stop_timer()
            
            # === 退選並重設當前選手與 UI 元件狀態 ===
            if hasattr(self, 'tree') and self.tree:
                try: self.tree.selection_remove(self.tree.selection())
                except: pass
            
            self.current_match_data = None
            self.current_match_uuid = None
            self.current_stage = 1
            self.update_court_label()
            self.lbl_no.config(text="籤號：", font=("Microsoft JhengHei", 28, "bold"))
            self.lbl_name_1.config(text="")
            self.lbl_name_2.config(text="")
            self.lbl_team_1.config(text="")
            self.lbl_team_2.config(text="")
            self.lbl_final_L.config(text="")
            self.lbl_final_R.config(text="")
            self.lbl_deduction_val.config(text="")
            self.lbl_deduction_val_R.config(text="")
            self.btn_ready.config(text="準備")
            self.btn_start.config(text="開始")
            if hasattr(self, 'btn_next'): self.btn_next.config(state="disabled")
            if hasattr(self, 'btn_end'): self.btn_end.config(state="disabled")
            if hasattr(self, 'btn_show_score'): self.btn_show_score.config(state="disabled", bg="#eee")
            
            for r in range(7):
                self.score_cells[r][0].config(text="")
                self.score_cells[r][1].config(text="")
                self.score_cells[r][2].config(text="")
                
            self.clear_left_panel_scores()
            self.reset_judges_display()
            if hasattr(self, 'combo_poomsae_1'): self.combo_poomsae_1.set("")
            if hasattr(self, 'combo_poomsae_2'): self.combo_poomsae_2.set("")
            
            # === 更新大螢幕投影跳回系統等待載入頁面 (Waiting...) ===
            self.update_proj_data("Waiting...", "", "", "")
            
            # === 清空當前暫存分數與重置裁判平板狀態 ===
            self.temp_scores_to_save = {}
            def trigger_cancel_reset():
                try:
                    import urllib.request
                    urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/reset_match", timeout=2, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
                except Exception as e:
                    with open("error.log", "a", encoding="utf-8") as f:
                        f.write(f"--- API CANCEL RESET ERROR: {e} ---\n")
            threading.Thread(target=trigger_cancel_reset, daemon=True).start()
        self.update_button_states()
    def toggle_start(self):
        if not self.is_locked:
            messagebox.showwarning("警告", "請先按下「準備」", parent=self.root)
            return
        current_text = self.btn_start.cget("text")
        if current_text == "開始":
            self.btn_start.config(text="暫停")
            self.has_clicked_start = True
            if not self.timer_running:
                self.timer_running = True
                if not current_state.get('is_scoring', False):
                    self.start_scoring()
                self._timer_tick()
                self.update_proj_data("Scoring...", self.current_match_data["C_Name"], self.current_match_data["C_Team"])
        else:
            self.stop_timer()
            self.btn_start.config(text="開始")
        self.update_button_states()
    def show_scores(self):
        if not self.current_match_data:
            return
            
        # 自動停止計時器
        if self.timer_running:
            self.stop_timer()
            self.btn_start.config(text="開始")
            
        self.last_proj_score_slide = 0
        self.last_proj_slide_finished = False
        score = ""
        if hasattr(self, 'left_merged_labels') and 4 in self.left_merged_labels:
            score = self.left_merged_labels[4].cget("text")
        if not score or score == "-":
            score = self.lbl_final_L.cget("text")
        if not score or score == "-":
            score = self.lbl_final_R.cget("text")
            
        if self.current_match_uuid and self.current_match_uuid in self.imported_matches:
            try:
                self.imported_matches[self.current_match_uuid]["final_score"] = float(score)
            except:
                pass
        self.update_proj_data("Final Score", self.current_match_data["C_Name"], self.current_match_data["C_Team"], score)
        self.update_button_states()
        
        # 注意：此時 mdata["final_score"] 已在上方由 GUI 標籤的文字設定為正確的分數，
        # 快取是有效且正確的，不應在這裡清除。
        # 快取失效應在 end_match 中、分數確實寫入資料庫之後才進行。
        
        # === 排名計算與 SocketIO 廣播移至背景執行緒，避免阻塞主畫面 ===
        import urllib.parse
        import urllib.request
        _score = score
        _player_name = self.current_match_data.get("C_Name", "") if self.current_match_data else ""
        
        def calc_rank_and_stop():
            rank_val = "-"
            try:
                self.invalidate_leaderboard_cache(uid=self.current_match_uuid)
                leaderboard = self.query_leaderboard_data()
                for item in leaderboard:
                    if item["name"] == _player_name:
                        rank_val = str(item.get("rank", "-"))
                        break
            except Exception as e:
                print(f"Error querying rank in background: {e}")
            
            # 於背景非同步請求本機 Flask 端點，將 SocketIO 廣播派發工作帶回 Flask 執行緒上下文中，確保執行緒安全
            params = urllib.parse.urlencode({
                'final_score': _score,
                'rank': rank_val
            })
            try:
                urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/stop_scoring?{params}", timeout=2, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
            except Exception as e:
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- API STOP TRIGGER ERROR: {e} ---\n")
        
        threading.Thread(target=calc_rank_and_stop, daemon=True).start()
    def next_round(self):
        pk_seq = int(config.system_settings.get('pk_sequence_mode', 0))
        seq_state = config.current_state.get('pk_seq_state', 0)
        
        # 依序上場：展示1R後按「下一品」→ 切換到紅方2R（state 2→3）
        if pk_seq == 2 and seq_state == 2:
            self.advance_seq_mode()  # 推進到紅方2R
            return
        
        self.current_stage = 2
        # 依序上場青方 1R→2R 時，保持青方（player_side=0）並推進狀態
        if pk_seq == 2 and seq_state == 0:
            config.current_state['pk_seq_state'] = 1  # 青方1R → 青方2R
            config.current_state['current_player_side'] = 0
        elif pk_seq != 2:
            config.current_state['current_player_side'] = 0  # 一般/交叉上場：重置為青方
        self.lbl_current_round.config(text="2R")
        self.lbl_1r_tag.config(bg="#eee", fg="#999")
        self.lbl_2r_tag.config(bg="white", fg="blue")
        self.timer_seconds = system_settings["countdown_sec"]
        self._update_timer_label()
        self.lbl_final_L.config(text="-")
        self.lbl_final_R.config(text="-")
        self.lbl_deduction_val.config(text="0.0")
        self.lbl_deduction_val_R.config(text="")
        self.btn_next.config(state="disabled", text="下一品")
        self.btn_start.config(text="開始")
        self.btn_show_score.config(state="disabled", bg="#eee")
        self.stop_timer()
        self.reset_judges_display()
        self.update_proj_data("Round 2 Ready", self.current_match_data["C_Name"], self.current_match_data["C_Team"], "-")
        self.start_scoring()
        self.update_button_states()
    
    def dispatch_next_action(self):
        """依據 btn_next 的目前文字，動態分發到對應的動作"""
        btn_text = self.btn_next.cget('text')
        pk_seq = int(config.system_settings.get('pk_sequence_mode', 0))
        if btn_text == '交換選手':
            if pk_seq == 2:
                self.advance_seq_mode()  # 依序上場：切換紅方
            else:
                self.switch_player_cross_mode()  # 交叉上場
        else:
            self.next_round()
    
    def advance_seq_mode(self):
        """
        依序上場專用狀態推進：
          state 0 (青方1R完成) → state 1: 青方2R
          state 1 (青方2R完成) → state 2: 紅方1R
          state 2 (紅方1R展示後) → state 3: 紅方2R
        """
        seq_state = config.current_state.get('pk_seq_state', 0)
        
        if seq_state == 0:
            # 青方1R → 青方2R（保持青方，進入第2品）
            config.current_state['pk_seq_state'] = 1
            config.current_state['current_player_side'] = 0
            self.current_stage = 2
            self.lbl_current_round.config(text="2R")
            self.lbl_1r_tag.config(bg="#eee", fg="#999")
            self.lbl_2r_tag.config(bg="white", fg="blue")
            self.timer_seconds = system_settings["countdown_sec"]
            self._update_timer_label()
            self.lbl_final_L.config(text="-")
            self.lbl_final_R.config(text="-")
            self.lbl_deduction_val.config(text="0.0")
            self.lbl_deduction_val_R.config(text="")
            self.btn_next.config(state="disabled", text="下一品")
            self.btn_start.config(text="開始")
            self.btn_show_score.config(state="disabled", bg="#eee")
            self.stop_timer()
            self.reset_judges_display()
            self.update_proj_data("Round 2 Ready", self.current_match_data["C_Name"], self.current_match_data["C_Team"], "-")
            self.start_scoring()
        
        elif seq_state == 1:
            # 青方2R → 紅方1R（換邊，切回第1品）
            config.current_state['pk_seq_state'] = 2
            config.current_state['current_player_side'] = 1  # 切換到紅方
            config.current_state['is_scoring'] = True
            self.current_stage = 1
            self.lbl_current_round.config(text="1R")
            self.lbl_1r_tag.config(bg="white", fg="blue")
            self.lbl_2r_tag.config(bg="#eee", fg="#999")
            for sid, jd in config.current_state['judges'].items():
                jd['submitted'] = False
                jd['hong_submitted'] = False
            self.btn_next.config(state="disabled", text="下一品")
            self.btn_show_score.config(state="disabled", bg="#eee")
            self.timer_seconds = system_settings["countdown_sec"]
            self._update_timer_label()
            self.lbl_final_L.config(text="-")
            self.lbl_final_R.config(text="-")
            self.lbl_deduction_val.config(text="0.0")
            self.lbl_deduction_val_R.config(text="")
            self.stop_timer()
            self.reset_judges_display()
            self.update_proj_data("Round 1 Ready", self.current_match_data["H_Name"], self.current_match_data["H_Team"], "-")
            self.update_live_scores()
            self.start_scoring()
        
        elif seq_state == 2:
            # 紅方1R展示後 → 紅方2R（保持紅方，進入第2品）
            config.current_state['pk_seq_state'] = 3
            config.current_state['current_player_side'] = 1  # 仍為紅方
            config.current_state['is_scoring'] = True
            self.current_stage = 2
            self.lbl_current_round.config(text="2R")
            self.lbl_1r_tag.config(bg="#eee", fg="#999")
            self.lbl_2r_tag.config(bg="white", fg="blue")
            for sid, jd in config.current_state['judges'].items():
                jd['submitted'] = False
                jd['hong_submitted'] = False
            self.btn_next.config(state="disabled", text="下一品")
            self.btn_show_score.config(state="disabled", bg="#eee")
            self.timer_seconds = system_settings["countdown_sec"]
            self._update_timer_label()
            self.lbl_final_L.config(text="-")
            self.lbl_final_R.config(text="-")
            self.lbl_deduction_val.config(text="0.0")
            self.lbl_deduction_val_R.config(text="")
            self.stop_timer()
            self.reset_judges_display()
            self.update_proj_data("Round 2 Ready", self.current_match_data["H_Name"], self.current_match_data["H_Team"], "-")
            self.update_live_scores()
            self.start_scoring()
        
        self.update_button_states()
    
    def switch_player_cross_mode(self):
        """
        PK 交叉上場：青方評分完成後，切換到紅方上場評分。
        - 將 current_player_side 設為 1（紅方）
        - 重置裁判評分狀態
        - 重新開始評分廣播（裁判端顯示紅方評分介面）
        """
        config.current_state['current_player_side'] = 1  # 切換到紅方
        config.current_state['is_scoring'] = True
        
        # 重置所有裁判的 submitted 狀態與紅方狀態，保留青方狀態
        for sid, jd in config.current_state['judges'].items():
            jd['submitted'] = False
            jd['hong_submitted'] = False
        
        # 按鈕狀態暫時重置
        self.btn_next.config(state="disabled", text="下一品")
        self.btn_show_score.config(state="disabled", bg="#eee")
        self.timer_seconds = system_settings["countdown_sec"]
        self._update_timer_label()
        self.stop_timer()
        self.update_live_scores()
        
        # 重新開始評分（裁判平板會看到紅方評分介面）
        self.start_scoring()
        self.update_button_states()
    def write_html_log(self):
        try:
            if not self.current_match_uuid or not self.current_match_data:
                return
                
            from datetime import datetime
            import sqlite3
            data = self.current_match_data
            match_uuid = self.current_match_uuid
            
            # 1. 取得時間
            now = datetime.now()
            end_time_str = now.strftime("%Y%m%d%H%M%S")
            log_time_str = now.strftime("%Y/%m/%d %H:%M:%S")
            date_str = now.strftime("%Y_%m_%d")
            os.makedirs("daily_logs", exist_ok=True)
            filename = os.path.join("daily_logs", f"log.1.0.14_{date_str}.html")
            
            # 2. 判斷是否為棄權 (Withdraw)
            # 優先從資料庫讀取完整數據，若無資料（如比賽未結束或棄權）才從 temp_scores 讀取
            rows = []
            conn = sqlite3.connect(database.get_db_path())
            c = conn.cursor()
            try:
                c.execute("""
                    SELECT round, judge_id, accuracy, presentation, total, p1, p2, p3, deduction, timestamp, player_side
                    FROM scores
                    WHERE match_uuid = ?
                    ORDER BY round, judge_id
                """, (match_uuid,))
                rows = c.fetchall()
            except Exception:
                try:
                    # 舊資料庫相容
                    c.execute("""
                        SELECT round, judge_id, accuracy, presentation, total, player_side
                        FROM scores
                        WHERE match_uuid = ?
                        ORDER BY round, judge_id
                    """, (match_uuid,))
                    rows = c.fetchall()
                except Exception:
                    pass
            conn.close()
            
            if not rows:
                temp_scores = config.current_state.get('temp_scores', getattr(self, 'temp_scores_to_save', {}))
                if temp_scores:
                    for r_num, scores_list in temp_scores.items():
                        for s in scores_list:
                            if s.get('match_uuid') == match_uuid:
                                rows.append((
                                    s['round_num'],
                                    s['judge_id'],
                                    s['acc'],
                                    s['pres'],
                                    s['total'],
                                    s.get('p1', 0.0),
                                    s.get('p2', 0.0),
                                    s.get('p3', 0.0),
                                    s.get('deduction', 0.0),
                                    None,
                                    s.get('player_side', 0)
                                ))
            
            is_withdraw = len(rows) == 0
            status_val = data.get("Status", "")
            if status_val == "End":
                result_status = "End"
            elif status_val == "Withdraw":
                result_status = "Withdraw"
            else:
                result_status = "" if is_withdraw else "End"
                
            if is_withdraw or status_val != "End":
                end_time_str = ""
            
            # 3. 欄位資料
            court = str(data.get("Court", ""))
            no = str(data.get("No", ""))
            
            # 處理 Game 可能是 None 或字串/整數
            game_val = data.get("Game", 0)
            try:
                game_method = "Cutoff" if int(game_val) == 0 else "Tournaments"
            except:
                game_method = "Cutoff"
                
            game_type = str(data.get("Type", ""))
            category = str(data.get("Category", ""))
            division = str(data.get("Division", ""))
            phase = str(data.get("Phase", ""))
            
            is_pk = False
            try: is_pk = (int(game_val) == 1)
            except: pass
            
            def get_player_side(r):
                if len(r) > 10: return r[10]
                elif len(r) > 5: return r[5]
                return 0
                
            has_chung_scores = any(get_player_side(r) == 0 for r in rows)
            has_hong_scores = any(get_player_side(r) == 1 for r in rows)
            
            c_noc = str(data.get("C_NOC", "")) if has_chung_scores else ""
            c_team = str(data.get("C_Team", ""))
            c_name = str(data.get("C_Name", ""))
            
            h_noc = str(data.get("H_NOC", "")) if (is_pk and has_hong_scores) else ""
            h_team = str(data.get("H_Team", ""))
            h_name = str(data.get("H_Name", ""))
            
            print(f"[HTML Log Debug] match_uuid={match_uuid}, is_pk={is_pk}, game_val={game_val}, rows_count={len(rows)}")

                
            def calc_avg(scores_list):
                if not scores_list:
                    return 0.0, 0.0, 0.0
                accs = [s[2] for s in scores_list]
                pres = [s[3] for s in scores_list]
                
                def calc_trimmed_avg(val_list):
                    if not val_list: return 0.0
                    if len(val_list) <= 3: return sum(val_list) / len(val_list)
                    val_list.sort()
                    valid = val_list[1:-1]
                    return sum(valid) / len(valid)
                
                avg_acc = calc_trimmed_avg(accs)
                avg_pres = calc_trimmed_avg(pres)
                total_raw = sum(accs) + sum(pres)
                return avg_acc, avg_pres, total_raw

            html_block = f"""<tr>
<td colspan='8' style='height: 5px'></td>
</tr><tr>
<td>{court}</td>
<td>{no}</td>
<td>{game_method}</td>
<td>{game_type}</td>
<td>{category}</td>
<td>{division}</td>
<td>{phase}</td>
<td>{end_time_str}</td>
</tr><tr>
<td>{c_noc}</td>
<td>{c_team}</td>
<td>{c_name}</td>
<td>{h_noc}</td>
<td>{h_team}</td>
<td>{h_name}</td>
<td>{result_status}</td>
<td>__RANK_OR_WINNER__</td>
</tr>"""

            if not is_withdraw:
                # 計算青方分數
                chung_side_rows = [r for r in rows if get_player_side(r) == 0]
                r1_scores_chung = [r for r in chung_side_rows if r[0] == 1]
                r2_scores_chung = [r for r in chung_side_rows if r[0] == 2]
                
                r1_avg_acc, r1_avg_pres, r1_total_raw = calc_avg(r1_scores_chung)
                r2_avg_acc, r2_avg_pres, r2_total_raw = calc_avg(r2_scores_chung)
                
                try:
                    deduction = float(self.lbl_deduction_val.cget("text"))
                except:
                    deduction = 0.0
                
                r1_avg = r1_avg_acc + r1_avg_pres - deduction if r1_scores_chung else 0.0
                r2_avg = r2_avg_acc + r2_avg_pres - deduction if r2_scores_chung else 0.0
                
                if r1_scores_chung and r2_scores_chung:
                    total_avg = (r1_avg + r2_avg) / 2
                    total_raw = r1_total_raw + r2_total_raw
                    avg_acc_chung = (r1_avg_acc + r2_avg_acc) / 2
                    avg_pres_chung = (r1_avg_pres + r2_avg_pres) / 2
                elif r1_scores_chung:
                    total_avg = r1_avg
                    total_raw = r1_total_raw
                    avg_acc_chung = r1_avg_acc
                    avg_pres_chung = r1_avg_pres
                else:
                    total_avg = 0.0
                    total_raw = 0.0
                    avg_acc_chung = 0.0
                    avg_pres_chung = 0.0
                    
                r1_text = f"{r1_avg_acc:.3f} / {r1_avg_pres:.3f} / {deduction:.1f} / {r1_avg:.3f} / {r1_total_raw:.1f}" if r1_scores_chung else ""
                r2_text = f"{r2_avg_acc:.3f} / {r2_avg_pres:.3f} / {deduction:.1f} / {r2_avg:.3f} / {r2_total_raw:.1f}" if r2_scores_chung else ""
                total_text = f"{avg_acc_chung:.3f} / {avg_pres_chung:.3f} / {deduction:.1f} / {total_avg:.3f} / {total_raw:.1f}" if r1_scores_chung else ""
                
                # 初始化紅方欄位
                r1_text_hong = ""
                r2_text_hong = ""
                total_text_hong = ""
                
                print(f"[HTML Log Debug] chung_rows={len(chung_side_rows)}, total_text={total_text}")
                
                if is_pk:
                    hong_side_rows = [r for r in rows if get_player_side(r) == 1]
                    r1_scores_hong = [r for r in hong_side_rows if r[0] == 1]
                    r2_scores_hong = [r for r in hong_side_rows if r[0] == 2]
                    
                    r1_avg_acc_h, r1_avg_pres_h, r1_total_raw_h = calc_avg(r1_scores_hong)
                    r2_avg_acc_h, r2_avg_pres_h, r2_total_raw_h = calc_avg(r2_scores_hong)
                    
                    try:
                        deduction_R = float(self.lbl_deduction_val_R.cget("text")) if hasattr(self, 'lbl_deduction_val_R') else 0.0
                    except:
                        deduction_R = 0.0
                        
                    r1_avg_h = r1_avg_acc_h + r1_avg_pres_h - deduction_R if r1_scores_hong else 0.0
                    r2_avg_h = r2_avg_acc_h + r2_avg_pres_h - deduction_R if r2_scores_hong else 0.0
                    
                    if r1_scores_hong and r2_scores_hong:
                        total_avg_h = (r1_avg_h + r2_avg_h) / 2
                        total_raw_h = r1_total_raw_h + r2_total_raw_h
                        avg_acc_hong = (r1_avg_acc_h + r2_avg_acc_h) / 2
                        avg_pres_hong = (r1_avg_pres_h + r2_avg_pres_h) / 2
                    elif r1_scores_hong:
                        total_avg_h = r1_avg_h
                        total_raw_h = r1_total_raw_h
                        avg_acc_hong = r1_avg_acc_h
                        avg_pres_hong = r1_avg_pres_h
                    else:
                        total_avg_h = 0.0
                        total_raw_h = 0.0
                        avg_acc_hong = 0.0
                        avg_pres_hong = 0.0
                        
                    r1_text_hong = f"{r1_avg_acc_h:.3f} / {r1_avg_pres_h:.3f} / {deduction_R:.1f} / {r1_avg_h:.3f} / {r1_total_raw_h:.1f}" if r1_scores_hong else ""
                    r2_text_hong = f"{r2_avg_acc_h:.3f} / {r2_avg_pres_h:.3f} / {deduction_R:.1f} / {r2_avg_h:.3f} / {r2_total_raw_h:.1f}" if r2_scores_hong else ""
                    total_text_hong = f"{avg_acc_hong:.3f} / {avg_pres_hong:.3f} / {deduction_R:.1f} / {total_avg_h:.3f} / {total_raw_h:.1f}" if r1_scores_hong else ""
                    print(f"[HTML Log Debug] hong_rows={len(hong_side_rows)}, total_text_hong={total_text_hong}")
                
                # 輸出成績資料行 (PK時同時印出青紅雙方成績)
                html_block += f"""<tr>
<td>{r1_text}</td>
<td>{r2_text}</td>
<td style='color: blue; font-weight:bold;'>{total_text}</td>
<td>{r1_text_hong}</td>
<td>{r2_text_hong}</td>
<td style='color: red; font-weight:bold;'>{total_text_hong}</td>
<td></td>
<td></td>
</tr>"""
                
                def format_judge_row(r_scores):
                    cells = []
                    for j_idx in range(1, 8):
                        j_score = None
                        for s in r_scores:
                            jid = s[1]
                            if jid == f"J{j_idx}" or jid == f"manual_J{j_idx}":
                                j_score = s
                                break
                        if j_score:
                            acc_val, pres_val = j_score[2], j_score[3]
                            if len(j_score) > 8 and j_score[5] is not None and (j_score[5] > 0 or j_score[6] > 0 or j_score[7] > 0):
                                p1, p2, p3 = j_score[5], j_score[6], j_score[7]
                            else:
                                p1 = min(2.0, pres_val / 3)
                                p2 = min(2.0, (pres_val - p1) / 2)
                                p3 = max(0.0, pres_val - p1 - p2)
                            cells.append(f"<td>{int(acc_val*10)} / {int(p1*10)} / {int(p2*10)} / {int(p3*10)}</td>")
                        else:
                            cells.append("<td></td>")
                    return "<tr>" + "".join(cells) + "<td></td></tr>"
                    
                # 青方 1R/2R 裁判明細
                html_block += format_judge_row(r1_scores_chung)
                html_block += format_judge_row(r2_scores_chung)
                
                # 紅方 1R/2R 裁判明細 (若為非 PK 則以兩列空行填充以對齊表頭)
                if is_pk:
                    html_block += format_judge_row(r1_scores_hong)
                    html_block += format_judge_row(r2_scores_hong)
                else:
                    html_block += "<tr>" + "<td></td>"*7 + "<td></td></tr>"
                    html_block += "<tr>" + "<td></td>"*7 + "<td></td></tr>"
                
                # 計算排名或勝利者
                rank_or_winner = ""
                if is_pk:
                    def val_eq(a, b):
                        return round(a, 4) == round(b, 4)
                    if total_avg > total_avg_h:
                        rank_or_winner = "BLUE"
                    elif total_avg_h > total_avg:
                        rank_or_winner = "RED"
                    else:
                        if val_eq(avg_pres_chung, avg_pres_hong):
                            if val_eq(total_raw, total_raw_h):
                                rank_or_winner = "DRAW"
                            elif total_raw > total_raw_h:
                                rank_or_winner = "BLUE"
                            else:
                                rank_or_winner = "RED"
                        elif avg_pres_chung > avg_pres_hong:
                            rank_or_winner = "BLUE"
                        else:
                            rank_or_winner = "RED"
                else:
                    # 一般賽制排名
                    rank_val = "-"
                    try:
                        leaderboard = self.query_leaderboard_data(cat=data.get("Category"), div=data.get("Division"), phase=data.get("Phase"))
                        for item in leaderboard:
                            if item["name"] == c_name:
                                rank_val = str(item.get("rank", "-"))
                                break
                    except Exception as e:
                        print(f"Error querying rank in write_html_log: {e}")
                    rank_or_winner = rank_val
                
                html_block = html_block.replace("__RANK_OR_WINNER__", rank_or_winner)
            else:
                html_block = html_block.replace("__RANK_OR_WINNER__", "")
                
            def append_to_file(path):
                header_template = f"<!DOCTYPE html><html><head><meta charset='UTF-8'><title>LOG</title><style>table {{table-layout: fixed; width: 100%; border-collapse: collapse;}} td {{height: 20px; text-align: center; word-wrap: break-word;}}</style></head><body>\n<h1 style='text-align: center; border-top: 2px solid #888; border-bottom: 2px solid #888; height: 50px; line-height: 50px;'> GAME RESULT </h1>\n<p style='text-align: right; '> TIME : {log_time_str}</p>\n<table style='width: 100%;border-spacing: 0px; font-size: 13px;' border='1'>\n<tr>\n<td style='width: 13%;'>Court</td>\n<td style='width: 13%;'>No.</td>\n<td style='width: 13%;'>Game method</td>\n<td style='width: 13%;'>Type</td>\n<td style='width: 13%;'>Category</td>\n<td style='width: 13%;'>Division</td>\n<td style='width: 13%;'>Phase</td>\n<td style='width: 9%;'>End Time</td>\n</tr><tr>\n<td>Noc (Chung)</td>\n<td>Team (Chung)</td>\n<td>Name (Chung)</td>\n<td>Noc (Hong)</td>\n<td>Team (Hong)</td>\n<td>Name (Hong)</td>\n<td>Result</td>\n<td></td>\n</tr><tr>\n<td>*1R (Chung) (A/ P/ D/ Avg/ Tot)</td>\n<td>*2R (Chung) (A/ P/ D/ Avg/ Tot)</td>\n<td>**Total (Chung) (A/ P/ D/ Avg/ Tot)</td>\n<td>*1R (Hong) (A/ P/ D/ Avg/ Tot)</td>\n<td>*2R (Hong) (A/ P/ D/ Avg/ Tot)</td>\n<td>**Total (Hong) (A/ P/ D/ Avg/ Tot)</td>\n<td></td>\n<td></td>\n</tr><tr>\n<td>1R J1(Chung)</td>\n<td>1R J2(Chung)</td>\n<td>1R J3(Chung)</td>\n<td>1R J4(Chung)</td>\n<td>1R J5(Chung)</td>\n<td>1R J6(Chung)</td>\n<td>1R J7(Chung)</td>\n<td></td>\n</tr><tr>\n<td>2R J1(Chung)</td>\n<td>2R J2(Chung)</td>\n<td>2R J3(Chung)</td>\n<td>2R J4(Chung)</td>\n<td>2R J5(Chung)</td>\n<td>2R J6(Chung)</td>\n<td>2R J7(Chung)</td>\n<td></td>\n</tr><tr>\n<td>1R J1(Hong)</td>\n<td>1R J2(Hong)</td>\n<td>1R J3(Hong)</td>\n<td>1R J4(Hong)</td>\n<td>1R J5(Hong)</td>\n<td>1R J6(Hong)</td>\n<td>1R J7(Hong)</td>\n<td></td>\n</tr><tr>\n<td>2R J1(Hong)</td>\n<td>2R J2(Hong)</td>\n<td>2R J3(Hong)</td>\n<td>2R J4(Hong)</td>\n<td>2R J5(Hong)</td>\n<td>2R J6(Hong)</td>\n<td>2R J7(Hong)</td>\n<td></td>\n</tr><tr>\n<td colspan='8' style='height: 5px'></td>\n</tr>\n"
                if os.path.exists(path):
                    try:
                        with open(path, 'r', encoding='utf-8') as f:
                            content = f.read()
                        if "</table>" in content:
                            parts = content.rsplit("</table>", 1)
                            content = parts[0] + html_block + "</table>" + parts[1]
                        else:
                            content += html_block
                        with open(path, 'w', encoding='utf-8') as f:
                            f.write(content)
                    except Exception as e:
                        print(f"Error appending HTML to {path}: {e}")
                else:
                    try:
                        with open(path, 'w', encoding='utf-8') as f:
                            f.write(header_template + html_block + "</table></body></html>")
                    except Exception as e:
                        print(f"Error creating HTML log {path}: {e}")
                        
            append_to_file(filename)
            
            if self.watch_directory and os.path.isdir(self.watch_directory):
                dest_path = os.path.join(self.watch_directory, os.path.basename(filename))
                append_to_file(dest_path)
        except Exception as err:
            with open("error.log", "a", encoding="utf-8") as err_f:
                import traceback
                err_f.write(f"--- ERROR AT {datetime.now()} ---\n")
                traceback.print_exc(file=err_f)

    def end_match(self):
        # 真正將暫存的分數一次性存入 SQLite 資料庫 (改用批次寫入)
        try:
            if hasattr(self, 'temp_scores_to_save') and self.temp_scores_to_save:
                flat_scores = []
                for r_num, scores_list in self.temp_scores_to_save.items():
                    flat_scores.extend(scores_list)
                database.save_scores_batch(flat_scores)
                self.temp_scores_to_save = {}
        except Exception as db_err:
            import traceback
            print(f"Error saving temp scores: {db_err}")
            with open("error.log", "a", encoding="utf-8") as f:
                f.write(f"--- DB SAVE TEMP SCORES ERROR AT {datetime.now()} ---\n")
                traceback.print_exc(file=f)
            
        # 將寫日誌與匯出報表移至背景執行緒執行，避免阻塞 GUI
        def run_background_logs():
            try:
                self.write_html_log()
            except Exception as log_err:
                import traceback
                print(f"Error writing HTML log in background: {log_err}")
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- WRITE HTML LOG ERROR AT {datetime.now()} ---\n")
                    traceback.print_exc(file=f)
                
            try:
                self.export_match_log(auto_open=False)
            except Exception as match_log_err:
                import traceback
                print(f"Error exporting match log in background: {match_log_err}")
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- EXPORT MATCH LOG ERROR AT {datetime.now()} ---\n")
                    traceback.print_exc(file=f)

        if self.current_match_uuid and self.current_match_uuid in self.imported_matches:
            self.imported_matches[self.current_match_uuid]["Status"] = "End"
            # 完賽後立即清除該選手的排行榜分數快取，強制下次查詢排行榜時重新從資料庫取得正確分數
            self.invalidate_leaderboard_cache(uid=self.current_match_uuid)
            self.update_tree_columns()

        threading.Thread(target=run_background_logs, daemon=True).start()
            
        # 於背景非同步請求本機 Flask 端點，將 SocketIO 廣播工作帶回 Flask 執行緒上下文中，確保執行緒安全
        def trigger_reset():
            try:
                import urllib.request
                urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/reset_match", timeout=2, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
            except Exception as e:
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- API RESET TRIGGER ERROR: {e} ---\n")
        threading.Thread(target=trigger_reset, daemon=True).start()
        
        self.current_match_data = None
        self.current_match_uuid = None
        self.is_locked = False
        self.current_stage = 1
        config.current_state['current_player_side'] = 0
        config.current_state['pk_seq_state'] = 0  # 重置依序上場狀態機
        
        # 重置倒數計時器
        self.timer_running = False
        self.timer_seconds = int(config.system_settings.get("countdown_sec", 90))
        self._update_timer_label()
        self.update_court_label()
        self.lbl_no.config(text="籤號：", font=("Microsoft JhengHei", 28, "bold"))
        self.lbl_name_1.config(text="")
        self.lbl_name_2.config(text="")
        self.lbl_team_1.config(text="")
        self.lbl_team_2.config(text="")
        self.lbl_final_L.config(text="")
        self.lbl_final_R.config(text="")
        self.lbl_deduction_val.config(text="")
        self.lbl_deduction_val_R.config(text="")
        self.btn_ready.config(text="準備")
        self.btn_start.config(text="開始")
        self.btn_next.config(state="disabled")
        self.btn_end.config(state="disabled")
        self.btn_show_score.config(state="disabled", bg="#eee")
        for r in range(7):
            self.score_cells[r][0].config(text="")
            self.score_cells[r][1].config(text="")
            self.score_cells[r][2].config(text="")
        self.clear_left_panel_scores()
        self.reset_judges_display()
        if hasattr(self, 'combo_poomsae_1'):
            self.combo_poomsae_1.set("")
        if hasattr(self, 'combo_poomsae_2'):
            self.combo_poomsae_2.set("")
        self.update_proj_data("Waiting...", "", "", "")
        self.update_button_states()
    def unselect_match(self):
        if self.is_locked:
            messagebox.showwarning("警告", "比賽進行中 (已準備)，無法取消", parent=self.root)
            return
        self.tree.selection_remove(self.tree.selection())
        
        self.current_match_data = None
        self.current_match_uuid = None
        self.is_locked = False
        self.current_stage = 1
        
        # 於背景非同步請求本機 Flask 端點，將 SocketIO 廣播工作帶回 Flask 執行緒上下文中，確保執行緒安全
        def trigger_reset():
            try:
                import urllib.request
                urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/reset_match", timeout=2, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
            except Exception as e:
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- API RESET TRIGGER ERROR: {e} ---\n")
        threading.Thread(target=trigger_reset, daemon=True).start()
        
        self.update_court_label()
        self.lbl_no.config(text="籤號：", font=("Microsoft JhengHei", 28, "bold"))
        self.lbl_name_1.config(text="")
        self.lbl_name_2.config(text="")
        self.lbl_team_1.config(text="")
        self.lbl_team_2.config(text="")
        self.lbl_final_L.config(text="")
        self.lbl_final_R.config(text="")
        self.lbl_deduction_val.config(text="")
        self.lbl_deduction_val_R.config(text="")
        
        for r in range(7):
            self.score_cells[r][0].config(text="")
            self.score_cells[r][1].config(text="")
            self.score_cells[r][2].config(text="")
        self.clear_left_panel_scores()
        self.reset_judges_display()
        if hasattr(self, 'combo_poomsae_1'):
            self.combo_poomsae_1.set("")
        if hasattr(self, 'combo_poomsae_2'):
            self.combo_poomsae_2.set("")
        self.update_proj_data("Waiting...", "", "", "")
        self.update_button_states()
    def reshow_scores(self):
        if not self.current_match_data:
            return
        self.last_proj_score_slide = 0
        self.last_proj_slide_finished = False
        if self.proj_window and self.proj_window.winfo_exists():
            self.proj_window.current_score_slide = 0
            self.proj_window.score_slide_show_finished = False
            self.proj_window.stop_score_slide_show()
        self.show_scores()

    def re_evaluate_round(self):
        if not self.current_match_data:
            return
            
        # 將大螢幕切回 Scoring
        name = self.current_match_data["C_Name"]
        team = self.current_match_data["C_Team"]
        self.update_proj_data("Scoring...", name, team, "-")
        
        # 呼叫 API 讓裁判平板回到分數送出前的狀態
        def trigger_resume():
            try:
                import urllib.request
                urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/resume_scoring", timeout=2, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
            except Exception as e:
                print(f"Error resuming scoring: {e}")
                
        threading.Thread(target=trigger_resume, daemon=True).start()
        
        # 本地端按鈕狀態更新
        self.btn_show_score.config(state="disabled") 
        self.update_button_states()
        
        # 強制系統重估送分狀態 (可讓展示分數按鈕在所有人都 submitted 時再度亮起)
        try:
            import web_server
            web_server.check_all_submitted()
        except:
            pass

    def select_next_player(self):
        if self.is_locked:
            return
            
        if not hasattr(self, 'tree_ready'):
            return
            
        children = self.tree_ready.get_children()
        if not children:
            messagebox.showinfo("提示", "目前沒有未比賽的選手！", parent=self.root)
            return
            
        next_iid = None
        if self.current_match_uuid and self.current_match_uuid in children:
            curr_idx = children.index(self.current_match_uuid)
            if curr_idx + 1 < len(children):
                next_iid = children[curr_idx + 1]
            else:
                next_iid = children[0]
        else:
            next_iid = children[0]
            
        if next_iid:
            self.tree_ready.selection_remove(self.tree_ready.selection())
            self.tree_ready.selection_set(next_iid)
            self.tree_ready.focus(next_iid)
            self.tree_ready.see(next_iid)
            
            class DummyEvent:
                def __init__(self, widget):
                    self.widget = widget
            
            self.on_match_select(DummyEvent(self.tree_ready))

    def update_proj_data(self, status, player="", team="", score="-"):
        self.current_proj_status = status
        if self.proj_window and self.proj_window.winfo_exists():
            self.proj_window.update_data(status, player, team, score)
    def open_log(self):
        self.export_match_log(auto_open=True)

    def export_match_log(self, auto_open=False):
        # 優先從當前比賽資料取得 SourceFile (場次)，若無則從下拉選單取得
        if self.current_match_data and self.current_match_data.get("SourceFile"):
            target_session = self.current_match_data.get("SourceFile", "").strip()
        else:
            target_session = self.cb_session_select.get().strip()
            
        if not target_session:
            if auto_open:
                messagebox.showwarning("提示", "請先選擇比賽場次以檢視分數紀錄。", parent=self.root)
            return
            
        group_players = []
        html_title = target_session
        for uid, mdata in self.imported_matches.items():
            if mdata.get("SourceFile") == target_session:
                group_players.append((uid, mdata))
                    
        if not group_players:
            if auto_open:
                messagebox.showwarning("提示", "找不到相關的選手資料。", parent=self.root)
            return
            
        import sqlite3
        from datetime import datetime
        
        detailed_list = []
        
        # === 優化：一次性從 SQLite 批次撈取同組所有選手的分數，避免在迴圈中重複開關資料庫 ===
        uids_to_query = [uid for uid, _ in group_players]
        scores_by_uid = {}
        if uids_to_query:
            conn = sqlite3.connect(database.get_db_path())
            c = conn.cursor()
            try:
                placeholders = ",".join(["?"] * len(uids_to_query))
                query_sql = f"""
                    SELECT match_uuid, round, judge_id, accuracy, presentation, total, p1, p2, p3, deduction, timestamp, player_side
                    FROM scores
                    WHERE match_uuid IN ({placeholders})
                    ORDER BY round, judge_id
                """
                c.execute(query_sql, uids_to_query)
                all_rows = c.fetchall()
                for r in all_rows:
                    muid = r[0]
                    if muid not in scores_by_uid:
                        scores_by_uid[muid] = []
                    # 排除 match_uuid，組裝成與原本 row 格式一致：
                    # round(0), judge_id(1), accuracy(2), presentation(3), total(4), p1(5), p2(6), p3(7), deduction(8), timestamp(9), player_side(10)
                    scores_by_uid[muid].append(r[1:])
            except Exception as e_batch:
                print(f"[Export Batch Log Error] Fallback to legacy structure query: {e_batch}")
                try:
                    # 備用相容性結構查詢 (相容無 p1, p2, p3 等欄位的舊版 scores 表)
                    # 注意：補齊缺少的欄位為 None，確保 row[1:] 後欄位結構符合：
                    # round(0), judge_id(1), accuracy(2), presentation(3), total(4), p1(5), p2(6), p3(7), deduction(8), timestamp(9), player_side(10)
                    placeholders = ",".join(["?"] * len(uids_to_query))
                    query_sql_legacy = f"""
                        SELECT match_uuid, round, judge_id, accuracy, presentation, total,
                               0.0 AS p1, 0.0 AS p2, 0.0 AS p3,
                               0.0 AS deduction, NULL AS timestamp, player_side
                        FROM scores
                        WHERE match_uuid IN ({placeholders})
                        ORDER BY round, judge_id
                    """
                    c.execute(query_sql_legacy, uids_to_query)
                    all_rows_legacy = c.fetchall()
                    for r in all_rows_legacy:
                        muid = r[0]
                        if muid not in scores_by_uid:
                            scores_by_uid[muid] = []
                        scores_by_uid[muid].append(r[1:])
                except Exception as e_fallback:
                    print(f"[Export Batch Log Error] Fallback query failed: {e_fallback}")
            finally:
                conn.close()
        
        for uid, mdata in group_players:
            total_avg = 0.0
            is_pk = False
            # 優先嘗試從記憶體暫存 (temp_scores) 中取得最新的評分
            rows = []
            temp_scores = config.current_state.get('temp_scores', getattr(self, 'temp_scores_to_save', {}))
            if temp_scores:
                for r_num, scores_list in temp_scores.items():
                    for s in scores_list:
                        if s.get('match_uuid') == uid:
                            rows.append((
                                s['round_num'],
                                s['judge_id'],
                                s['acc'],
                                s['pres'],
                                s['total'],
                                s.get('p1', 0.0),
                                s.get('p2', 0.0),
                                s.get('p3', 0.0),
                                s.get('deduction', 0.0),
                                None,
                                s.get('player_side', 0)
                            ))
            if not rows:
                # 直接從一次性撈取的字典中拿取，無須再連接 SQLite
                rows = scores_by_uid.get(uid, [])
            
            status_val = mdata.get("Status", "")
            if status_val != "End" and status_val != "Withdraw":
                rows = []
            is_withdraw = (len(rows) == 0)
            if status_val == "End":
                result_status = "End"
            elif status_val == "Withdraw":
                result_status = "Withdraw"
            else:
                result_status = ""
            
            court = str(mdata.get("Court", ""))
            no = str(mdata.get("No", ""))
            
            game_val = mdata.get("Game", 0)
            try: game_method = "Cutoff" if int(game_val) == 0 else "Tournaments"
            except: game_method = "Cutoff"
                
            game_type = str(mdata.get("Type", ""))
            category = str(mdata.get("Category", ""))
            division = str(mdata.get("Division", ""))
            phase = str(mdata.get("Phase", ""))
            
            is_pk = False
            try: is_pk = (int(game_val) == 1)
            except: pass

            def get_player_side(r):
                if len(r) > 10: return r[10]
                elif len(r) > 5: return r[5]
                return 0

            has_chung_scores = any(get_player_side(r) == 0 for r in rows)
            has_hong_scores = any(get_player_side(r) == 1 for r in rows)
            
            c_noc = str(mdata.get("C_NOC", "")) if has_chung_scores else ""
            c_team = str(mdata.get("C_Team", ""))
            c_name = str(mdata.get("C_Name", ""))
            
            h_noc = str(mdata.get("H_NOC", "")) if (is_pk and has_hong_scores) else ""
            h_team = str(mdata.get("H_Team", ""))
            h_name = str(mdata.get("H_Name", ""))

                
            def get_deduction(side_rows, r_num):
                scores = [r for r in side_rows if r[0] == r_num]
                if len(scores) > 0 and len(scores[0]) > 8:
                    deds = [r[8] for r in scores if r[8] is not None]
                    if deds: return deds[0]
                return 0.0

            def calc_avg(scores_list):
                if not scores_list:
                    return 0.0, 0.0, 0.0
                accs = [s[2] for s in scores_list]
                pres = [s[3] for s in scores_list]
                
                def calc_trimmed_avg(val_list):
                    if not val_list: return 0.0
                    if len(val_list) <= 3: return sum(val_list) / len(val_list)
                    val_list.sort()
                    return sum(val_list[1:-1]) / len(val_list[1:-1])
                
                avg_acc = calc_trimmed_avg(accs)
                avg_pres = calc_trimmed_avg(pres)
                total_raw = sum(accs) + sum(pres)
                return avg_acc, avg_pres, total_raw

            end_time_str = ""
            if status_val == "End" and not is_withdraw:
                if rows and len(rows[0]) > 9 and rows[0][9]:
                    try:
                        ts_str = rows[0][9]
                        if isinstance(ts_str, str):
                            dt = datetime.strptime(ts_str.split(".")[0], "%Y-%m-%d %H:%M:%S")
                        else:
                            dt = ts_str
                        end_time_str = dt.strftime("%Y%m%d%H%M%S")
                    except:
                        end_time_str = datetime.now().strftime("%Y%m%d%H%M%S")
                else:
                    end_time_str = datetime.now().strftime("%Y%m%d%H%M%S")
                
            html_block = f"""<tr>
<td colspan='8' style='height: 5px'></td>
</tr><tr>
<td>{court}</td>
<td>{no}</td>
<td>{game_method}</td>
<td>{game_type}</td>
<td>{category}</td>
<td>{division}</td>
<td>{phase}</td>
<td>{end_time_str}</td>
</tr><tr>
<td>{c_noc}</td>
<td>{c_team}</td>
<td>{c_name}</td>
<td>{h_noc}</td>
<td>{h_team}</td>
<td>{h_name}</td>
<td>{result_status}</td>
<td>__RANK_OR_WINNER__</td>
</tr>"""

            if not is_withdraw:
                # 篩選青方分數
                chung_side_rows = [r for r in rows if get_player_side(r) == 0]
                r1_scores_chung = [r for r in chung_side_rows if r[0] == 1]
                r2_scores_chung = [r for r in chung_side_rows if r[0] == 2]
                
                r1_avg_acc, r1_avg_pres, r1_total_raw = calc_avg(r1_scores_chung)
                r2_avg_acc, r2_avg_pres, r2_total_raw = calc_avg(r2_scores_chung)
                
                ded_1r_chung = get_deduction(chung_side_rows, 1)
                ded_2r_chung = get_deduction(chung_side_rows, 2)
                
                r1_avg = r1_avg_acc + r1_avg_pres - ded_1r_chung if r1_scores_chung else 0.0
                r2_avg = r2_avg_acc + r2_avg_pres - ded_2r_chung if r2_scores_chung else 0.0
                
                if r1_scores_chung and r2_scores_chung:
                    total_avg = (r1_avg + r2_avg) / 2
                    total_raw = r1_total_raw + r2_total_raw
                    avg_acc_chung = (r1_avg_acc + r2_avg_acc) / 2
                    avg_pres_chung = (r1_avg_pres + r2_avg_pres) / 2
                elif r1_scores_chung:
                    total_avg = r1_avg
                    total_raw = r1_total_raw
                    avg_acc_chung = r1_avg_acc
                    avg_pres_chung = r1_avg_pres
                else:
                    total_avg = 0.0
                    total_raw = 0.0
                    avg_acc_chung = 0.0
                    avg_pres_chung = 0.0
                    
                r1_text = f"{r1_avg_acc:.3f} / {r1_avg_pres:.3f} / {ded_1r_chung:.1f} / {r1_avg:.3f} / {r1_total_raw:.1f}" if r1_scores_chung else ""
                r2_text = f"{r2_avg_acc:.3f} / {r2_avg_pres:.3f} / {ded_2r_chung:.1f} / {r2_avg:.3f} / {r2_total_raw:.1f}" if r2_scores_chung else ""
                total_text = f"{avg_acc_chung:.3f} / {avg_pres_chung:.3f} / {ded_1r_chung:.1f} / {total_avg:.3f} / {total_raw:.1f}" if r1_scores_chung else ""
                
                # 初始化紅方欄位
                r1_text_hong = ""
                r2_text_hong = ""
                total_text_hong = ""
                
                is_pk = False
                try: is_pk = (int(game_val) == 1)
                except: pass
                
                if is_pk:
                    hong_side_rows = [r for r in rows if get_player_side(r) == 1]
                    r1_scores_hong = [r for r in hong_side_rows if r[0] == 1]
                    r2_scores_hong = [r for r in hong_side_rows if r[0] == 2]
                    
                    r1_avg_acc_h, r1_avg_pres_h, r1_total_raw_h = calc_avg(r1_scores_hong)
                    r2_avg_acc_h, r2_avg_pres_h, r2_total_raw_h = calc_avg(r2_scores_hong)
                    
                    ded_1r_hong = get_deduction(hong_side_rows, 1)
                    ded_2r_hong = get_deduction(hong_side_rows, 2)
                    
                    r1_avg_h = r1_avg_acc_h + r1_avg_pres_h - ded_1r_hong if r1_scores_hong else 0.0
                    r2_avg_h = r2_avg_acc_h + r2_avg_pres_h - ded_2r_hong if r2_scores_hong else 0.0
                    
                    if r1_scores_hong and r2_scores_hong:
                        total_avg_h = (r1_avg_h + r2_avg_h) / 2
                        total_raw_h = r1_total_raw_h + r2_total_raw_h
                        avg_acc_hong = (r1_avg_acc_h + r2_avg_acc_h) / 2
                        avg_pres_hong = (r1_avg_pres_h + r2_avg_pres_h) / 2
                    elif r1_scores_hong:
                        total_avg_h = r1_avg_h
                        total_raw_h = r1_total_raw_h
                        avg_acc_hong = r1_avg_acc_h
                        avg_pres_hong = r1_avg_pres_h
                    else:
                        total_avg_h = 0.0
                        total_raw_h = 0.0
                        avg_acc_hong = 0.0
                        avg_pres_hong = 0.0
                        
                    r1_text_hong = f"{r1_avg_acc_h:.3f} / {r1_avg_pres_h:.3f} / {ded_1r_hong:.1f} / {r1_avg_h:.3f} / {r1_total_raw_h:.1f}" if r1_scores_hong else ""
                    r2_text_hong = f"{r2_avg_acc_h:.3f} / {r2_avg_pres_h:.3f} / {ded_2r_hong:.1f} / {r2_avg_h:.3f} / {r2_total_raw_h:.1f}" if r2_scores_hong else ""
                    total_text_hong = f"{avg_acc_hong:.3f} / {avg_pres_hong:.3f} / {ded_1r_hong:.1f} / {total_avg_h:.3f} / {total_raw_h:.1f}" if r1_scores_hong else ""
                
                # 輸出成績資料行 (PK時同時印出青紅雙方成績)
                html_block += f"""<tr>
<td>{r1_text}</td>
<td>{r2_text}</td>
<td style='color: blue; font-weight:bold;'>{total_text}</td>
<td>{r1_text_hong}</td>
<td>{r2_text_hong}</td>
<td style='color: red; font-weight:bold;'>{total_text_hong}</td>
<td></td>
<td></td>
</tr>"""
                
                def format_judge_row(r_scores):
                    cells = []
                    for j_idx in range(1, 8):
                        j_score = None
                        for s in r_scores:
                            jid = s[1]
                            if jid == f"J{j_idx}" or jid == f"manual_J{j_idx}":
                                j_score = s
                                break
                        if j_score:
                            acc_val, pres_val = j_score[2], j_score[3]
                            if len(j_score) > 8 and j_score[5] is not None and (j_score[5] > 0 or j_score[6] > 0 or j_score[7] > 0):
                                p1, p2, p3 = j_score[5], j_score[6], j_score[7]
                            else:
                                p1 = min(2.0, pres_val / 3)
                                p2 = min(2.0, (pres_val - p1) / 2)
                                p3 = max(0.0, pres_val - p1 - p2)
                            cells.append(f"<td>{int(acc_val*10)} / {int(p1*10)} / {int(p2*10)} / {int(p3*10)}</td>")
                        else:
                            cells.append("<td></td>")
                    return "<tr>" + "".join(cells) + "<td></td></tr>"
                    
                # 青方 1R/2R 裁判明細
                html_block += format_judge_row(r1_scores_chung)
                html_block += format_judge_row(r2_scores_chung)
                
                # 紅方 1R/2R 裁判明細 (若為非 PK 則以兩列空行填充以對齊表頭)
                if is_pk:
                    html_block += format_judge_row(r1_scores_hong)
                    html_block += format_judge_row(r2_scores_hong)
                else:
                    html_block += "<tr>" + "<td></td>"*7 + "<td></td></tr>"
                    html_block += "<tr>" + "<td></td>"*7 + "<td></td></tr>"
                
            pk_winner = ""
            if is_pk and not is_withdraw:
                def val_eq(a, b):
                    return round(a, 4) == round(b, 4)
                t_avg_chung = total_avg
                t_avg_hong = total_avg_h
                pres_chung = avg_pres_chung
                pres_hong = avg_pres_hong
                raw_chung = total_raw
                raw_hong = total_raw_h
                
                if t_avg_chung > t_avg_hong:
                    pk_winner = "BLUE"
                elif t_avg_hong > t_avg_chung:
                    pk_winner = "RED"
                else:
                    if val_eq(pres_chung, pres_hong):
                        if val_eq(raw_chung, raw_hong):
                            pk_winner = "DRAW"
                        elif raw_chung > raw_hong:
                            pk_winner = "BLUE"
                        else:
                            pk_winner = "RED"
                    elif pres_chung > pres_hong:
                        pk_winner = "BLUE"
                    else:
                        pk_winner = "RED"

            detailed_list.append({
                "score": total_avg,
                "presentation_score": mdata.get("presentation_score", 0.0),
                "raw_total_score": mdata.get("raw_total_score", 0.0),
                "no": no,
                "is_end": mdata.get("Status") == "End",
                "has_score": not is_withdraw,
                "is_pk": is_pk,
                "pk_winner": pk_winner,
                "html": html_block
            })
            
        # 根據是否為 PK 賽制決定排序邏輯
        has_pk = any(item["is_pk"] for item in detailed_list)
        if has_pk:
            def pk_sort_key(item):
                try:
                    no_val = int(item["no"])
                except:
                    no_val = 999999
                # (True, no_val) -> 讓未結束的 (not is_end = True) 排在已結束的 (not is_end = False) 後面
                return (not item["is_end"], no_val)
            detailed_list.sort(key=pk_sort_key)
        else:
            # 一般賽制排序：
            # 1. 有成績的選手（has_score=True）排在前面，依據 WT 決勝規則：最終得分→表現力→原始總分 (皆降序)
            # 2. 尚未出賽的選手（has_score=False）排在後面，依籤號升序
            def nonpk_sort_key(item):
                try:
                    no_val = int(item["no"])
                except:
                    no_val = 999999
                if item["has_score"]:
                    # 有成績：以負值讓 sort 以降序排列三個決勝欄位
                    return (0, -item["score"], -item["presentation_score"], -item["raw_total_score"], no_val)
                else:
                    # 無成績：排在最後，再按籤號升序
                    return (1, 0, 0, 0, no_val)
            detailed_list.sort(key=nonpk_sort_key)
        
        # 計算非 PK 賽制的排名並填入 __RANK_OR_WINNER__
        if not has_pk:
            def scores_eq(a, b):
                return round(a, 4) == round(b, 4)
            has_score_items = [item for item in detailed_list if item["has_score"]]
            for idx, item in enumerate(has_score_items):
                if idx > 0 and (
                    scores_eq(item["score"], has_score_items[idx - 1]["score"]) and
                    scores_eq(item["presentation_score"], has_score_items[idx - 1]["presentation_score"]) and
                    scores_eq(item["raw_total_score"], has_score_items[idx - 1]["raw_total_score"])
                ):
                    item["rank"] = has_score_items[idx - 1]["rank"]
                else:
                    item["rank"] = idx + 1

        for item in detailed_list:
            if item["is_pk"]:
                val = item.get("pk_winner", "")
            else:
                if item["has_score"]:
                    val = str(item.get("rank", ""))
                else:
                    val = ""
            item["html"] = item["html"].replace("__RANK_OR_WINNER__", val)

        all_html_blocks = "".join([item["html"] for item in detailed_list])
        
        log_time_str = datetime.now().strftime("%Y/%m/%d %H:%M:%S")
        header_template = f"<!DOCTYPE html><html><head><meta charset='UTF-8'><title>LOG</title><style>table {{table-layout: fixed; width: 100%; border-collapse: collapse;}} td {{height: 20px; text-align: center; word-wrap: break-word;}}</style></head><body>\n<h1 style='text-align: center; border-top: 2px solid #888; border-bottom: 2px solid #888; height: 50px; line-height: 50px;'> GAME RESULT </h1>\n<p style='text-align: right; '> TIME : {log_time_str}</p>\n<table style='width: 100%;border-spacing: 0px; font-size: 13px;' border='1'>\n<tr>\n<td style='width: 13%;'>Court</td>\n<td style='width: 13%;'>No.</td>\n<td style='width: 13%;'>Game method</td>\n<td style='width: 13%;'>Type</td>\n<td style='width: 13%;'>Category</td>\n<td style='width: 13%;'>Division</td>\n<td style='width: 13%;'>Phase</td>\n<td style='width: 9%;'>End Time</td>\n</tr><tr>\n<td>Noc (Chung)</td>\n<td>Team (Chung)</td>\n<td>Name (Chung)</td>\n<td>Noc (Hong)</td>\n<td>Team (Hong)</td>\n<td>Name (Hong)</td>\n<td>Result</td>\n<td></td>\n</tr><tr>\n<td>*1R (Chung) (A/ P/ D/ Avg/ Tot)</td>\n<td>*2R (Chung) (A/ P/ D/ Avg/ Tot)</td>\n<td>**Total (Chung) (A/ P/ D/ Avg/ Tot)</td>\n<td>*1R (Hong) (A/ P/ D/ Avg/ Tot)</td>\n<td>*2R (Hong) (A/ P/ D/ Avg/ Tot)</td>\n<td>**Total (Hong) (A/ P/ D/ Avg/ Tot)</td>\n<td></td>\n<td></td>\n</tr><tr>\n<td>1R J1(Chung)</td>\n<td>1R J2(Chung)</td>\n<td>1R J3(Chung)</td>\n<td>1R J4(Chung)</td>\n<td>1R J5(Chung)</td>\n<td>1R J6(Chung)</td>\n<td>1R J7(Chung)</td>\n<td></td>\n</tr><tr>\n<td>2R J1(Chung)</td>\n<td>2R J2(Chung)</td>\n<td>2R J3(Chung)</td>\n<td>2R J4(Chung)</td>\n<td>2R J5(Chung)</td>\n<td>2R J6(Chung)</td>\n<td>2R J7(Chung)</td>\n<td></td>\n</tr><tr>\n<td>1R J1(Hong)</td>\n<td>1R J2(Hong)</td>\n<td>1R J3(Hong)</td>\n<td>1R J4(Hong)</td>\n<td>1R J5(Hong)</td>\n<td>1R J6(Hong)</td>\n<td>1R J7(Hong)</td>\n<td></td>\n</tr><tr>\n<td>2R J1(Hong)</td>\n<td>2R J2(Hong)</td>\n<td>2R J3(Hong)</td>\n<td>2R J4(Hong)</td>\n<td>2R J5(Hong)</td>\n<td>2R J6(Hong)</td>\n<td>2R J7(Hong)</td>\n<td></td>\n</tr><tr>\n<td colspan='8' style='height: 5px'></td>\n</tr>\n"
        
        import os
        os.makedirs("match_logs", exist_ok=True)
        safe_title = "".join([c for c in html_title if c.isalnum() or c in (' ', '-', '_')]).strip()
        match_log_path = os.path.join("match_logs", f"log_{safe_title}.html")
        
        with open(match_log_path, "w", encoding="utf-8") as f:
            f.write(header_template + all_html_blocks + "</table></body></html>")
        if auto_open:
            webbrowser.open(match_log_path)
    def toggle_projection(self):
        if self.proj_window is None or not self.proj_window.winfo_exists():
            target_x = 0; target_y = 0; target_w = 1920; target_h = 1080
            if get_monitors:
                try:
                    monitors = get_monitors()
                    target_m = None
                    for m in monitors:
                        if not m.is_primary: target_m = m; break
                    if not target_m and len(monitors) > 0: target_m = monitors[0]
                    if target_m: target_x = target_m.x; target_y = target_m.y; target_w = target_m.width; target_h = target_m.height
                except Exception as e: print("Screeninfo detection failed:", e)
            else: target_x = self.root.winfo_screenwidth()
            
            if hasattr(self, 'mode_var') and self.mode_var.get() == 1:
                from projection_pk import PKProjectionWindow
                self.proj_window = PKProjectionWindow(self.root, x=target_x, y=target_y, width=target_w, height=target_h)
            else:
                self.proj_window = ProjectionWindow(self.root, x=target_x, y=target_y, width=target_w, height=target_h)
                
            self.proj_window.main_gui = self
            self.btn_proj.normal_bg = self.colors["btn_orange"]
            self.btn_proj.config(text="關閉投影", bg=self.colors["btn_orange"])
            if self.current_match_data:
                saved_status = getattr(self, 'current_proj_status', "Ready")
                saved_slide = getattr(self, 'last_proj_score_slide', 0)
                saved_finished = getattr(self, 'last_proj_slide_finished', False)
                
                self.proj_window.status_text = saved_status
                self.proj_window.current_score_slide = saved_slide
                self.proj_window.score_slide_show_finished = saved_finished
                if saved_status == "Final Score":
                    self.proj_window.last_is_showing_score = True
                else:
                    self.proj_window.last_is_showing_score = False
                    
                score = ""
                if hasattr(self, 'left_merged_labels') and 4 in self.left_merged_labels:
                    score = self.left_merged_labels[4].cget("text")
                if not score or score == "-":
                    score = self.lbl_final_L.cget("text")
                if not score or score == "-":
                    score = self.lbl_final_R.cget("text")
                    
                self.proj_window.update_data(saved_status, self.current_match_data["C_Name"], self.current_match_data["C_Team"], score)
            else:
                self.proj_window.update_data("Waiting...", "", "", "")
            try: self.root.state('zoomed')
            except: self.root.attributes('-fullscreen', True)
            self.root.focus_force()
        else:
            self.proj_window.destroy()
            self.proj_window = None
            self.btn_proj.normal_bg = self.colors["btn_yellow"]
            self.btn_proj.config(text="投影畫面", bg=self.colors["btn_yellow"])
            
    def switch_projection_mode(self):
        """當賽制切換時，無縫重建大螢幕投影視窗以防畫面閃爍並露出桌面"""
        if hasattr(self, 'proj_window') and self.proj_window is not None and self.proj_window.winfo_exists():
            try:
                # 1. 暫存舊視窗並取得螢幕幾何資訊
                old_proj = self.proj_window
                target_x = old_proj.winfo_x()
                target_y = old_proj.winfo_y()
                target_w = old_proj.winfo_width()
                target_h = old_proj.winfo_height()

                # 2. 建立對應新賽制的新投影視窗
                if hasattr(self, 'mode_var') and self.mode_var.get() == 1:
                    from projection_pk import PKProjectionWindow
                    new_proj = PKProjectionWindow(self.root, x=target_x, y=target_y, width=target_w, height=target_h)
                else:
                    from projection import ProjectionWindow
                    new_proj = ProjectionWindow(self.root, x=target_x, y=target_y, width=target_w, height=target_h)

                new_proj.main_gui = self

                # 3. 同步資料與投影片播放狀態
                saved_status = getattr(self, 'current_proj_status', "Ready")
                saved_slide = getattr(self, 'last_proj_score_slide', 0)
                saved_finished = getattr(self, 'last_proj_slide_finished', False)

                new_proj.status_text = saved_status
                new_proj.current_score_slide = saved_slide
                new_proj.score_slide_show_finished = saved_finished
                if saved_status == "Final Score":
                    new_proj.last_is_showing_score = True
                else:
                    new_proj.last_is_showing_score = False

                score = ""
                if hasattr(self, 'left_merged_labels') and 4 in self.left_merged_labels:
                    score = self.left_merged_labels[4].cget("text")
                if not score or score == "-":
                    score = self.lbl_final_L.cget("text")
                if not score or score == "-":
                    score = self.lbl_final_R.cget("text")

                if self.current_match_data:
                    new_proj.update_data(saved_status, self.current_match_data["C_Name"], self.current_match_data["C_Team"], score)
                else:
                    new_proj.update_data("Waiting...", "", "", "")

                # 4. 強制繪製新視窗並拉抬至最上層以覆蓋舊視窗
                new_proj.update()
                new_proj.lift()

                # 5. 更新 GUI 變數指向新視窗，並銷毀舊視窗
                self.proj_window = new_proj
                old_proj.destroy()

                # 6. 確保視窗保持全螢幕
                try: self.root.state('zoomed')
                except: self.root.attributes('-fullscreen', True)
                self.root.focus_force()

            except Exception as e:
                # 若無縫切換失敗，回退至普通重建方式
                print(f"[Projection] 無縫切換失敗，執行回退重建: {e}")
                self.toggle_projection()  # 關閉舊投影
                self.toggle_projection()  # 重新開啟新投影
                
    def sync_timer_to_flask(self):
        def trigger_sync():
            try:
                import urllib.request
                urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/timer_sync?seconds={self.timer_seconds}&running={1 if self.timer_running else 0}", timeout=1, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
            except Exception as e:
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- API TIMER SYNC ERROR: {e} ---\n")
        threading.Thread(target=trigger_sync, daemon=True).start()
    def stop_timer(self):
        self.timer_running = False
        if hasattr(self, 'timer_after_id') and self.timer_after_id:
            self.root.after_cancel(self.timer_after_id)
            self.timer_after_id = None
        self.sync_timer_to_flask()
    def _timer_tick(self):
        if self.timer_running and self.timer_seconds > 0:
            self.timer_seconds -= 1
            self._update_timer_label()
            if self.proj_window and self.proj_window.winfo_exists():
                self.proj_window.refresh()
            self.timer_after_id = self.root.after(1000, self._timer_tick)
        elif self.timer_seconds == 0:
            self.timer_running = False
            self.timer_after_id = None
            self.btn_start.config(text="開始")
            if self.proj_window and self.proj_window.winfo_exists():
                self.proj_window.refresh()
            self.update_button_states()
            self.sync_timer_to_flask()
    def _update_timer_label(self):
        mins, secs = divmod(self.timer_seconds, 60)
        self.lbl_timer.config(text=f"{mins:02}:{secs:02}")
        if self.timer_seconds <= 10: self.lbl_timer.config(fg="red")
        else: self.lbl_timer.config(fg=self.colors["timer_fg"])
        self.sync_timer_to_flask()
    def get_local_ip(self):
        try:
            s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
            s.connect(("8.8.8.8", 80))
            ip = s.getsockname()[0]
            s.close()
            return ip
        except: return "127.0.0.1"
    def start_scoring(self):
        current_state['is_scoring'] = True
        
        # 1. 取得選手與型場資訊
        player_name = current_state.get('current_player', "TEST PLAYER")
        mode = self.mode_var.get() if hasattr(self, 'mode_var') else 0
        stage = self.current_stage if hasattr(self, 'current_stage') else 1
        
        team = ""
        no = ""
        category = ""
        division = ""
        phase = ""
        poomsae_name = ""
        poomsae_1 = ""
        poomsae_2 = ""
        chung_player = ""
        chung_team = ""
        hong_player = ""
        hong_team = ""
        match_type = ""
        
        if hasattr(self, 'current_match_data') and self.current_match_data:
            mdata = self.current_match_data
            team = mdata.get("C_Team", "")
            no = mdata.get("No", "")
            category = mdata.get("Category", "")
            division = mdata.get("Division", "")
            phase = mdata.get("Phase", "")
            match_type = mdata.get("Type", "")
            
            # 青方為 C_Name / C_Team
            chung_player = mdata.get("C_Name", "")
            chung_team = mdata.get("C_Team", "")
            # 紅方為 H_Name / H_Team（PK 模式才有紅方）
            hong_player = mdata.get("H_Name", "")
            hong_team = mdata.get("H_Team", "")
            
            poomsae_1 = self.combo_poomsae_1.get() if hasattr(self, 'combo_poomsae_1') else ""
            poomsae_2 = self.combo_poomsae_2.get() if hasattr(self, 'combo_poomsae_2') else ""
            
            if stage == 1:
                poomsae_name = poomsae_1
            else:
                poomsae_name = poomsae_2
        
        # 2. 組成 URL 參數
        import urllib.parse
        player_side = config.current_state.get('current_player_side', 0)
        pk_sequence_mode = config.system_settings.get('pk_sequence_mode', 0)
        params = urllib.parse.urlencode({
            'player': player_name,
            'team': team,
            'no': no,
            'mode': mode,
            'stage': stage,
            'category': category,
            'division': division,
            'phase': phase,
            'poomsae': poomsae_name,
            'poomsae_1': poomsae_1,
            'poomsae_2': poomsae_2,
            'match_type': match_type,
            'chung_player': chung_player,
            'chung_team': chung_team,
            'hong_player': hong_player,
            'hong_team': hong_team,
            'player_side': player_side,
            'pk_sequence_mode': pk_sequence_mode
        })
        
        # 3. 於背景非同步請求本機 Flask 端點，將 SocketIO 廣播派發工作帶回 Flask 執行緒上下文中
        def trigger():
            try:
                import urllib.request
                urllib.request.urlopen(f"{config.INTERNAL_SCHEME}://127.0.0.1:{config.PORT}/api/start_scoring?{params}", timeout=2, context=config.INTERNAL_SSL_CTX if config.USE_SSL else None)
            except Exception as e:
                with open("error.log", "a", encoding="utf-8") as f:
                    f.write(f"--- API TRIGGER ERROR: {e} ---\n")
        threading.Thread(target=trigger, daemon=True).start()
    def update_final_score(self, score):
        if "Final Points" in self.result_vars:
            self.root.after(0, lambda: self.lbl_final_L.config(text=score))

    def get_tiebreaker_metrics(self, uid, mdata, player_side=None):
        """計算選手的同分打破指標：(最終得分, 表現力去尾平均, 所有裁判原始分數的加總總分)"""
        if player_side is None:
            player_side = 0
            
        import sqlite3
        try:
            conn = sqlite3.connect(database.get_db_path())
            c = conn.cursor()
            try:
                c.execute("""
                    SELECT round, judge_id, accuracy, presentation, deduction, total
                    FROM scores
                    WHERE match_uuid = ? AND player_side = ?
                """, (uid, player_side))
                rows = c.fetchall()
            except Exception:
                # 舊資料庫相容：如果無 deduction 欄位則不讀取，且 total 欄位用 (accuracy + presentation) 計算
                c.execute("""
                    SELECT round, judge_id, accuracy, presentation, 0.0, (accuracy + presentation)
                    FROM scores
                    WHERE match_uuid = ? AND player_side = ?
                """, (uid, player_side))
                rows = c.fetchall()
            conn.close()
            
            temp_scores = config.current_state.get('temp_scores', getattr(self, 'temp_scores_to_save', {}))
            if temp_scores:
                existing_keys = {(row[0], str(row[1])) for row in rows}
                for r_num, scores_list in temp_scores.items():
                    for s in scores_list:
                        if s.get('match_uuid') == uid and s.get('player_side', 0) == player_side:
                            key = (s.get('round_num'), str(s.get('judge_id')))
                            if key not in existing_keys:
                                rows.append((
                                    s.get('round_num'),
                                    s.get('judge_id'),
                                    s.get('acc', 0.0),
                                    s.get('pres', 0.0),
                                    s.get('deduction', 0.0),
                                    s.get('total', 0.0)
                                ))
                                existing_keys.add(key)
            
            if not rows:
                return (0.0, 0.0, 0.0)
                
            r_scores = {}
            raw_sum_total = 0.0
            for row in rows:
                r_num = row[0]
                acc = row[2]
                pres = row[3]
                ded = row[4] if len(row) > 4 else 0.0
                total = row[5] if len(row) > 5 else (acc + pres)
                
                raw_sum_total += total
                
                if r_num not in r_scores:
                    r_scores[r_num] = {"acc": [], "pres": [], "ded": []}
                r_scores[r_num]["acc"].append(acc)
                r_scores[r_num]["pres"].append(pres)
                r_scores[r_num]["ded"].append(ded)
                
            def calc_avg(scores):
                if not scores: return 0.0
                if len(scores) <= 3: return sum(scores) / len(scores)
                else:
                    scores_sorted = sorted(scores)
                    valid = scores_sorted[1:-1]
                    return sum(valid) / len(valid)
                    
            r_averages = []
            r_pres_averages = []
            for r_num, sdata in r_scores.items():
                avg_acc = calc_avg(sdata["acc"])
                avg_pres = calc_avg(sdata["pres"])
                deduction = sdata["ded"][0] if sdata["ded"] else 0.0
                
                r_averages.append(avg_acc + avg_pres - deduction)
                r_pres_averages.append(avg_pres)
                
            if not r_averages:
                final_score = 0.0
                pres_avg = 0.0
            elif len(r_averages) == 1:
                final_score = r_averages[0]
                pres_avg = r_pres_averages[0]
            else:
                final_score = sum(r_averages) / len(r_averages)
                pres_avg = sum(r_pres_averages) / len(r_pres_averages)
                
            final_score = round(final_score, 3)
            pres_avg = round(pres_avg, 3)
            raw_sum_total = round(raw_sum_total, 2)
            
            return (final_score, pres_avg, raw_sum_total)
        except Exception as e:
            from datetime import datetime
            import traceback
            with open("error.log", "a", encoding="utf-8") as err_f:
                err_f.write(f"--- GET_TIEBREAKER_METRICS ERROR AT {datetime.now()} ---\n")
                traceback.print_exc(file=err_f)
            return (0.0, 0.0, 0.0)

    def get_final_score(self, uid, mdata, player_side=None):
        """獲取已完賽選手的最終得分，支援快取與 SQLite 歷史查詢 (包含扣分 deduction 計算)"""
        if player_side is None:
            player_side = 0
            
        cache_key = f"final_score_{player_side}"
        pres_key = f"presentation_score_{player_side}"
        raw_key = f"raw_total_score_{player_side}"
        
        # 舊代碼快取相容
        if player_side == 0 and "final_score" in mdata:
            # 確保決勝欄位也已填入，避免剛結束比賽的選手因早期返回而導致排名用 0.0 計算
            if "presentation_score" not in mdata or "raw_total_score" not in mdata:
                _, pres_avg, raw_sum = self.get_tiebreaker_metrics(uid, mdata, player_side)
                mdata["presentation_score"] = pres_avg
                mdata["raw_total_score"] = raw_sum
            return mdata["final_score"]
            
        if cache_key in mdata:
            return mdata[cache_key]
            
        final_score, pres_avg, raw_sum = self.get_tiebreaker_metrics(uid, mdata, player_side)
        
        # 寫入多方位快取
        mdata[cache_key] = final_score
        mdata[pres_key] = pres_avg
        mdata[raw_key] = raw_sum
        
        # 同時寫入舊欄位供舊相容代碼使用 (預設為 0)
        if player_side == 0:
            mdata["final_score"] = final_score
            mdata["presentation_score"] = pres_avg
            mdata["raw_total_score"] = raw_sum
            
        return final_score

    def update_button_states(self):
        # 1. 取得基本狀態
        has_player = (self.current_match_uuid is not None)
        p1 = self.combo_poomsae_1.get() if hasattr(self, 'combo_poomsae_1') else ""
        p2 = self.combo_poomsae_2.get() if hasattr(self, 'combo_poomsae_2') else ""
        is_2r_active = hasattr(self, 'combo_poomsae_2') and str(self.combo_poomsae_2['state']) != 'disabled'
        
        poomsae_ok = False
        if has_player:
            if is_2r_active:
                poomsae_ok = (p1 != "" and p2 != "")
            else:
                poomsae_ok = (p1 != "")
                
        is_ready = self.is_locked
        timer_running = self.timer_running
        timer_seconds = self.timer_seconds
        
        # 判斷分數是否已經送齊
        judge_count = int(system_settings["judge_count"])
        
        unique_judges = {}
        for sid, jdata in current_state['judges'].items():
            jid = jdata.get('id', '')
            if not jid: continue
            if jid not in unique_judges or sid.startswith("manual_"):
                unique_judges[jid] = jdata
        active_judges = list(unique_judges.values())
        submitted_count = sum(1 for j in active_judges if j.get('submitted'))
        scores_all_submitted = (submitted_count >= judge_count)
        
        # 判斷是否已經展示分數 (以當前設定的投影狀態為準)
        score_shown = (getattr(self, 'current_proj_status', '') == "Final Score")
                
        # 預設按鈕配置：灰色且不可點選，文字為中灰色
        btn_states = {
            "draw": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d"},
            "ready": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d", "text": "準備"},
            "start": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d", "text": "開始"},
            "show_score": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d"},
            "re_eval": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d"},
            "next": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d", "text": "下一品"},
            "end": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d"},
            "next_player": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d"},
            "reshow": {"state": "disabled", "bg": "#e0e0e0", "fg": "#7f8c8d"}
        }
        
        ready_has_children = False
        if hasattr(self, 'tree_ready'):
            ready_has_children = len(self.tree_ready.get_children()) > 0
        if not is_ready and ready_has_children:
            btn_states["next_player"] = {"state": "normal", "bg": "#1abc9c", "fg": "#ffffff"}
        
        if not has_player:
            # 階段 0：未載入任何選手
            pass
        elif not poomsae_ok:
            # 階段 1：已載入選手，尚未抽型場 -> 胡蘿蔔橘引導
            btn_states["draw"] = {"state": "normal", "bg": "#e67e22", "fg": "#ffffff"}
        elif not is_ready:
            # 階段 2：已抽型場，尚未準備 -> 抽型場變淡灰，準備變亮藍色引導
            btn_states["draw"] = {"state": "normal", "bg": "#bdc3c7", "fg": "#ffffff"}
            btn_states["ready"] = {"state": "normal", "bg": "#3498db", "fg": "#ffffff", "text": "準備"}
        else:
            # 已準備
            btn_states["ready"] = {"state": "normal", "bg": "#e74c3c", "fg": "#ffffff", "text": "取消"}
            
            # --- START 按鈕邏輯獨立判斷 ---
            if not score_shown:
                if timer_running:
                    # 階段 4：評分進行中 -> 橘色暫停
                    btn_states["start"] = {"state": "normal", "bg": "#e67e22", "fg": "#ffffff", "text": "暫停"}
                elif timer_seconds > 0:
                    # 階段 3 或者是 5：計分未開始 / 暫停中，等待開始 -> 綠色開始引導
                    btn_states["start"] = {"state": "normal", "bg": "#2ecc71", "fg": "#ffffff", "text": "開始"}
                # timer_seconds == 0 則維持預設 disabled (等待送分中)

            if score_shown:
                # 階段 8：分數已展示，準備進入下一輪或結束
                # 「本輪重評」按鈕在分數展示後立刻可以使用，無須等待投影片輪播完畢
                btn_states["re_eval"] = {"state": "normal", "bg": "#e67e22", "fg": "#ffffff"} # 橘色重評
                btn_states["reshow"] = {"state": "normal", "bg": "#34495e", "fg": "#ffffff"}
                
                # 只有在大螢幕輪播播放完指定秒數 (或沒有開啟大螢幕) 的情況下，下一品或結束按鈕才變為可以按的狀態
                proj_active = (self.proj_window is not None and self.proj_window.winfo_exists())
                slide_finished = self.last_proj_slide_finished if proj_active else True
                
                if slide_finished:
                    try: rounds = int(self.current_match_data["Round"])
                    except: rounds = 2
                    if self.current_stage == 1 and rounds == 2:
                        btn_states["next"] = {"state": "normal", "bg": "#f1c40f", "fg": "#2c3e50"} # 黃色下一品
                    else:
                        btn_states["end"] = {"state": "normal", "bg": "#c0392b", "fg": "#ffffff"} # 紅色結束
            elif scores_all_submitted:
                # 階段 7：分數已送齊
                mode = int(self.mode_var.get()) if hasattr(self, 'mode_var') else 0
                pk_seq = int(config.system_settings.get('pk_sequence_mode', 0))
                current_side = config.current_state.get('current_player_side', 0)
                seq_state = config.current_state.get('pk_seq_state', 0)
                
                if mode == 1 and pk_seq == 2:
                    # ── 依序上場專用邏輯 ──
                    if seq_state == 0:
                        # 青方1R完成：有2R則「下一品」，無2R則直接「交換選手」
                        try: rounds = int(self.current_match_data["Round"])
                        except: rounds = 2
                        if rounds >= 2:
                            btn_states["next"] = {"state": "normal", "bg": "#f1c40f", "fg": "#2c3e50", "text": "下一品"}
                        else:
                            # 1R only：直接換到紅方
                            btn_states["next"] = {"state": "normal", "bg": "#e67e22", "fg": "#ffffff", "text": "交換選手"}
                    elif seq_state == 1:
                        # 青方2R完成 → 「交換選手」換到紅方
                        btn_states["next"] = {"state": "normal", "bg": "#e67e22", "fg": "#ffffff", "text": "交換選手"}
                    else:
                        # 紅方1R或2R完成 → 展示分數
                        btn_states["show_score"] = {"state": "normal", "bg": "#9b59b6", "fg": "#ffffff"}
                elif mode == 1 and pk_seq == 1 and current_side == 0:
                    # 交叉上場且青方剛完成：「交換選手」
                    btn_states["next"] = {"state": "normal", "bg": "#e67e22", "fg": "#ffffff", "text": "交換選手"}
                else:
                    # 一般賽制 / PK 同時上場 / PK 交叉紅方已完成：紫色展示分數
                    btn_states["show_score"] = {"state": "normal", "bg": "#9b59b6", "fg": "#ffffff"}
                
        # 同步鎖定或解鎖場次選擇下拉選單
        if hasattr(self, 'cb_session_select'):
            if is_ready:
                self.cb_session_select.config(state="disabled")
            else:
                self.cb_session_select.config(state="normal")
                
        # 3. 更新按鈕狀態與背景色與前景文字色
        if hasattr(self, 'btn_draw'):
            self.btn_draw.normal_bg = btn_states["draw"]["bg"]
            self.btn_draw.config(state=btn_states["draw"]["state"], bg=btn_states["draw"]["bg"], fg=btn_states["draw"]["fg"])
        if hasattr(self, 'btn_ready'):
            self.btn_ready.normal_bg = btn_states["ready"]["bg"]
            self.btn_ready.config(state=btn_states["ready"]["state"], bg=btn_states["ready"]["bg"], fg=btn_states["ready"]["fg"], text=btn_states["ready"]["text"])
        if hasattr(self, 'btn_start'):
            self.btn_start.normal_bg = btn_states["start"]["bg"]
            self.btn_start.config(state=btn_states["start"]["state"], bg=btn_states["start"]["bg"], fg=btn_states["start"]["fg"], text=btn_states["start"]["text"])
        if hasattr(self, 'btn_show_score'):
            self.btn_show_score.normal_bg = btn_states["show_score"]["bg"]
            self.btn_show_score.config(state=btn_states["show_score"]["state"], bg=btn_states["show_score"]["bg"], fg=btn_states["show_score"]["fg"])
        if hasattr(self, 'btn_next'):
            self.btn_next.normal_bg = btn_states["next"]["bg"]
            self.btn_next.config(
                state=btn_states["next"]["state"],
                bg=btn_states["next"]["bg"],
                fg=btn_states["next"]["fg"],
                text=btn_states["next"].get("text", "下一品")
            )
        if hasattr(self, 'btn_end'):
            self.btn_end.normal_bg = btn_states["end"]["bg"]
            self.btn_end.config(state=btn_states["end"]["state"], bg=btn_states["end"]["bg"], fg=btn_states["end"]["fg"])
        if hasattr(self, 'btn_re_eval'):
            self.btn_re_eval.normal_bg = btn_states["re_eval"]["bg"]
            self.btn_re_eval.config(state=btn_states["re_eval"]["state"], bg=btn_states["re_eval"]["bg"], fg=btn_states["re_eval"]["fg"])
        if hasattr(self, 'btn_next_player'):
            self.btn_next_player.normal_bg = btn_states["next_player"]["bg"]
            self.btn_next_player.config(state=btn_states["next_player"]["state"], bg=btn_states["next_player"]["bg"], fg=btn_states["next_player"]["fg"])
        if hasattr(self, 'btn_reshow'):
            self.btn_reshow.normal_bg = btn_states["reshow"]["bg"]
            self.btn_reshow.config(state=btn_states["reshow"]["state"], bg=btn_states["reshow"]["bg"], fg=btn_states["reshow"]["fg"])

# === SSL 設定：自動偵測 IP 並在必要時重新生成憑證 ===
import ssl as _ssl_module
import ipaddress as _ipaddress
import datetime as _dt

_CERT_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'cert.pem')
_KEY_FILE  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'key.pem')

def _get_local_ip():
    """取得目前本機在區網中的 IP 位址"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        ip = s.getsockname()[0]
        s.close()
        return ip
    except Exception:
        return "127.0.0.1"

def _get_cert_ip(cert_path):
    """讀取 cert.pem 中的第一個 IP SAN，若無則回傳 None"""
    try:
        from cryptography import x509 as _cx509
        with open(cert_path, "rb") as f:
            cert_data = f.read()
        cert_obj = _cx509.load_pem_x509_certificate(cert_data)
        san_ext = cert_obj.extensions.get_extension_for_class(_cx509.SubjectAlternativeName)
        ip_list = san_ext.value.get_values_for_type(_cx509.IPAddress)
        # 回傳第一個非 127.0.0.1 的 IP，否則回傳第一個 IP
        for ip in ip_list:
            if str(ip) != "127.0.0.1":
                return str(ip)
        return str(ip_list[0]) if ip_list else None
    except Exception:
        return None

def _generate_ssl_cert(local_ip):
    """使用 cryptography 套件生成自簽憑證，並加入正確的 IP SAN"""
    from cryptography import x509 as _cx509
    from cryptography.x509.oid import NameOID as _NameOID, ExtendedKeyUsageOID as _EKU
    from cryptography.hazmat.primitives import hashes as _hashes, serialization as _ser
    from cryptography.hazmat.primitives.asymmetric import rsa as _rsa

    print(f"[SSL] Generating new certificate for IP: {local_ip}")
    private_key = _rsa.generate_private_key(public_exponent=65537, key_size=2048)

    subject = issuer = _cx509.Name([
        _cx509.NameAttribute(_NameOID.COUNTRY_NAME, "TW"),
        _cx509.NameAttribute(_NameOID.ORGANIZATION_NAME, "Poomsae Scoring System"),
        _cx509.NameAttribute(_NameOID.COMMON_NAME, local_ip),
    ])

    san_list = [
        _cx509.IPAddress(_ipaddress.ip_address(local_ip)),
        _cx509.IPAddress(_ipaddress.ip_address("127.0.0.1")),
        _cx509.DNSName("localhost"),
    ]

    cert = (
        _cx509.CertificateBuilder()
        .subject_name(subject)
        .issuer_name(issuer)
        .public_key(private_key.public_key())
        .serial_number(_cx509.random_serial_number())
        .not_valid_before(_dt.datetime.now(_dt.timezone.utc))
        .not_valid_after(_dt.datetime.now(_dt.timezone.utc) + _dt.timedelta(days=3 * 365))
        .add_extension(_cx509.SubjectAlternativeName(san_list), critical=False)
        .add_extension(_cx509.BasicConstraints(ca=True, path_length=None), critical=True)
        .add_extension(
            _cx509.KeyUsage(
                digital_signature=True, key_cert_sign=True, crl_sign=True,
                content_commitment=False, key_encipherment=False, data_encipherment=False,
                key_agreement=False, encipher_only=False, decipher_only=False,
            ), critical=True,
        )
        .add_extension(_cx509.ExtendedKeyUsage([_EKU.SERVER_AUTH]), critical=False)
        .sign(private_key, _hashes.SHA256())
    )

    with open(_CERT_FILE, "wb") as f:
        f.write(cert.public_bytes(_ser.Encoding.PEM))
    with open(_KEY_FILE, "wb") as f:
        f.write(private_key.private_bytes(
            encoding=_ser.Encoding.PEM,
            format=_ser.PrivateFormat.TraditionalOpenSSL,
            encryption_algorithm=_ser.NoEncryption(),
        ))
    print(f"[SSL] Certificate saved to {_CERT_FILE}")

def _ensure_ssl_cert():
    local_ip = _get_local_ip()
    print(f"[HTTP] 本地 HTTPS 已停用，專為 Cloudflare Tunnel 優化。")
    return None, None, local_ip

# 啟動時執行自動憑證檢查
_cert_result = _ensure_ssl_cert()
_LOCAL_IP = _cert_result[2] if len(_cert_result) == 3 else _get_local_ip()
USE_SSL = (_cert_result[0] is not None)
INTERNAL_SCHEME = 'https' if USE_SSL else 'http'

# SSL context 供內部 urllib 呼叫使用（略過自簽憑證驗證）
_INTERNAL_SSL_CTX = _ssl_module.SSLContext(_ssl_module.PROTOCOL_TLS_CLIENT)
_INTERNAL_SSL_CTX.check_hostname = False
_INTERNAL_SSL_CTX.verify_mode = _ssl_module.CERT_NONE

